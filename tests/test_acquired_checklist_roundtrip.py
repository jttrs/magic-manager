"""Round-trip tests for the ADD-mode checklist column/bracket contract
(2026-09-03): the generator writes a single ``acquired_qty`` column (XLSX) /
``[A:n]`` bracket (MD), and the parser reads it back into
``JumpstartRow.acquired_qty``. Pins the writer↔parser wiring so a column rename
can't silently break ingest. Precon MODIFY keeps its three-column layout.
"""

from __future__ import annotations

import openpyxl


def _fake_jumpstart(monkeypatch):
    """Stub the network seams the jumpstart writer touches: variant list, per-deck
    JSON (for rollups), and the front-card sync."""
    from magic_manager import mtgjson, front_cards
    variants = [
        {"fileName": "IronMan_MSH", "name": "Iron Man", "type": "Jumpstart", "code": "MSH"},
        {"fileName": "Wild_MSH", "name": "Wild", "type": "Jumpstart", "code": "MSH"},
    ]
    monkeypatch.setattr(mtgjson, "jumpstart_variants", lambda code: list(variants))
    monkeypatch.setattr(mtgjson, "deck", lambda fn: {
        "name": next(v["name"] for v in variants if v["fileName"] == fn),
        "type": "Jumpstart",
        "mainBoard": [{"name": "Card", "count": 1, "isFoil": False, "setCode": "MSH",
                       "number": "1", "identifiers": {"scryfallId": "x"}}],
    })
    monkeypatch.setattr(front_cards, "sync_front_cards", lambda anchor: 0)


def test_jumpstart_xlsx_has_single_acquired_column(tmp_db, tmp_path, monkeypatch):
    """The jumpstart XLSX writer emits ONE acquired_qty column (no keep_qty /
    deconstructed_qty), and _meta declares kind=jumpstart, mode=add."""
    from magic_manager import sets as sets_mod, parsers
    _fake_jumpstart(monkeypatch)

    out = tmp_path / "msh-jumpstart-checklist.xlsx"
    n = sets_mod.write_jumpstart_list_xlsx("msh", out, slug="msh-jumpstart")
    assert n == 2

    wb = openpyxl.load_workbook(out)
    header = [c.value for c in wb["checklist"][1]]
    assert "acquired_qty" in header
    assert "keep_qty" not in header
    assert "deconstructed_qty" not in header

    # Fill an acquired_qty cell, save, and parse it back.
    ws = wb["checklist"]
    ac = header.index("acquired_qty") + 1
    fn_col = header.index("file_name") + 1
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=fn_col).value == "IronMan_MSH":
            ws.cell(row=row, column=ac).value = 3
    wb.save(out)

    parsed = parsers.parse_jumpstart_list_xlsx(out)
    assert parsed.meta["kind"] == "jumpstart"
    assert parsed.meta["mode"] == "add"
    by_fn = {r.file_name: r for r in parsed.rows}
    assert by_fn["IronMan_MSH"].acquired_qty == 3
    assert by_fn["Wild_MSH"].acquired_qty == 0
    assert [r.file_name for r in parsed.filled_rows] == ["IronMan_MSH"]


def test_jumpstart_md_roundtrips_A_bracket(tmp_db, tmp_path, monkeypatch):
    """The MD writer emits an [A:0] bracket; editing it to [A:2] parses back to
    acquired_qty=2."""
    from magic_manager import sets as sets_mod, parsers
    _fake_jumpstart(monkeypatch)

    out = tmp_path / "msh-jumpstart-checklist.md"
    sets_mod.write_jumpstart_list_md("msh", out, slug="msh-jumpstart")
    text = out.read_text()
    assert "[A:0]" in text
    assert "[K:" not in text

    out.write_text(text.replace("IronMan_MSH — Iron Man", "IronMan_MSH — Iron Man").replace(
        "[A:0]", "[A:2]", 1))
    parsed = parsers.parse_jumpstart_list_md(out)
    # The first data line got [A:2]; both parse, one is filled.
    filled = {r.file_name: r.acquired_qty for r in parsed.filled_rows}
    assert sum(filled.values()) == 2


def test_md_parser_accepts_both_brackets():
    """The shared MD regex reads add-mode [A:n] AND modify-mode [C:c D:d P:p]."""
    from magic_manager import parsers
    import tempfile
    from pathlib import Path

    body = (
        "---\nkind: precon\nmode: add\n---\n\n"
        "- IronMan_MSH — Iron Man — R — 20 cards — $7.00 [A:3]\n"
        "- Wild_MSH — Wild — G — 20 cards — $8.00 [C:1 D:2 P:0]\n"
    )
    with tempfile.TemporaryDirectory() as d:
        p = Path(d) / "x.md"
        p.write_text(body)
        parsed = parsers.parse_jumpstart_list_md(p)
    by_fn = {r.file_name: r for r in parsed.rows}
    assert by_fn["IronMan_MSH"].acquired_qty == 3
    assert (by_fn["IronMan_MSH"].keep_qty, by_fn["IronMan_MSH"].deconstructed_qty) == (0, 0)
    # Modify-style line: construct/decon/pool populated, acquired stays 0.
    assert by_fn["Wild_MSH"].acquired_qty == 0
    assert (by_fn["Wild_MSH"].keep_qty, by_fn["Wild_MSH"].deconstructed_qty,
            by_fn["Wild_MSH"].pool_qty) == (1, 2, 0)

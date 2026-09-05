"""Deterministic card-value estimate for a sealed MTG product.

Identifies a sealed product (Booster Box, Bundle, Intro Pack, Clash Pack, …)
from MTGJSON, walks its contents recursively, and reports two independent
valuations per node:

  - **EV / deck / singles (intrinsic)** — the value of the cards inside,
    computed from MTGJSON's per-card booster weights (`ev.booster_ev`), precon
    deck singles (`sets._rollup_deck_prices`), and explicit card refs. Fully
    deterministic and offline (needs only the local synced `cards` prices).
  - **market** — an external sealed price from a pluggable provider
    (tcgcsv/tcgapi). Default `null` → market shows `(manual)` with the
    TCGplayer link. eBay (`--ebay`) is ADVISORY only (non-deterministic).

For a container (a Box = 36 Packs), it reports the whole's own market price AND
the sum of its component packs' market prices — value the whole and the parts.

Usage:
    uv run python scripts/sealed_value.py m15 "booster box"
    uv run python scripts/sealed_value.py m15 "will of the masses"
    uv run python scripts/sealed_value.py fdn --list-boosters
    uv run python scripts/sealed_value.py m15 "clash pack" --format xlsx
    uv run python scripts/sealed_value.py m15 "booster box" --market tcgcsv

Exit codes:
    0 — report written
    2 — bad invocation / product not found / no sealed data
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

from magic_manager import construct, ev, mtgjson, sealed, sets, util  # noqa: E402

QUERIES_DIR = ROOT / "queries"


# ---------- rendering ----------

def _fmt(v):
    return util.fmt_usd(v)


def _render_tree(node: sealed.ProductNode, *, depth: int = 0) -> list[str]:
    """Indented per-node lines: count× name [kind]   market / EV|deck|singles."""
    indent = "  " * depth
    count = f"{node.count}× " if node.count != 1 else ""
    market = _fmt(node.market_usd) if node.market_usd is not None else "(manual)"
    intr = _fmt(node.intrinsic_usd)
    kind_tag = {"ev": "EV", "deck": "deck", "singles": "singles",
                "variable": "≈var", "sum-of-children": "Σ"}.get(node.intrinsic_kind, "")
    line = f"{indent}{count}{node.name}"
    line = f"{line:<48} market {market:>10}   {kind_tag:>7} {intr:>10}"
    lines = [line]
    for c in node.children:
        lines.extend(_render_tree(c, depth=depth + 1))
    return lines


# ---------- top singles (reuses the construct engine — DRY) ----------

def _compute_top_singles(code: str, product_substr: str | None):
    """Expand the product into deterministic per-card singles, sorted by value
    desc, by REUSING ``construct``. Returns ``(rows, packs_skipped, error)``:
    ``rows`` is a list of ``construct.NetRow`` (finish-aware unit prices),
    ``packs_skipped`` the random-booster labels excluded, ``error`` a string if
    expansion failed (the sealed report degrades gracefully, never crashes).

    Market is forced ``null`` here — the sealed *market* price is already shown
    in the tree above; this section is the deterministic local-price singles
    breakdown ("which cards carry the value")."""
    try:
        exp = construct.expand_sealed(code, product_substr, market="null")
    except Exception as e:  # noqa: BLE001 — never break the sealed report
        return [], [], str(e)
    rows = construct.net_against_loose(exp.needs)
    return rows, exp.packs_skipped, None


def _render_top_singles(rows, packs_skipped, error, *, top_n: int = 15) -> list[str]:
    """Markdown 'Top singles' section for chat + txt. Shows the top ``top_n`` by
    unit value, the full priced-singles total, and notes for random boosters or
    a pure-booster product."""
    lines = ["", "### Top singles (by value)"]
    if error:
        lines.append(f"  (singles unavailable: {error})")
        return lines
    priced = [r for r in rows if r.unit_usd is not None]
    if not priced:
        lines.append("  No fixed singles — this product is all random boosters; "
                     "see per-booster EV above.")
        if packs_skipped:
            lines.append(f"  ({len(packs_skipped)} random booster(s): "
                         + _summarize_packs(packs_skipped) + ")")
        return lines
    lines.append("")
    lines.append("| # | Card | Set | CN | Finish | Unit $ |")
    lines.append("|---:|---|---|---|---|---:|")
    for i, r in enumerate(priced[:top_n], 1):
        link = f"[{r.name}]({construct.scryfall_url(r.set_code, r.collector_number)})"
        lines.append(f"| {i} | {link} | {r.set_code.upper()} | {r.collector_number} | "
                     f"{r.finish} | {_fmt(r.unit_usd)} |")
    total = sum(r.unit_usd * r.need_qty for r in priced)
    n_cards = sum(r.need_qty for r in priced)
    lines.append("")
    lines.append(f"Deterministic singles total {_fmt(round(total, 2))} across {n_cards} card(s)"
                 + (f"; showing top {top_n}." if len(priced) > top_n else "."))
    if packs_skipped:
        lines.append(f"  · {len(packs_skipped)} random booster(s) excluded from the singles "
                     f"table (a random pack can't be itemized) — see EV above: "
                     f"{_summarize_packs(packs_skipped)}.")
    return lines


def _summarize_packs(packs: list[str]) -> str:
    """Collapse a repetitive booster-label list to ``label ×N`` counts, e.g.
    36 identical labels → 'M15 draft booster ×36'."""
    from collections import Counter
    counts = Counter(packs)
    return ", ".join(f"{label} ×{n}" if n > 1 else label
                     for label, n in counts.items())


# ---------- XLSX artifact ----------

def _flatten(node: sealed.ProductNode, depth: int = 0, out=None):
    out = [] if out is None else out
    out.append((depth, node))
    for c in node.children:
        _flatten(c, depth + 1, out)
    return out


def _write_xlsx(node: sealed.ProductNode, market_source: str, out_path: Path,
                singles_rows=None) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "tree"
    headers = ["depth", "name", "kind", "count", "category", "market_usd",
               "market_source", "ev_usd", "deck_usd", "singles_usd",
               "ebay_advisory_usd", "tcgplayer_product_id", "purchase_url",
               "diagnostics"]
    ws.append(headers)
    for depth, n in _flatten(node):
        ev_usd = n.intrinsic_usd if n.intrinsic_kind == "ev" else None
        deck_usd = n.intrinsic_usd if n.intrinsic_kind == "deck" else None
        singles_usd = n.intrinsic_usd if n.intrinsic_kind in ("singles", "variable") else None
        ws.append([
            depth, n.name, n.intrinsic_kind or n.kind, n.count, n.category,
            n.market_usd, market_source if n.market_usd is not None else None,
            ev_usd, deck_usd, singles_usd, n.ebay_advisory_usd,
            n.tcgplayer_product_id, n.purchase_url, "; ".join(n.diagnostics),
        ])
    for row_idx in range(2, ws.max_row + 1):
        for col in (6, 8, 9, 10, 11):
            ws.cell(row=row_idx, column=col).number_format = '"$"#,##0.00'
        ws.cell(row=row_idx, column=12).number_format = "@"  # product id as text

    # Second sheet: the auditable per-sheet EV breakdown for every pack node.
    ws2 = wb.create_sheet("sheets")
    ws2.append(["node", "booster_type", "sheet", "foil", "total_weight",
                "n_cards", "n_unpriced", "ev_per_pull"])
    for _, n in _flatten(node):
        if n.ev_detail is None:
            continue
        for sname, se in n.ev_detail.sheets.items():
            ws2.append([n.name, n.ev_detail.booster_type, sname, se.foil,
                        se.total_weight, se.n_cards, se.n_unpriced,
                        round(se.ev_per_pull, 4)])

    for ws_, widths in ((ws, {2: 42, 3: 14, 5: 18, 7: 16, 13: 40, 14: 40}),
                        (ws2, {1: 30, 2: 14, 3: 16})):
        for ci, w in widths.items():
            ws_.column_dimensions[get_column_letter(ci)].width = w
        for col in range(1, ws_.max_column + 1):
            ws_.cell(row=1, column=col).font = Font(bold=True)
            ws_.cell(row=1, column=col).alignment = Alignment(horizontal="left")
        ws_.freeze_panes = "A2"
        util.apply_base_font_size(ws_)

    # Third sheet: the FULL deterministic singles table (top-value first), the
    # per-card complement to the tree's summed deck/singles nodes.
    if singles_rows:
        ws3 = wb.create_sheet("singles")
        ws3.append(["rank", "name", "set_code", "collector_number", "finish",
                    "unit_usd", "scryfall_url"])
        rank = 0
        for r in singles_rows:
            if r.unit_usd is None:
                continue
            rank += 1
            ws3.append([rank, r.name, r.set_code.upper(), r.collector_number,
                        r.finish, r.unit_usd,
                        construct.scryfall_url(r.set_code, r.collector_number)])
        for row_idx in range(2, ws3.max_row + 1):
            ws3.cell(row=row_idx, column=6).number_format = '"$"#,##0.00'
        for ci, w in {2: 34, 7: 46}.items():
            ws3.column_dimensions[get_column_letter(ci)].width = w
        for col in range(1, ws3.max_column + 1):
            ws3.cell(row=1, column=col).font = Font(bold=True)
            ws3.cell(row=1, column=col).alignment = Alignment(horizontal="left")
        ws3.freeze_panes = "A2"
        util.apply_base_font_size(ws3)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ---------- sync helper ----------

def _sync_referenced_sets(node: sealed.ProductNode) -> None:
    """Sync every set code referenced anywhere in the tree so local prices
    resolve. Best-effort — a sync failure just under-reports (surfaced as low
    coverage)."""
    codes = sealed.referenced_set_codes(node)
    unsynced = sets.unsynced_set_codes(codes)
    if unsynced:
        print(f"Syncing {len(unsynced)} referenced set(s): {', '.join(sorted(unsynced))}…")
        try:
            sets.sync(unsynced)
        except Exception as e:  # noqa: BLE001
            print(f"  ! sync failed: {e} (prices may under-report)", file=sys.stderr)


# ---------- main ----------

def main() -> int:
    ap = argparse.ArgumentParser(
        description="Estimate the card value of a sealed MTG product (EV + deck + singles + market).")
    ap.add_argument("set_code", help="Set code (e.g. m15, fdn).")
    ap.add_argument("product", nargs="?", default=None,
                    help="Product name substring (e.g. 'booster box'). Omit if the set has one product.")
    ap.add_argument("--market", choices=["null", "tcgcsv", "tcgapi", "chain", "compare"],
                    default="null", help="Market price source (default: null → manual link).")
    ap.add_argument("--ebay", action="store_true",
                    help="Also fetch eBay sold-comp advisory prices (NON-deterministic).")
    ap.add_argument("--list-boosters", action="store_true",
                    help="List the set's booster types with per-type EV, then exit.")
    ap.add_argument("--format", choices=["txt", "xlsx", "all"], default="all",
                    help="Artifact(s) to write (default: all).")
    ap.add_argument("--out-dir", type=Path, default=QUERIES_DIR,
                    help=f"Output dir (default: {QUERIES_DIR.relative_to(ROOT)}).")
    args = ap.parse_args()
    code = args.set_code.lower()

    # --list-boosters: enumerate booster types + per-type EV (used by characterize-set).
    if args.list_boosters:
        set_data = mtgjson.set_file(code)
        types = ev.booster_types(set_data)
        if not types:
            print(f"{code.upper()}: no booster data in MTGJSON.", file=sys.stderr)
            return 2
        # Sync the parent + every set any booster type pulls from (sourceSetCodes),
        # so cross-set sheets (e.g. AFR collector → AFC, set → PLST) price fully.
        boosters = set_data.get("booster") or {}
        wanted = {code}
        for t in types:
            for c in (boosters.get(t) or {}).get("sourceSetCodes") or []:
                if c:
                    wanted.add(c.lower())
        try:
            sets.sync(sets.unsynced_set_codes(wanted))
        except Exception as e:  # noqa: BLE001
            print(f"  ! sync failed: {e}", file=sys.stderr)
        set_data = mtgjson.set_file(code)
        print(f"{code.upper()} booster types ({len(types)}):")
        for t in types:
            # No shared uuid_price: let each type build its own cross-set map.
            b = ev.booster_ev(set_data, t)
            print(f"  {t:14} EV {util.fmt_usd(round(b.ev_usd, 2)):>8}  "
                  f"(coverage {b.coverage:.1%}, {b.n_configs} layout(s))")
        return 0

    # Identify the product.
    try:
        product = sealed.identify_product(code, args.product)
    except LookupError as e:
        print(f"error: {e}", file=sys.stderr)
        return 2

    market_provider = sealed.make_market_provider(args.market)
    ebay_provider = None
    if args.ebay:
        try:
            from magic_manager import ebay
            ebay_provider = ebay.EbayAdvisoryProvider()
        except Exception as e:  # noqa: BLE001
            print(f"  ! eBay provider unavailable: {e}", file=sys.stderr)

    # Build once with a null provider to discover referenced sets, sync, rebuild.
    scout = sealed.build_product_tree(code, product)
    _sync_referenced_sets(scout)
    node = sealed.build_product_tree(
        code, product, market_provider=market_provider, ebay_provider=ebay_provider)
    totals = sealed.aggregate(node)

    meta = mtgjson.meta()
    print(f"\n## Sealed value — {node.name}"
          f"   [prices as of {meta.get('date', '?')}]")
    for line in _render_tree(node):
        print(line)
    print()
    parts = _fmt(totals.market_sum_of_parts) if totals.market_sum_of_parts is not None else "—"
    whole = _fmt(totals.market_whole) if totals.market_whole is not None else "(manual)"
    print(f"TOTALS  market(whole) {whole}   market(parts) {parts}   "
          f"intrinsic {_fmt(totals.intrinsic)}   coverage {totals.coverage:.1%}")
    if node.market_usd is None and node.purchase_url:
        print(f"Market: manual — buy/price at {node.purchase_url}")
    elif args.market != "null":
        src = getattr(market_provider, "last_source", None) or market_provider.name
        print(f"Market source: {src}")
    if node.ebay_advisory_usd is not None:
        print(f"eBay advisory (non-deterministic): {_fmt(node.ebay_advisory_usd)}")
    # Compare mode: show each provider's price side-by-side for accuracy checking.
    if isinstance(market_provider, sealed.CompareMarketProvider) and market_provider.seen:
        prov_names = [p.name for p in market_provider.providers]
        print(f"Provider compare ({' vs '.join(prov_names)}):")
        header = "  " + "product".ljust(44) + "".join(n.rjust(12) for n in prov_names)
        print(header)
        for key, row in market_provider.seen.items():
            nm = (row.get("name") or str(key))[:43]
            cells = "".join(_fmt(row.get(n)).rjust(12) for n in prov_names)
            print(f"  {nm:<44}{cells}")
    if totals.diagnostics:
        print("Diagnostics:")
        for d in totals.diagnostics[:12]:
            print(f"  · {d}")

    # Top singles — always shown: reuse the construct engine to itemize which
    # cards carry the value (the per-card complement to the summed tree above).
    singles_rows, packs_skipped, singles_err = _compute_top_singles(code, args.product)
    singles_lines = _render_top_singles(singles_rows, packs_skipped, singles_err)
    for line in singles_lines:
        print(line)

    # Artifacts.
    args.out_dir.mkdir(parents=True, exist_ok=True)
    from datetime import UTC, datetime
    ts = datetime.now(UTC).strftime("%Y-%m-%d-%H%M%S")
    slug = "".join(c if c.isalnum() else "-" for c in node.name.lower()).strip("-")
    slug = "-".join(filter(None, slug.split("-")))[:60]
    written: list[Path] = []
    if args.format in ("txt", "all"):
        p = args.out_dir / f"sealed-value-{code}-{slug}-{ts}.txt"
        # txt carries the tree + the FULL singles table (no top-N truncation).
        full_singles = _render_top_singles(singles_rows, packs_skipped, singles_err,
                                           top_n=10**9)
        p.write_text("\n".join(_render_tree(node) + full_singles) + "\n", encoding="utf-8")
        written.append(p)
    if args.format in ("xlsx", "all"):
        p = args.out_dir / f"sealed-value-{code}-{slug}-{ts}.xlsx"
        src = getattr(market_provider, "last_source", None) or market_provider.name
        _write_xlsx(node, src, p, singles_rows=singles_rows)
        written.append(p)
    for p in written:
        print(f"  → {p}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

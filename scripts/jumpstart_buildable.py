"""Deterministic "buildable set" missing-cards report for a Jumpstart set.

Goal: hold the MINIMUM cards to have one built copy of every theme coexisting,
while still being able to assemble any *version* of a theme on demand — one
version of each theme constructed, plus the unique/extra cards from that theme's
other versions. This reports the cards still MISSING to reach that target, as
missing-set-style artifacts (ManaPool txt + TCGplayer txt + XLSX).

A "theme" is a Jumpstart variant name minus its trailing version suffix,
parenthesized or bare: ``Angels (1)`` / ``Angels (2)`` → theme ``Angels``;
``Corruption 1`` / ``Corruption 2`` (ONE-style naming) → theme ``Corruption``.

Target math (per scryfall_id):
  - within a theme: MAX count across that theme's versions (union at max
    multiplicity → any single version is buildable, reusing shared cards);
  - across themes:  SUM of each theme's target (all themes built at once, so a
    card used by K themes needs K copies).
  → target[card] = Σ_themes max_versions(count_in_version).

Owned = TOTAL inventory quantity per card (summed across finishes, INCLUDING
copies pledged to already-built packs — you can deconstruct to reuse them).
Finish is not tracked: a foil copy satisfies the need; the buy list is nonfoil.
missing[card] = max(0, target − owned). Basics are included.

Usage:
    uv run python scripts/jumpstart_buildable.py j25
    uv run python scripts/jumpstart_buildable.py j25 --format manapool
    uv run python scripts/jumpstart_buildable.py j25 --out-dir /tmp

Exit codes:
    0 — report written (even if nothing missing)
    2 — bad invocation / no Jumpstart variants for the set
"""
from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

from magic_manager import db, exports, mtgjson, selectors, sets, util  # noqa: E402

QUERIES_DIR = ROOT / "queries"

# A version suffix is a trailing integer, with or without parentheses:
#   'Angels (1)'    → 'Angels'   (parenthesized — e.g. J25/MSH naming)
#   'Corruption 1'  → 'Corruption' (bare space-number — e.g. ONE naming)
# The `$` anchor guarantees only a TRAILING version token is stripped, so a
# digit inside a theme name (none observed, but e.g. 'Squad 5' would keep '5')
# is only removed when it's the final token — which is exactly the version
# marker in every Jumpstart set's DeckList naming to date.
_VERSION_SUFFIX = re.compile(r"\s*\(?\d+\)?\s*$")


def theme_of(variant_name: str) -> str:
    """Strip a trailing version suffix, parenthesized or bare:
    'Angels (1)' → 'Angels', 'Corruption 1' → 'Corruption'."""
    return _VERSION_SUFFIX.sub("", variant_name or "").strip()


def build_target(variants_boards: dict[str, list[dict[str, int]]]) -> dict[str, int]:
    """Pure target computation — no DB, no network (unit-testable).

    ``variants_boards`` maps a variant NAME to a list of per-version card-count
    dicts ``{scryfall_id: count}`` (one dict per version of any theme; the key
    is the full variant name so we can group by theme here). Returns
    ``{scryfall_id: target_count}`` = Σ_themes max_versions(count).
    """
    # theme -> list of {scryfall_id: count} (one per version)
    by_theme: dict[str, list[dict[str, int]]] = defaultdict(list)
    for name, version_count in variants_boards.items():
        by_theme[theme_of(name)].append(version_count)

    target: dict[str, int] = defaultdict(int)
    for versions in by_theme.values():
        # max count per scryfall_id across this theme's versions
        theme_max: dict[str, int] = defaultdict(int)
        for vc in versions:
            for sid, n in vc.items():
                theme_max[sid] = max(theme_max[sid], n)
        # sum the theme's target into the grand total (all themes coexist)
        for sid, n in theme_max.items():
            target[sid] += n
    return dict(target)


def _variant_boards(set_code: str) -> tuple[dict[str, dict[str, int]], dict[str, str], int]:
    """For each Jumpstart variant of ``set_code``, sum card copies per
    scryfall_id across all boards. Returns ``(boards, names_by_sid, n_variants)``
    where ``boards[variant_name] = {scryfall_id: count}`` and ``names_by_sid``
    maps scryfall_id → card name (fallback display for cards absent from the
    local cards table)."""
    variants = mtgjson.jumpstart_variants(set_code)
    boards: dict[str, dict[str, int]] = {}
    names_by_sid: dict[str, str] = {}
    for v in variants:
        deck_data = mtgjson.deck(v["fileName"])
        counts: dict[str, int] = defaultdict(int)
        for board_key in ("commander", "mainBoard", "sideBoard"):
            for entry in deck_data.get(board_key) or []:
                sid = (entry.get("identifiers") or {}).get("scryfallId")
                if not sid:
                    continue
                counts[sid] += int(entry.get("count", 1) or 1)
                if entry.get("name"):
                    names_by_sid.setdefault(sid, entry["name"])
        boards[v["name"]] = dict(counts)
    return boards, names_by_sid, len(variants)


def _owned_totals(scryfall_ids: list[str]) -> dict[str, int]:
    """Total inventory quantity per scryfall_id, summed across finishes
    (includes copies pledged to built decks — deconstructable, so they count)."""
    if not scryfall_ids:
        return {}
    with db.connect() as conn:
        placeholders = ",".join("?" for _ in scryfall_ids)
        return {
            r["scryfall_id"]: r["q"]
            for r in conn.execute(
                f"SELECT scryfall_id, SUM(quantity) AS q FROM inventory "
                f"WHERE scryfall_id IN ({placeholders}) GROUP BY scryfall_id",
                scryfall_ids,
            ).fetchall()
        }


def _card_dicts(scryfall_ids: list[str]) -> dict[str, dict]:
    """Local cards-table rows normalized via selectors._card_dict, by scryfall_id."""
    if not scryfall_ids:
        return {}
    with db.connect() as conn:
        placeholders = ",".join("?" for _ in scryfall_ids)
        return {
            r["scryfall_id"]: selectors._card_dict(r)
            for r in conn.execute(
                f"SELECT {selectors._CARD_COLS} FROM cards c "
                f"WHERE c.scryfall_id IN ({placeholders})",
                scryfall_ids,
            ).fetchall()
        }


def _write_xlsx(rows: list, out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "buildable-missing"
    headers = ["set", "collector_number", "name", "rarity", "finish",
               "qty", "unit_usd", "line_value"]
    ws.append(headers)
    for r in rows:
        c = r.card
        unit = c.get("prices_usd")
        line = (unit * r.quantity) if unit is not None else None
        ws.append([
            (c.get("set") or "").upper(),
            c.get("collector_number"),
            c.get("name"),
            c.get("rarity"),
            r.finish,
            r.quantity,
            unit,
            line,
        ])
        ws.cell(row=ws.max_row, column=2).number_format = "@"  # CN as text
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=7).number_format = '"$"#,##0.00'
        ws.cell(row=row_idx, column=8).number_format = '"$"#,##0.00'
    for ci, w in {1: 6, 2: 8, 3: 40, 4: 10, 5: 9, 6: 6, 7: 10, 8: 11}.items():
        ws.column_dimensions[get_column_letter(ci)].width = w
    for col in range(1, ws.max_column + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="left")
    ws.freeze_panes = "A2"
    util.apply_base_font_size(ws)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Report cards missing to make every Jumpstart theme buildable.")
    ap.add_argument("set_code", help="Jumpstart set code (e.g. j25).")
    ap.add_argument("--format", choices=["manapool", "tcgplayer", "xlsx", "all"],
                    default="all", help="Which artifact(s) to write (default: all).")
    ap.add_argument("--out-dir", type=Path, default=QUERIES_DIR,
                    help=f"Output dir (default: {QUERIES_DIR.relative_to(ROOT)}).")
    args = ap.parse_args()
    code = args.set_code.lower()

    variants = mtgjson.jumpstart_variants(code)
    if not variants:
        print(f"error: no Jumpstart variants found for set {code!r}. "
              f"Check `mm mtgjson decks --set {code}`.", file=sys.stderr)
        return 2

    # Sync the family so cards/prices resolve locally.
    print(f"Syncing {code.upper()} family for prices…")
    try:
        sets.sync(sets.resolve(code).filtered_codes())
    except Exception as e:  # noqa: BLE001 — best-effort, mirrors sets.py contract
        print(f"  ! sync failed: {e} (prices/names may under-report)", file=sys.stderr)

    boards, names_by_sid, n_variants = _variant_boards(code)
    n_themes = len({theme_of(n) for n in boards})
    target = build_target(boards)

    owned = _owned_totals(list(target))
    missing = {sid: target[sid] - owned.get(sid, 0)
               for sid in target if target[sid] - owned.get(sid, 0) > 0}

    cards = _card_dicts(list(missing))
    rows: list = []
    skipped: list[str] = []
    for sid, qty in missing.items():
        card = cards.get(sid)
        if card is None:
            skipped.append(names_by_sid.get(sid, sid))
            continue
        rows.append(selectors.MaterializedRow(
            scryfall_id=sid, quantity=qty, finish="nonfoil", card=card))
    # Deterministic order: (set, collector-number).
    rows.sort(key=lambda r: ((r.card.get("set") or ""),
                             util.cn_sort_key(r.card.get("collector_number"))))

    args.out_dir.mkdir(parents=True, exist_ok=True)
    from datetime import UTC, datetime
    ts = datetime.now(UTC).strftime("%Y-%m-%d-%H%M%S")
    written: list[Path] = []
    if args.format in ("manapool", "all"):
        p = args.out_dir / f"buildable-{code}-manapool-{ts}.txt"
        p.write_text(exports.build("manapool", rows), encoding="utf-8")
        written.append(p)
    if args.format in ("tcgplayer", "all"):
        p = args.out_dir / f"buildable-{code}-tcgplayer-{ts}.txt"
        p.write_text(exports.build("tcgplayer", rows), encoding="utf-8")
        written.append(p)
    if args.format in ("xlsx", "all"):
        p = args.out_dir / f"buildable-{code}-checklist-{ts}.xlsx"
        _write_xlsx(rows, p)
        written.append(p)

    target_total = sum(target.values())
    missing_total = sum(missing.values())
    buy_usd = sum((r.card.get("prices_usd") or 0.0) * r.quantity for r in rows)
    print(f"\n{code.upper()} buildable set — {n_themes} theme(s) across {n_variants} variant(s)")
    print(f"  target:  {len(target)} distinct cards / {target_total} copies")
    print(f"  owned:   {sum(owned.values())} copies of target cards")
    print(f"  missing: {len(rows)} distinct / {missing_total} copies · ${buy_usd:,.2f} to buy")
    if skipped:
        print(f"  ! {len(skipped)} missing card(s) not in local cards table (name-only): "
              f"{', '.join(sorted(set(skipped))[:10])}"
              + (" …" if len(set(skipped)) > 10 else ""), file=sys.stderr)
    for p in written:
        print(f"  → {p}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

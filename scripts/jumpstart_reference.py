"""Deterministic two-sheet XLSX reference of Jumpstart pack versions.

Answers "which version of a Jumpstart pack is this?" (Angels (1) vs Angels (2),
etc.) and "what's in each version?". Emits ``reference/jumpstart-versions.xlsx``
with two sheets:

  - **packs** — one row per Jumpstart pack variant:
    ``set, theme, color, top_card, top_card_usd, card_count, usd_total``.
    (A reference, not an ingestible checklist — no keep/deconstruct qty fields.)
  - **cards** — one row per distinct card in each pack:
    ``set, theme, color, card_name, card_value, count`` (count = copies in pack).

Both sheets sort by COLOR then NAME (theme / card) A→Z. Color order is
``C → W → U → B → R → G`` for single symbols, then every multicolor code as one
trailing block (ordered by its letter sequence) — i.e. mono-first, then multi.

Color is the deck/card convention (actual WUBRG letters, no 'M' collapse; a
colorless card/pack → 'C'), matching the checklist writers. Card value is the
card's SHIPPED finish (foil price if the pack ships it foil, else nonfoil) —
the same basis the pack ``usd_total`` uses. Pack ``usd_total`` also folds in
the decorative front/title card price (as the checklist does); the front card
itself is NOT listed on either sheet (it isn't a version signal).

Prices/colors come from the local ``cards`` table, so the script syncs each
referenced set's family first (Jumpstart contents span the parent expansion).

Usage:
    uv run python scripts/jumpstart_reference.py            # ALL jumpstart sets
    uv run python scripts/jumpstart_reference.py j25        # one set
    uv run python scripts/jumpstart_reference.py --out /tmp/jr.xlsx

Exit codes:
    0 — written
    2 — bad invocation / no Jumpstart variants found
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

from magic_manager import db, mtgjson, sets, util  # noqa: E402

DEFAULT_OUT = ROOT / "reference" / "jumpstart-versions.xlsx"

# Color-code sort: single symbols rank C<W<U<B<R<G; any multicolor code sorts
# as one trailing block (rank 6), ordered among itself by its letter sequence.
_MONO_RANK = {"C": 0, "W": 1, "U": 2, "B": 3, "R": 4, "G": 5}


def _color_sort_key(code: str) -> tuple[int, str]:
    """Sort key for a color code. Mono/colorless first in C,W,U,B,R,G order;
    every multicolor code (2+ letters) after, grouped and ordered by sequence."""
    if code in _MONO_RANK:
        return (_MONO_RANK[code], "")
    return (6, code)  # multicolor block ("M" tier), ordered by letters


def _all_jumpstart_set_codes() -> list[str]:
    """Every set code that publishes ``type: Jumpstart`` products, from the
    global DeckList (deduped, lowercased, sorted)."""
    codes = {
        (d.get("code") or "").lower()
        for d in mtgjson.deck_list()
        if d.get("type") == "Jumpstart" and d.get("code")
    }
    return sorted(codes)


def _sync_family(code: str) -> None:
    """Sync ``code``'s family into the local cards table so prices/colors
    resolve (Jumpstart contents span the parent expansion). Best-effort:
    a resolution/sync failure leaves the set's rows under-reported, not fatal."""
    try:
        resolved = sets.resolve(code)
        sets.sync(resolved.filtered_codes())
    except Exception as e:  # noqa: BLE001 — best-effort, mirrors sets.py contract
        print(f"  ! sync failed for {code}: {e}", file=sys.stderr)


def _card_rows_for_variant(code: str, variant: dict, summary: dict) -> tuple[list[dict], int]:
    """Per-distinct-card rows for one pack. Returns (rows, n_skipped) where
    n_skipped counts scryfall_ids absent from the local cards table."""
    deck_data = mtgjson.deck(variant["fileName"])
    theme = summary["theme"]

    # Accumulate copies per (scryfall_id, is_foil): the shipped finish is part
    # of the identity, so a card shipped both foil and nonfoil (rare) is two
    # rows. count = summed copies of that (printing, finish) across boards.
    counts: dict[tuple[str, bool], int] = {}
    order: list[tuple[str, bool]] = []
    for board_key in ("commander", "mainBoard", "sideBoard"):
        for entry in deck_data.get(board_key) or []:
            sid = (entry.get("identifiers") or {}).get("scryfallId")
            if not sid:
                continue
            key = (sid, bool(entry.get("isFoil")))
            if key not in counts:
                order.append(key)
            counts[key] = counts.get(key, 0) + int(entry.get("count", 1) or 1)

    sids = [k[0] for k in order]
    card_data: dict[str, tuple] = {}
    if sids:
        with db.connect() as conn:
            placeholders = ",".join("?" for _ in sids)
            card_data = {
                r["scryfall_id"]: (r["name"], r["prices_usd"],
                                   r["prices_usd_foil"], r["color_identity"])
                for r in conn.execute(
                    f"SELECT scryfall_id, name, prices_usd, prices_usd_foil, "
                    f"color_identity FROM cards WHERE scryfall_id IN ({placeholders})",
                    sids,
                ).fetchall()
            }

    rows: list[dict] = []
    n_skipped = 0
    for (sid, is_foil) in order:
        data = card_data.get(sid)
        if data is None:
            n_skipped += 1
            continue
        name, nonfoil, foil, ci = data
        price = foil if is_foil else nonfoil
        rows.append({
            "set": code.upper(),
            "theme": theme,
            "color": util.format_color_identity(ci, collapse_multicolor=False),
            "card_name": name,
            "card_value": float(price) if price is not None else None,
            "count": counts[(sid, is_foil)],
        })
    return rows, n_skipped


def _gather(codes: list[str]) -> tuple[list[dict], list[dict]]:
    """Build (pack_rows, card_rows) across the given set codes."""
    pack_rows: list[dict] = []
    card_rows: list[dict] = []
    total_skipped = 0
    for code in codes:
        variants = mtgjson.jumpstart_variants(code)
        if not variants:
            print(f"  (no Jumpstart variants for {code.upper()}, skipping)", file=sys.stderr)
            continue
        print(f"  {code.upper()}: syncing family + {len(variants)} pack(s)…")
        _sync_family(code)
        for v in variants:
            summary = sets._jumpstart_variant_summary(v, anchor=code)
            pack_rows.append({
                "set": code.upper(),
                "theme": summary["theme"],
                "color": summary["color"],
                "top_card": summary["top_card"],
                "top_card_usd": summary["top_card_usd"],
                "card_count": summary["card_count"],
                "usd_total": summary["usd_total"],
            })
            crows, skipped = _card_rows_for_variant(code, v, summary)
            card_rows.extend(crows)
            total_skipped += skipped
    if total_skipped:
        print(f"  ! {total_skipped} card printing(s) not found locally, omitted from cards sheet",
              file=sys.stderr)
    return pack_rows, card_rows


def _write_xlsx(pack_rows: list[dict], card_rows: list[dict], out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    # Sort both sheets: color (C→W→U→B→R→G→multi), then name A→Z.
    pack_rows.sort(key=lambda r: (_color_sort_key(r["color"]), (r["theme"] or "").lower()))
    card_rows.sort(key=lambda r: (_color_sort_key(r["color"]), (r["card_name"] or "").lower()))

    wb = Workbook()

    ws_p = wb.active
    ws_p.title = "packs"
    p_headers = ["set", "theme", "color", "top_card", "top_card_usd",
                 "card_count", "usd_total"]
    ws_p.append(p_headers)
    for r in pack_rows:
        ws_p.append([r["set"], r["theme"], r["color"], r["top_card"],
                     r["top_card_usd"], r["card_count"], r["usd_total"]])
    # currency: top_card_usd (col 5), usd_total (col 7)
    for row_idx in range(2, ws_p.max_row + 1):
        ws_p.cell(row=row_idx, column=5).number_format = '"$"#,##0.00'
        ws_p.cell(row=row_idx, column=7).number_format = '"$"#,##0.00'
    _p_widths = {1: 6, 2: 26, 3: 8, 4: 28, 5: 12, 6: 11, 7: 11}
    for ci, w in _p_widths.items():
        ws_p.column_dimensions[get_column_letter(ci)].width = w

    ws_c = wb.create_sheet("cards")
    c_headers = ["set", "theme", "color", "card_name", "card_value", "count"]
    ws_c.append(c_headers)
    for r in card_rows:
        ws_c.append([r["set"], r["theme"], r["color"], r["card_name"],
                     r["card_value"], r["count"]])
    for row_idx in range(2, ws_c.max_row + 1):
        ws_c.cell(row=row_idx, column=5).number_format = '"$"#,##0.00'
    _c_widths = {1: 6, 2: 26, 3: 8, 4: 34, 5: 12, 6: 7}
    for ci, w in _c_widths.items():
        ws_c.column_dimensions[get_column_letter(ci)].width = w

    for ws in (ws_p, ws_c):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="left")
        ws.freeze_panes = "A2"
        util.apply_base_font_size(ws)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> int:
    ap = argparse.ArgumentParser(description="Generate the Jumpstart versions reference XLSX.")
    ap.add_argument("set_code", nargs="?", default=None,
                    help="Jumpstart set code (e.g. j25). Omit for ALL Jumpstart sets.")
    ap.add_argument("--out", type=Path, default=DEFAULT_OUT,
                    help=f"Output path (default: {DEFAULT_OUT.relative_to(ROOT)}).")
    args = ap.parse_args()

    if args.set_code:
        codes = [args.set_code.lower()]
    else:
        print("Discovering all Jumpstart sets from MTGJSON DeckList…")
        codes = _all_jumpstart_set_codes()
        if not codes:
            print("error: no Jumpstart sets found in DeckList.", file=sys.stderr)
            return 2
        print(f"  {len(codes)} set(s): {', '.join(c.upper() for c in codes)}")

    pack_rows, card_rows = _gather(codes)
    if not pack_rows:
        print(f"error: no Jumpstart variants found for {codes}.", file=sys.stderr)
        return 2

    _write_xlsx(pack_rows, card_rows, args.out)
    print(f"\nWrote {len(pack_rows)} pack row(s) + {len(card_rows)} card row(s) "
          f"across {len(codes)} set(s) to {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

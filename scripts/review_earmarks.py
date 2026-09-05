"""Deterministic review of earmarked sealed products (a watchlist deal table).

Reads the earmark watchlist (`earmarks.earmark_list`) and, for each product,
**recomputes** its live market + intrinsic value by REUSING the sealed engine
(`sealed.build_product_tree` / `aggregate`, the same path `sealed_value.py`
drives) — the DB stores only the non-derivable asking-price snapshot, never
derived values. Emits a markdown deal table (product names hyperlinked to their
storefronts, collated across stores) + a txt/xlsx to `queries/`.

Columns: product (+ per-store links & asking prices), set, category, release,
best asking $, live market $, live intrinsic $, deal delta (market − best
asking), and the age of the asking-price snapshot. Sorted by deal delta desc
(best deals first).

Usage:
    uv run python scripts/review_earmarks.py
    uv run python scripts/review_earmarks.py --market compare
    uv run python scripts/review_earmarks.py --format txt

Exit codes:
    0 — report written (or nothing earmarked)
    2 — unexpected error
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

from magic_manager import earmarks, sealed, sets, util  # noqa: E402

QUERIES_DIR = ROOT / "queries"


def _fmt(v) -> str:
    return util.fmt_usd(v)


def _age_days(captured_at: str, today: str) -> int | None:
    """Whole days between an ISO ``captured_at`` and an ISO ``today`` (both
    ``YYYY-MM-DD…``). ``None`` if either is unparseable. Date math only — no
    ``datetime.now()`` in the script (today is passed in, keeping it
    deterministic/testable)."""
    from datetime import date
    try:
        c = date.fromisoformat(captured_at[:10])
        t = date.fromisoformat(today[:10])
    except (ValueError, TypeError):
        return None
    return (t - c).days


# ---------- per-product live valuation (reuses the sealed engine) ----------

def _value_product(set_code: str, product_name: str, market_provider) -> dict:
    """Recompute market + intrinsic for one product via the sealed engine.

    Mirrors ``sealed_value.py``'s flow: identify → scout-build to discover
    referenced sets → sync → rebuild with the market provider → aggregate.
    Returns ``{"market": float|None, "intrinsic": float|None, "error": str|None}``.
    """
    try:
        product = sealed.identify_product(set_code, product_name)
    except LookupError as e:
        return {"market": None, "intrinsic": None, "error": str(e)}

    # Discover + sync referenced sets so local prices resolve (best-effort).
    scout = sealed.build_product_tree(set_code, product)
    codes: set[str] = set()

    def _walk(n):
        codes.add(n.set_code.lower())
        for c in n.children:
            _walk(c)
    _walk(scout)
    unsynced = sets.unsynced_set_codes(codes)
    if unsynced:
        try:
            sets.sync(unsynced)
        except Exception as e:  # noqa: BLE001 — a sync failure just under-reports
            print(f"  ! sync failed for {set_code}: {e}", file=sys.stderr)

    node = sealed.build_product_tree(set_code, product, market_provider=market_provider)
    totals = sealed.aggregate(node)
    return {"market": totals.market_whole, "intrinsic": totals.intrinsic, "error": None}


# ---------- rendering ----------

def _product_cell(p) -> str:
    """The Product column: name linked to its cheapest store, plus a parenthetical
    list of the other stores with their asking prices."""
    if not p.links:
        return p.product_name
    # links are pre-sorted cheapest-first by earmark_list
    primary = p.links[0]
    cell = f"[{p.product_name}]({primary.store_url})"
    store_bits = []
    for l in p.links:
        px = _fmt(l.asking_price) if l.asking_price is not None else "—"
        store_bits.append(f"[{l.store_name or 'store'}]({l.store_url}) {px}")
    return cell + "<br>" + " · ".join(store_bits)


def _build_rows(products, market_provider, today: str) -> list[dict]:
    """Value every product and assemble sortable row dicts."""
    rows = []
    for p in products:
        val = _value_product(p.set_code, p.product_name, market_provider)
        best = p.best_asking
        market = val["market"]
        delta = (market - best) if (market is not None and best is not None) else None
        ages = [d for d in (_age_days(l.captured_at, today) for l in p.links) if d is not None]
        rows.append({
            "product": p,
            "market": market,
            "intrinsic": val["intrinsic"],
            "best_asking": best,
            "delta": delta,
            "age": min(ages) if ages else None,
            "error": val["error"],
        })
    # Best deals first: rows with a delta sort desc; None-delta rows sink to the
    # bottom (ordered by name for stability).
    rows.sort(key=lambda r: (r["delta"] is None,
                             -(r["delta"] if r["delta"] is not None else 0.0),
                             r["product"].product_name))
    return rows


def _render_lines(rows, today: str) -> list[str]:
    lines = [f"## Earmarked products — deal review   [as of {today}]", ""]
    lines.append("| Product / stores | Set | Category | Release | Best ask | Market | Intrinsic | Deal Δ | Ask age |")
    lines.append("|---|---|---|---|---:|---:|---:|---:|---:|")
    for r in rows:
        p = r["product"]
        delta = r["delta"]
        delta_cell = _fmt(delta) if delta is not None else "—"
        if delta is not None and delta > 0:
            delta_cell = f"**+{delta_cell.lstrip('$')}**" if delta_cell.startswith("$") else delta_cell
        age = f"{r['age']}d" if r["age"] is not None else "—"
        lines.append(
            f"| {_product_cell(p)} | {p.set_code.upper()} | {p.category or '—'} | "
            f"{p.release_date or '—'} | {_fmt(r['best_asking'])} | {_fmt(r['market'])} | "
            f"{_fmt(r['intrinsic'])} | {delta_cell} | {age} |"
        )
    lines.append("")
    n = len(rows)
    good = sum(1 for r in rows if r["delta"] is not None and r["delta"] > 0)
    lines.append(f"TOTALS  {n} product(s) earmarked   {good} priced below live market "
                 f"(positive Deal Δ = market exceeds asking = a deal)")
    for r in rows:
        if r["error"]:
            lines.append(f"  · {r['product'].product_name}: {r['error']}")
    return lines


# ---------- XLSX artifact ----------

def _write_xlsx(rows, today: str, out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "earmarks"
    headers = ["product_name", "set_code", "category", "release_date", "best_asking",
               "market", "intrinsic", "deal_delta", "ask_age_days", "n_stores",
               "store_urls"]
    ws.append(headers)
    for r in rows:
        p = r["product"]
        ws.append([
            p.product_name, p.set_code.upper(), p.category, p.release_date,
            r["best_asking"], r["market"], r["intrinsic"], r["delta"], r["age"],
            len(p.links), " | ".join(l.store_url for l in p.links),
        ])
    for row_idx in range(2, ws.max_row + 1):
        for col in (5, 6, 7, 8):
            ws.cell(row=row_idx, column=col).number_format = '"$"#,##0.00'
    for ci, w in {1: 42, 3: 14, 11: 60}.items():
        ws.column_dimensions[get_column_letter(ci)].width = w
    for col in range(1, ws.max_column + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="left")
    ws.freeze_panes = "A2"
    util.apply_base_font_size(ws)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ---------- main ----------

def main() -> int:
    ap = argparse.ArgumentParser(
        description="Review earmarked sealed products: live market/intrinsic vs asking price.")
    ap.add_argument("--market", choices=["null", "tcgcsv", "tcgapi", "chain", "compare"],
                    default="tcgcsv", help="Live market price source (default: tcgcsv).")
    ap.add_argument("--format", choices=["txt", "xlsx", "all"], default="all",
                    help="Artifact(s) to write (default: all).")
    ap.add_argument("--out-dir", type=Path, default=QUERIES_DIR,
                    help=f"Output dir (default: {QUERIES_DIR.relative_to(ROOT)}).")
    args = ap.parse_args()

    products = earmarks.earmark_list()
    if not products:
        print("(no earmarked products — use /earmark-product <store-URL> to add one)")
        return 0

    from datetime import UTC, datetime
    now = datetime.now(UTC)
    today = now.strftime("%Y-%m-%d")

    market_provider = sealed.make_market_provider(args.market)
    rows = _build_rows(products, market_provider, today)
    lines = _render_lines(rows, today)
    print("\n" + "\n".join(lines))

    args.out_dir.mkdir(parents=True, exist_ok=True)
    ts = now.strftime("%Y-%m-%d-%H%M%S")
    written: list[Path] = []
    if args.format in ("txt", "all"):
        p = args.out_dir / f"earmarks-review-{ts}.txt"
        p.write_text("\n".join(lines) + "\n", encoding="utf-8")
        written.append(p)
    if args.format in ("xlsx", "all"):
        p = args.out_dir / f"earmarks-review-{ts}.xlsx"
        _write_xlsx(rows, today, p)
        written.append(p)
    for p in written:
        print(f"  → {p}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

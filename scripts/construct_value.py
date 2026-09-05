"""Value the cost to CONSTRUCT a decklist or sealed product from singles.

Given a sealed product OR a decklist, reports three valuations and a per-card
table:

  1. **sealed**          — cost to buy the sealed product (sealed input only).
  2. **scratch**         — buy every card as a net-new single.
  3. **with-collection** — use my LOOSE (unpledged) copies first, buy net-new
                           only for the shortfall.

All valuation logic lives in ``magic_manager.construct`` (which reuses the
sealed engine, the deck-price rollup, the shared price map and the loose-
inventory helper). This script is a thin CLI: parse args → expand → net →
summarize → render table + write artifacts. No arithmetic lives here.

Input forms (exactly one):
    construct_value.py <set_code> [product-substr]     # a sealed product
    construct_value.py --deck-file AncientArsenal_ACR   # an MTGJSON precon
    construct_value.py --slug atraxa-superfriends        # a local deck
    construct_value.py --decklist <path|->               # a pasted block

Deck URLs (Moxfield/Archidekt/MTGGoldfish) are handled by the skill's Claude-
side recipe (WebFetch → normalize to text → --decklist <tmp>); this script stays
deterministic and offline apart from the price/market sync.

Usage:
    uv run python scripts/construct_value.py acr "Assassins Creed Starter Kit" --market compare
    uv run python scripts/construct_value.py --deck-file AncientArsenal_ACR
    uv run python scripts/construct_value.py --decklist - < list.txt

Exit codes:
    0 — report written
    2 — bad invocation / product or deck not found / nothing to value
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

from magic_manager import construct, mtgjson, util  # noqa: E402

QUERIES_DIR = ROOT / "queries"


def _fmt(v):
    return util.fmt_usd(v)


def _scry_url(set_code: str, cn: str) -> str:
    return f"https://scryfall.com/card/{set_code.lower()}/{cn}"


# ---------- rendering ----------

def _render_lines(exp: construct.Expansion, rows: list[construct.NetRow],
                  summary: dict, *, is_sealed: bool) -> list[str]:
    """Markdown-ish plain text: a header, the card table, the TOTALS line, and
    diagnostics. The same body is printed to stdout and written to the .txt."""
    lines: list[str] = []
    meta = mtgjson.meta()
    kind = "sealed product" if is_sealed else "decklist"
    lines.append(f"## Construct value — {exp.label}  ({kind})"
                 f"   [prices as of {meta.get('date', '?')}]")
    lines.append("")
    lines.append("| Card | Set | CN | Finish | Need | Loose | Buy | Unit $ | Buy $ |")
    lines.append("|---|---|---|---|---:|---:|---:|---:|---:|")
    for r in rows:
        name_link = f"[{r.name}]({_scry_url(r.set_code, r.collector_number)})"
        unit = _fmt(r.unit_usd) if r.unit_usd is not None else "—"
        buy = _fmt(r.buy_usd) if r.buy_usd is not None else "—"
        lines.append(
            f"| {name_link} | {r.set_code.upper()} | {r.collector_number} | "
            f"{r.finish} | {r.need_qty} | {r.loose_qty} | {r.buy_qty} | {unit} | {buy} |"
        )
    lines.append("")
    sealed_cell = _fmt(summary["sealed"]) if summary["sealed"] is not None else (
        "(manual)" if is_sealed else "n/a")
    lines.append(
        f"TOTALS  sealed {sealed_cell}   scratch {_fmt(summary['scratch'])}   "
        f"with-collection {_fmt(summary['with_collection'])}   "
        f"coverage {summary['coverage']:.1%}"
    )
    if summary["n_unpriced"]:
        lines.append(f"  ! {summary['n_unpriced']} card(s) unpriced "
                     f"(scratch/with-collection under-report by their value)")
    for d in exp.diagnostics:
        lines.append(f"  · {d}")
    return lines


# ---------- XLSX artifact ----------

def _write_xlsx(exp: construct.Expansion, rows: list[construct.NetRow],
                summary: dict, *, is_sealed: bool, out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "cards"
    headers = ["name", "set_code", "collector_number", "finish", "need_qty",
               "loose_qty", "buy_qty", "unit_usd", "scratch_usd", "buy_usd",
               "scryfall_url"]
    ws.append(headers)
    for r in rows:
        ws.append([
            r.name, r.set_code.upper(), r.collector_number, r.finish,
            r.need_qty, r.loose_qty, r.buy_qty, r.unit_usd,
            r.scratch_usd, r.buy_usd, _scry_url(r.set_code, r.collector_number),
        ])
    for row_idx in range(2, ws.max_row + 1):
        for col in (8, 9, 10):
            ws.cell(row=row_idx, column=col).number_format = '"$"#,##0.00'

    # Second sheet: the three headline totals + diagnostics.
    ws2 = wb.create_sheet("summary")
    ws2.append(["metric", "value"])
    ws2.append(["input", exp.label])
    ws2.append(["kind", "sealed product" if is_sealed else "decklist"])
    ws2.append(["sealed_usd", summary["sealed"]])
    ws2.append(["scratch_usd", summary["scratch"]])
    ws2.append(["with_collection_usd", summary["with_collection"]])
    ws2.append(["coverage", round(summary["coverage"], 4)])
    ws2.append(["n_unpriced", summary["n_unpriced"]])
    ws2.append(["total_need", summary["total_need"]])
    for label in exp.packs_skipped:
        ws2.append(["random_pack_excluded", label])
    for d in exp.diagnostics:
        ws2.append(["diagnostic", d])

    for ws_, widths in ((ws, {1: 34, 2: 10, 4: 10, 11: 46}),
                        (ws2, {1: 24, 2: 46})):
        for ci, w in widths.items():
            ws_.column_dimensions[get_column_letter(ci)].width = w
        for col in range(1, ws_.max_column + 1):
            ws_.cell(row=1, column=col).font = Font(bold=True)
            ws_.cell(row=1, column=col).alignment = Alignment(horizontal="left")
        ws_.freeze_panes = "A2"
        util.apply_base_font_size(ws_)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ---------- main ----------

def _read_text_or_stdin(source: str) -> str:
    if source == "-":
        return sys.stdin.read()
    p = Path(source)
    if not p.exists():
        raise FileNotFoundError(f"no such file: {source}")
    return p.read_text(encoding="utf-8")


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Cost to construct a decklist or sealed product from singles "
                    "(scratch vs using your loose collection).")
    ap.add_argument("set_code", nargs="?", default=None,
                    help="Set code for a sealed product (e.g. acr, m15).")
    ap.add_argument("product", nargs="?", default=None,
                    help="Sealed product name substring (omit if the set has one product).")
    ap.add_argument("--deck-file", default=None,
                    help="MTGJSON precon deck fileName (e.g. AncientArsenal_ACR).")
    ap.add_argument("--slug", default=None,
                    help="Local deck slug (uses its stored deck_cards recipe).")
    ap.add_argument("--decklist", default=None,
                    help="Path to a Moxfield-style block, or '-' for stdin.")
    ap.add_argument("--market", choices=["null", "tcgcsv", "tcgapi", "chain", "compare"],
                    default="tcgcsv",
                    help="Sealed market price source (default: tcgcsv; sealed input only).")
    ap.add_argument("--format", choices=["txt", "xlsx", "all"], default="all",
                    help="Artifact(s) to write (default: all).")
    ap.add_argument("--out-dir", type=Path, default=QUERIES_DIR,
                    help=f"Output dir (default: {QUERIES_DIR.relative_to(ROOT)}).")
    args = ap.parse_args()

    # Exactly one input form.
    forms = [bool(args.set_code), bool(args.deck_file), bool(args.slug), bool(args.decklist)]
    if sum(forms) != 1:
        print("error: provide exactly one input — a <set_code> [product], "
              "--deck-file, --slug, or --decklist.", file=sys.stderr)
        return 2

    is_sealed = bool(args.set_code)
    try:
        if is_sealed:
            exp = construct.expand_sealed(args.set_code.lower(), args.product,
                                          market=args.market)
        elif args.deck_file:
            exp = construct.expand_deck_file(args.deck_file)
        elif args.slug:
            exp = construct.expand_slug(args.slug)
        else:
            text = _read_text_or_stdin(args.decklist)
            label = "decklist" if args.decklist == "-" else Path(args.decklist).stem
            exp = construct.expand_decklist_text(text, label=label)
    except (LookupError, FileNotFoundError) as e:
        print(f"error: {e}", file=sys.stderr)
        return 2

    if not exp.needs:
        print(f"error: no constructable cards found for {exp.label!r}"
              + (" (only random boosters?)" if exp.packs_skipped else ""), file=sys.stderr)
        for d in exp.diagnostics:
            print(f"  · {d}", file=sys.stderr)
        return 2

    rows = construct.net_against_loose(exp.needs)
    summary = construct.summarize(rows, exp.sealed_market)

    lines = _render_lines(exp, rows, summary, is_sealed=is_sealed)
    print("\n" + "\n".join(lines))

    # Artifacts.
    args.out_dir.mkdir(parents=True, exist_ok=True)
    from datetime import UTC, datetime
    ts = datetime.now(UTC).strftime("%Y-%m-%d-%H%M%S")
    slug = "".join(c if c.isalnum() else "-" for c in exp.label.lower()).strip("-")
    slug = "-".join(filter(None, slug.split("-")))[:60] or "construct"
    written: list[Path] = []
    if args.format in ("txt", "all"):
        p = args.out_dir / f"construct-value-{slug}-{ts}.txt"
        p.write_text("\n".join(lines) + "\n", encoding="utf-8")
        written.append(p)
    if args.format in ("xlsx", "all"):
        p = args.out_dir / f"construct-value-{slug}-{ts}.xlsx"
        _write_xlsx(exp, rows, summary, is_sealed=is_sealed, out_path=p)
        written.append(p)
    for p in written:
        print(f"  → {p}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

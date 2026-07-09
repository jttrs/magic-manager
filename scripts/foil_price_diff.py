"""Rank a list of Scryfall printings by the price gap between simple-foil and
nonfoil finishes, sorted ascending by percent difference.

"Simple foil" means the plain foil finish — this script EXCLUDES fancy foils
(surgefoil, etched, textured, rainbowfoil, chocobotrackfoil, etc., anything
that ``treatments.compute_treatment()`` maps to the ``ff`` keyword). Those are
separate premium products; the foil-vs-nonfoil decision they represent is
different (you're not choosing a finish, you're choosing a whole product tier).

Every price is fetched live via Scryfall's ``/cards/collection`` batch
endpoint through the project's rate-limited wrapper. Prices are cached
inside the wrapper (24h TTL) so re-runs the same day are instant.

Input modes:
  - stdin (canonical): Moxfield-style lines like ``1 Card Name (SET) CN``
    or ``1 Card Name (SET) CN *F*``. Quantity and foil marker are ignored;
    the script dedupes on (set, cn) and always evaluates BOTH finishes.
  - ``--file PATH`` where PATH ends in ``.txt`` → same Moxfield parse.
  - ``--file PATH`` where PATH ends in ``.xlsx`` → reads the visible sheet
    and takes the ``set`` + ``collector_number`` columns. Works for both
    master-list checklists and missing-set result files.

Output:
  - stdout: a markdown table sorted ascending by % diff. Columns:
    Card (hyperlinked), Nonfoil, Foil, % diff, $ diff.
  - stderr: a one-line summary of counts by bucket
    (fancy-foil, foil-only, nonfoil-only, unpriced, unresolved, filtered).

Optional post-sort filter flags (all inclusive; combine with AND):
  --min-pct / --max-pct     percent-diff bounds (in percent, e.g. -14, 100)
  --min-raw / --max-raw     dollar-diff bounds (in USD, e.g. -0.50, 10)
  --drop-expensive PCT:RAW  drop rows where BOTH %-diff > PCT AND $-diff >= RAW
                            (e.g. 100:10 keeps cheap-but-high-multiple rows
                             like $0.30→$2.36 = +594%/+$2.02 while dropping
                             expensive-and-high-multiple ones like
                             $195→$423 = +117%/+$228)

Exit codes:
  0 — ran to completion (even if no ranked rows; empty stdout, summary on stderr).
  2 — bad invocation or unreadable file.

Determinism notes:
  - Sort key is (round(pct, 4), name, set, cn_int_or_str_padded). Rounding
    before compare stabilizes ordering across Scryfall's 24h cache windows
    when prices flip by pennies.
  - URLs are constructed manually (no ?utm_source query string).
  - CN sort tries int() first; letter-suffix CNs fall through to lex.
  - Bucket precedence is fixed: fancy-foil > foil-only > nonfoil-only > unpriced
    > included. Each card lands in exactly one bucket.

Usage:
    printf '1 Nazgûl (LTR) 100\\n' | uv run python scripts/foil_price_diff.py
    uv run python scripts/foil_price_diff.py --file queries/missing-ltr-manapool-<ts>.txt
    uv run mm export moxfield 'set:ltr+related missing' | uv run python scripts/foil_price_diff.py
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

from magic_manager import parsers, scryfall, treatments  # noqa: E402


def _identifiers_from_stdin() -> list[tuple[str, str]]:
    text = sys.stdin.read()
    return _identifiers_from_text(text)


def _identifiers_from_text(text: str) -> list[tuple[str, str]]:
    result = parsers.parse_text(text)
    out: list[tuple[str, str]] = []
    for e in result.entries:
        if e.set and e.collector_number:
            out.append((e.set.lower(), e.collector_number))
    return out


def _identifiers_from_xlsx(path: Path) -> list[tuple[str, str]]:
    """Read set + collector_number columns from any sheet-based checklist.

    Handles both master-list XLSX (columns include ``set``, ``collector_number``,
    ``qty_normal``, ``qty_foil``) and missing-set XLSX (columns include ``set``,
    ``collector_number``, ``finish``). We only need the two identifier columns.
    """
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), data_only=True)
    # Pick the first non-underscore sheet as the data sheet.
    ws = None
    for name in wb.sheetnames:
        if not name.startswith("_"):
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return []
    header_lower = [str(h).strip().lower() if h is not None else "" for h in header]
    try:
        i_set = header_lower.index("set")
        i_cn = header_lower.index("collector_number")
    except ValueError:
        return []

    out: list[tuple[str, str]] = []
    for row in rows:
        if not row:
            continue
        set_val = row[i_set] if i_set < len(row) else None
        cn_val = row[i_cn] if i_cn < len(row) else None
        if set_val is None or cn_val is None:
            continue
        s = str(set_val).strip().lower()
        c = str(cn_val).strip()
        if s and c:
            out.append((s, c))
    return out


def _dedup(pairs: list[tuple[str, str]]) -> list[tuple[str, str]]:
    seen: set[tuple[str, str]] = set()
    out: list[tuple[str, str]] = []
    for p in pairs:
        if p in seen:
            continue
        seen.add(p)
        out.append(p)
    return out


def _cn_sort_key(cn: str) -> tuple[int, str]:
    """Numeric-first CN sort. ``9`` < ``10`` < ``10a``."""
    try:
        return (int(cn), "")
    except ValueError:
        # Split off leading numeric portion if possible for letter-suffix CNs.
        digits = ""
        for ch in cn:
            if ch.isdigit():
                digits += ch
            else:
                break
        try:
            n = int(digits) if digits else 999999
        except ValueError:
            n = 999999
        return (n, cn)


def _scryfall_url(set_code: str, cn: str) -> str:
    """Stable printing URL. No query string; no ?utm_source suffix."""
    return f"https://scryfall.com/card/{set_code.lower()}/{cn}"


def _fmt_usd(v: float) -> str:
    return f"${v:.2f}"


def _fmt_pct(p: float) -> str:
    """Signed percent with one decimal. ``+12.3%`` / ``-4.1%`` / ``+0.0%``."""
    return f"{p * 100:+.1f}%"


def _fmt_raw(v: float) -> str:
    sign = "+" if v >= 0 else "-"
    return f"{sign}${abs(v):.2f}"


def _price(card: dict, key: str) -> float | None:
    """Extract a nested Scryfall price. ``card["prices"][key]`` returns a
    string or None; we coerce to float or None."""
    prices = card.get("prices") or {}
    v = prices.get(key)
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _is_fancy_foil(card: dict) -> bool:
    return "ff" in treatments.compute_treatment(card).split("|")


def _parse_drop_expensive(s: str) -> tuple[float, float]:
    """Parse ``PCT:RAW`` (both floats). Both must be present."""
    try:
        pct_s, raw_s = s.split(":", 1)
        return (float(pct_s), float(raw_s))
    except (ValueError, AttributeError) as e:
        raise argparse.ArgumentTypeError(
            f"--drop-expensive expects PCT:RAW (e.g. 100:10), got {s!r}: {e}"
        ) from e


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Rank cards by simple-foil vs nonfoil price gap.",
    )
    ap.add_argument(
        "--file", type=Path, default=None,
        help="Path to a .txt (Moxfield-style) or .xlsx (checklist) file. "
             "If omitted, reads Moxfield-style lines from stdin.",
    )
    ap.add_argument(
        "--min-pct", type=float, default=None,
        help="Drop rows whose percent-diff is less than this (in percent, e.g. "
             "'--min-pct 0' hides any row where foil is cheaper than nonfoil).",
    )
    ap.add_argument(
        "--max-pct", type=float, default=None,
        help="Drop rows whose percent-diff is greater than this (in percent, "
             "e.g. '--max-pct 100' hides rows where foil is >2x nonfoil).",
    )
    ap.add_argument(
        "--min-raw", type=float, default=None,
        help="Drop rows whose dollar-diff is less than this (in USD, e.g. "
             "'--min-raw 0' hides rows where foil is cheaper than nonfoil).",
    )
    ap.add_argument(
        "--max-raw", type=float, default=None,
        help="Drop rows whose dollar-diff is greater than this (in USD, e.g. "
             "'--max-raw 10' hides rows where the foil upgrade costs more than $10).",
    )
    ap.add_argument(
        "--drop-expensive", type=_parse_drop_expensive, default=None,
        metavar="PCT:RAW",
        help="Compound filter: drop rows where BOTH percent-diff > PCT AND "
             "dollar-diff >= RAW. Use for 'cheap-upgrade' rankings that want "
             "to keep cheap-but-high-multiple rows (e.g. $0.30 → $2.36 foil = "
             "594%% is fine) but drop expensive-and-high-multiple ones "
             "(e.g. Gandalf $195 → $423 = 117%% + $228 is not). "
             "Example: --drop-expensive 100:10 keeps everything unless "
             "%%-diff exceeds 100 AND $-diff is $10 or more.",
    )
    args = ap.parse_args()

    if args.file is not None:
        if not args.file.exists():
            print(f"error: file not found: {args.file}", file=sys.stderr)
            return 2
        suffix = args.file.suffix.lower()
        if suffix == ".xlsx":
            pairs = _identifiers_from_xlsx(args.file)
        else:
            pairs = _identifiers_from_text(args.file.read_text(encoding="utf-8"))
    else:
        pairs = _identifiers_from_stdin()

    pairs = _dedup(pairs)

    if not pairs:
        print("Ranked 0 cards. Excluded: fancy-foil=0, foil-only=0, "
              "nonfoil-only=0, unpriced=0, unresolved=0, filtered=0.",
              file=sys.stderr)
        return 0

    identifiers = [{"set": s, "collector_number": c} for (s, c) in pairs]
    try:
        found, not_found = scryfall.collection(identifiers)
    except scryfall.ScryfallError as e:
        print(f"error: scryfall lookup failed: {e}", file=sys.stderr)
        return 2

    unresolved = len(not_found)
    fancy_foil = 0
    foil_only = 0
    nonfoil_only = 0
    unpriced = 0
    ranked: list[dict] = []

    for card in found:
        finishes = card.get("finishes") or []
        # Bucket precedence: fancy-foil first, so etched-foil-only prints
        # count as fancy-foil (their diff isn't computable AND they're the
        # premium tier the user isn't asking about).
        if _is_fancy_foil(card):
            fancy_foil += 1
            continue
        if finishes == ["foil"]:
            foil_only += 1
            continue
        if "foil" not in finishes:
            nonfoil_only += 1
            continue
        nonfoil = _price(card, "usd")
        foil = _price(card, "usd_foil")
        if nonfoil is None or foil is None or nonfoil == 0:
            unpriced += 1
            continue
        pct = (foil - nonfoil) / nonfoil
        raw = foil - nonfoil
        ranked.append({
            "name": card.get("name") or "",
            "set": (card.get("set") or "").lower(),
            "cn": card.get("collector_number") or "",
            "nonfoil": nonfoil,
            "foil": foil,
            "pct": pct,
            "pct_key": round(pct, 4),
            "raw": raw,
        })

    ranked.sort(key=lambda r: (r["pct_key"], r["name"], r["set"], _cn_sort_key(r["cn"])))

    # Post-sort filters. Applied after ranking so the table always renders in
    # sort order regardless of which rows the filters trim. All bounds are
    # inclusive on the "keep" side. Bounds are in percent for pct, USD for raw.
    filtered = 0
    if any(v is not None for v in (args.min_pct, args.max_pct,
                                    args.min_raw, args.max_raw,
                                    args.drop_expensive)):
        kept: list[dict] = []
        for r in ranked:
            pct_pct = r["pct"] * 100  # store as fraction; flags are in percent
            raw = r["raw"]
            if args.min_pct is not None and pct_pct < args.min_pct:
                filtered += 1; continue
            if args.max_pct is not None and pct_pct > args.max_pct:
                filtered += 1; continue
            if args.min_raw is not None and raw < args.min_raw:
                filtered += 1; continue
            if args.max_raw is not None and raw > args.max_raw:
                filtered += 1; continue
            if args.drop_expensive is not None:
                pct_thresh, raw_thresh = args.drop_expensive
                if pct_pct > pct_thresh and raw >= raw_thresh:
                    filtered += 1; continue
            kept.append(r)
        ranked = kept

    # Emit summary to stderr FIRST so it lands even if stdout is redirected.
    print(
        f"Ranked {len(ranked)} cards. Excluded: "
        f"fancy-foil={fancy_foil}, foil-only={foil_only}, "
        f"nonfoil-only={nonfoil_only}, unpriced={unpriced}, "
        f"unresolved={unresolved}, filtered={filtered}.",
        file=sys.stderr,
    )

    if not ranked:
        return 0

    # Markdown table. Left-align text, right-align numbers.
    print("| Card | Nonfoil | Foil | % diff | $ diff |")
    print("|---|---:|---:|---:|---:|")
    for r in ranked:
        link_text = f"{r['name']} ({r['set'].upper()}) {r['cn']}"
        # Escape pipe chars in card names (rare — e.g. split cards use ` // `).
        safe = link_text.replace("|", "\\|")
        url = _scryfall_url(r["set"], r["cn"])
        print(
            f"| [{safe}]({url}) | {_fmt_usd(r['nonfoil'])} | "
            f"{_fmt_usd(r['foil'])} | {_fmt_pct(r['pct'])} | "
            f"{_fmt_raw(r['raw'])} |"
        )

    return 0


if __name__ == "__main__":
    sys.exit(main())

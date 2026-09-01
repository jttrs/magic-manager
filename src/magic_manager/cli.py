"""``mm`` CLI: orchestrates set syncing, master-list generation, list import,
and exports.

Run via ``uv run mm …`` from the repo root.
"""

from __future__ import annotations

import hashlib
import json
import sys
from datetime import datetime
from pathlib import Path

import typer

from . import (
    db,
    decks as decks_mod,
    exports,
    front_cards as front_cards_mod,
    intake as intake_mod,
    inventory as inv_mod,
    mtgjson as mtgjson_mod,
    selectors as sel_mod,
    sets as sets_mod,
    util,
    wishlist as wishlist_mod,
)

CHECKLISTS_DIR = Path("checklists")
PROCESSED_DIR = CHECKLISTS_DIR / "processed"
# Backwards-compat alias — older code paths and migration helpers reference
# INPUT_DIR. Both names point at the same Path; the directory itself was
# renamed from ``input/`` → ``checklists/`` in V1.6.
INPUT_DIR = CHECKLISTS_DIR

# Exit codes used by master-list collision detection. The `generate-set-checklist`
# skill reads these to decide whether to prompt for ingest-or-force.
EXIT_UNPROCESSED_INTAKE = 3
# Ingest collision: file SHA matches a prior successful ingest_log row.
EXIT_DUPLICATE_INGEST = 4

VALID_RARITIES = ("mythic", "rare", "uncommon", "common", "bonus", "special")

app = typer.Typer(no_args_is_help=True, add_completion=False,
                  help="Local-first MTG collection / set / wishlist manager.")
set_app = typer.Typer(no_args_is_help=True, help="Set sync and master-list generation.")
inventory_app = typer.Typer(no_args_is_help=True, help="Cards I physically own (V2 fact table).")
wishlist_app = typer.Typer(no_args_is_help=True, help="Cards I want, organized by free-text category.")
deck_app = typer.Typer(no_args_is_help=True, help="Decks: compositions independent of ownership.")
query_app = typer.Typer(no_args_is_help=True,
                        help="Run V2 selector queries against the local DB (show/value/xlsx/url/top/total/multiples/stats).")

checklists_app = typer.Typer(no_args_is_help=True,
                             help="Inspect inventory checklists in checklists/.")
mtgjson_app = typer.Typer(no_args_is_help=True,
                          help="Read MTGJSON.com data (precon decks, set files, etc.).")
db_app = typer.Typer(no_args_is_help=True,
                     help="Manage the local SQLite DB: snapshots, restore, integrity.")
audit_app = typer.Typer(no_args_is_help=True,
                        help="Consistency checks + repair for the local DB.")

app.add_typer(set_app, name="set")
app.add_typer(inventory_app, name="inventory")
app.add_typer(wishlist_app, name="wishlist")
app.add_typer(deck_app, name="deck")
app.add_typer(query_app, name="query")
app.add_typer(checklists_app, name="checklists")
# Back-compat alias for muscle memory: ``mm input list`` still works.
app.add_typer(checklists_app, name="input")
app.add_typer(mtgjson_app, name="mtgjson")
app.add_typer(db_app, name="db")
app.add_typer(audit_app, name="audit")


def _slug(s: str) -> str:
    raw = "".join(c if c.isalnum() else "-" for c in s.lower())
    # Collapse runs of hyphens so "Final Fantasy: Through the Ages" → "final-fantasy-through-the-ages"
    # rather than "final-fantasy--through-the-ages".
    while "--" in raw:
        raw = raw.replace("--", "-")
    return raw.strip("-")


# ---------- set ----------

@set_app.command("list-related")
def set_list_related(name_or_code: str = typer.Argument(...)):
    """Show parent + sibling/child sets for the user's confirmation step."""
    try:
        r = sets_mod.resolve(name_or_code)
    except LookupError as e:
        typer.echo(f"error: {e}", err=True); raise typer.Exit(2)
    typer.echo(f"Parent: {r.code} ({r.name})")
    typer.echo("Related sets:")
    for s in r.related:
        marker = "  *" if s["code"] == r.code else "   "
        typer.echo(f"{marker} {s['code']:6}  {s.get('set_type','?'):14}  "
                   f"{s.get('card_count','?'):>5} cards  {s['name']}")


@set_app.command("sync")
def set_sync(
    name_or_code: str = typer.Argument(...),
    include_related: bool = typer.Option(False, "--include-related",
                                         help="Sync parent + every sibling/child set."),
    only: list[str] = typer.Option(None, "--only", help="Restrict to these set codes (comma-separated)."),
):
    """Resolve and sync set(s) into the local cards table."""
    try:
        r = sets_mod.resolve(name_or_code)
    except LookupError as e:
        typer.echo(f"error: {e}", err=True); raise typer.Exit(2)

    codes = r.all_codes if include_related else [r.code]
    if only:
        wanted = {c.strip().lower() for raw in only for c in raw.split(",")}
        codes = [c for c in codes if c in wanted]
        if not codes:
            typer.echo("error: --only filtered out all sets", err=True); raise typer.Exit(2)

    typer.echo(f"Syncing {len(codes)} set(s): {' '.join(codes)}")
    n = sets_mod.sync(codes)
    typer.echo(f"  → {n} cards upserted")


@set_app.command("is-synced")
def set_is_synced(
    name_or_code: str = typer.Argument(...),
):
    """Report how many cards are synced per family set-code. Exits non-zero if
    the family has zero cards locally.

    Answers "do I need to `mm set sync` before importing a precon / running a
    query?" without a hand-written ``SELECT COUNT(*)``. Resolves the +related
    family and prints one line per code with its local card count.
    """
    try:
        r = sets_mod.resolve(name_or_code)
    except LookupError as e:
        typer.echo(f"error: {e}", err=True); raise typer.Exit(2)

    codes = r.all_codes
    placeholders = ",".join("?" for _ in codes)
    with db.connect() as conn:
        rows = conn.execute(
            f"SELECT set_code, COUNT(*) AS n FROM cards "
            f"WHERE set_code IN ({placeholders}) GROUP BY set_code",
            [c.lower() for c in codes],
        ).fetchall()
    counts = {row["set_code"]: row["n"] for row in rows}
    total = sum(counts.values())
    typer.echo(f"{r.name} (anchor {r.code}) — {total} cards across {len(codes)} code(s):")
    for c in codes:
        n = counts.get(c.lower(), 0)
        mark = " " if n else " (not synced)"
        typer.echo(f"  {c:8} {n:>5}{mark}")
    if total == 0:
        raise typer.Exit(1)


def _slice_suffix(*, only_codes: list[str], rarities: list[str]) -> str:
    """Build the filename slice suffix from optional set-code and rarity slices.

    No slice → ``""`` (empty — the unsliced default; filename has no slice token).
    Codes only → ``codes-joined-by-plus``.
    Rarities only → ``rarities-joined-by-plus``.
    Both → ``codes-rarities``.
    """
    parts: list[str] = []
    if only_codes:
        parts.append("+".join(only_codes))
    if rarities:
        parts.append("+".join(rarities))
    return "-".join(parts)


def _intake_path(slug: str, slice_suffix: str = "", ext: str = "xlsx", mode: str = "add") -> Path:
    # Filename shape: ``<slug>[-<slice>]-<mode>-checklist.<ext>``.
    # ``mode`` is ``add`` (default; blank-qty checklist for additive ingest)
    # or ``modify`` (prefilled-qty checklist for replace ingest). The mode
    # token in the filename is critical because cmux/Finder show filenames
    # without exposing _meta — surface intent on disk. Slice suffix encodes
    # the optional ``--only`` and ``--rarity`` filters; pre-V1.6 used an
    # explicit ``master`` token instead of an empty slice.
    middle = f"-{slice_suffix}" if slice_suffix else ""
    return CHECKLISTS_DIR / f"{slug}{middle}-{mode}-checklist.{ext}"


def _processed_path(slug: str, slice_suffix: str = "",
                    when: datetime | None = None, ext: str = "xlsx") -> Path:
    when = when or datetime.now()
    middle = f"-{slice_suffix}" if slice_suffix else ""
    return PROCESSED_DIR / f"{slug}{middle}-checklist-{when:%Y-%m-%d-%H%M%S}.{ext}"


def _file_sha256(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _split_csv(values: list[str] | None) -> list[str]:
    """Flatten a list of strings (each potentially comma-separated) into a
    deduplicated, lowercased list. Lets ``--include token,memorabilia`` and
    ``--include token --include memorabilia`` mean the same thing."""
    out: list[str] = []
    seen: set[str] = set()
    for raw in values or []:
        for part in raw.split(","):
            v = part.strip().lower()
            if v and v not in seen:
                seen.add(v)
                out.append(v)
    return out


def _resolve_codes(name_or_code: str, *, include_kinds: list[str], only: list[str]
                   ) -> tuple[sets_mod.ResolvedSet, list[str]]:
    r = sets_mod.resolve(name_or_code)
    only_codes = _split_csv(only)
    kinds = _split_csv(include_kinds)
    if only_codes:
        # Honor explicit --only codes verbatim, even if they aren't in the
        # parent's Scryfall related-set graph. UB families like SPM ship with
        # a separate masterpiece root (mar/Marvel Universe) whose
        # parent_set_code is null — the user knows it belongs in the family
        # checklist even though Scryfall doesn't link it.
        codes = list(only_codes)
    else:
        codes = r.filtered_codes(include_kinds=kinds)
    if not codes:
        raise typer.BadParameter("set selection produced 0 codes")
    return r, codes


@set_app.command("master-list")
def set_master_list(
    name_or_code: str = typer.Argument(...),
    only: list[str] = typer.Option(
        None, "--only",
        help="Hard subset of codes (comma-separated). Bypasses the default set-type filter.",
    ),
    include: list[str] = typer.Option(
        None, "--include",
        help="Opt extra set_types into the family beyond the default "
             "(expansion/commander/masterpiece/promo/eternal). E.g. --include token,memorabilia.",
    ),
    rarity: list[str] = typer.Option(
        None, "--rarity",
        help="Slice by rarity (repeatable, comma-OK). Values: "
             "mythic|rare|uncommon|common|bonus|special. Output filename gets a "
             "rarity suffix and ingest of this file only touches rows of the "
             "given rarity.",
    ),
    out: Path = typer.Option(
        None, "--out",
        help="Override output path. When set, collision detection is skipped.",
    ),
    force: bool = typer.Option(
        False, "--force",
        help="Overwrite an existing inventory checklist without prompting.",
    ),
    fmt: str = typer.Option(
        "xlsx", "--format",
        help="Output format: 'xlsx' (default; clickable Scryfall hyperlinks) "
             "or 'md' (markdown checklist editable in any text editor).",
    ),
    include_variants: bool = typer.Option(
        False, "--include-variants",
        help="Include prerelease, store-stamped, japanshowcase, serialized, "
             "and white/yellow-bordered variants. Off by default — these are "
             "filtered out of the inventory checklist AND the seeded set:<anchor> list "
             "so set-missing math doesn't count them.",
    ),
    mode: str = typer.Option(
        "add", "--mode",
        help="'add' (default): blank checklist for additive ingest — qty>0 "
             "cells sum into existing inventory; safe (cannot zero rows). Use "
             "for new acquisitions (booster packs, precons, trade-ins). "
             "'modify': prefilled checklist for replace ingest — in-partition "
             "cells overwrite DB qty AND missing in-partition rows zero out. "
             "Use to correct existing records (sold cards, miscounts, audits).",
    ),
):
    """Build the inventory checklist for a release family or a slice of it.

    The default family is the anchor set + every related set whose set_type
    is in {expansion, commander, masterpiece, promo, eternal}. Tokens,
    memorabilia (art series, scene boxes), and other set_types are excluded
    by default; opt them in with ``--include token,memorabilia``.

    The checklist lands at ``input/<slug>-<slice>.<ext>`` where ``<slice>``
    encodes the optional ``--only`` and ``--rarity`` filters (or ``master``
    if neither is given) and ``<ext>`` is ``xlsx`` (default) or ``md``.
    There can be at most one active inventory checklist per slice + format
    at a time; if one exists, the command refuses with exit
    ``EXIT_UNPROCESSED_INTAKE`` (3). Either ingest that file first via
    ``mm set ingest`` or pass ``--force``.
    """
    fmt = fmt.lower()
    if fmt not in ("xlsx", "md"):
        typer.echo(f"error: --format must be 'xlsx' or 'md', got {fmt!r}", err=True)
        raise typer.Exit(2)
    mode = mode.lower()
    if mode not in ("add", "modify"):
        typer.echo(f"error: --mode must be 'add' or 'modify', got {mode!r}", err=True)
        raise typer.Exit(2)
    try:
        r, codes = _resolve_codes(name_or_code, include_kinds=list(include or []), only=list(only or []))
    except (LookupError, typer.BadParameter) as e:
        typer.echo(f"error: {e}", err=True); raise typer.Exit(2)

    only_codes = _split_csv(only)
    rarities = _split_csv(rarity)
    bad = [rr for rr in rarities if rr not in VALID_RARITIES]
    if bad:
        typer.echo(f"error: invalid --rarity value(s): {bad}; expected one of {VALID_RARITIES}", err=True)
        raise typer.Exit(2)

    slug = _slug(r.name)
    slice_suffix = _slice_suffix(only_codes=only_codes, rarities=rarities)
    out_path = out or _intake_path(slug, slice_suffix, ext=fmt, mode=mode)

    # Collision detection: only when the user is using the default path.
    if out is None and out_path.exists() and not force:
        typer.echo(f"refusing to overwrite existing inventory checklist: {out_path}", err=True)
        # Snapshot the current inventory rows that fall in this set's family.
        inv_rows = [r for r in inv_mod.inventory_show() if r.set_code in codes]
        if inv_rows:
            total_qty = sum(r.quantity for r in inv_rows)
            total_value = sum((r.line_value or 0.0) for r in inv_rows)
            typer.echo(
                f"  {len(inv_rows)} (card,finish) row(s) currently owned in this family, "
                f"qty {total_qty}, value ${total_value:.2f}",
                err=True,
            )
            top = sorted(inv_rows, key=lambda x: (x.line_value or 0.0), reverse=True)[:5]
            if top:
                typer.echo("  top by value:", err=True)
                for row in top:
                    price = f"${row.unit_price:.2f}" if row.unit_price is not None else "—"
                    typer.echo(
                        f"    {row.quantity}x {row.display_name} ({row.set_code.upper()}) "
                        f"{row.collector_number} [{row.finish}] @ {price}",
                        err=True,
                    )
        else:
            typer.echo(f"  no inventory rows yet in family {codes}", err=True)
        typer.echo("", err=True)
        typer.echo("To proceed, either:", err=True)
        typer.echo(f"  - Finish editing the existing XLSX, then: mm set ingest {name_or_code!r}", err=True)
        typer.echo(f"  - Discard partial edits and regenerate: mm set master-list {name_or_code!r} --force", err=True)
        raise typer.Exit(EXIT_UNPROCESSED_INTAKE)

    typer.echo(f"Syncing {len(codes)} set(s): {' '.join(codes)}")
    n_synced = sets_mod.sync(codes)
    typer.echo(f"  → {n_synced} cards upserted")

    target_result = sets_mod.register_set_target(
        r.code, codes, include_variants=include_variants, rarity_filter=rarities or None,
    )
    typer.echo(f"Registered set_target {r.code!r} ({target_result['action']}, "
               f"{len(target_result['related_codes'])} related code(s))")

    if force and out_path.exists():
        typer.echo(f"  ! --force: overwriting {out_path}", err=True)

    writer = (
        sets_mod.write_master_list_md if fmt == "md"
        else sets_mod.write_master_list_xlsx
    )
    n_rows, prefilled = writer(
        codes, out_path,
        # Tokens and memorabilia are governed by the family filter, not by a
        # second flag. If the user --included them they're in `codes` already.
        include_tokens=True,
        # mode='modify' prefills qty cells from current inventory (intended
        # for replace-style ingest). mode='add' leaves them blank (intended
        # for additive ingest of new acquisitions).
        prepopulate_from_inventory=(mode == "modify"),
        rarity_filter=rarities or None,
        anchor_code=r.code,
        slug=slug,
        include_variants=include_variants,
        mode=mode,
    )
    typer.echo(f"Wrote {n_rows} rows to {out_path} (mode={mode})")
    if prefilled:
        typer.echo(f"  → {prefilled} qty cell(s) pre-filled from inventory")
    typer.echo()
    typer.echo("Next steps:")
    if mode == "add":
        verb = "fill in qty_normal / qty_foil for the cards you're ADDING (cells start blank)"
    else:
        verb = "edit qty_normal / qty_foil to MODIFY existing inventory (prefilled values shown)"
    if fmt == "md":
        typer.echo(f"  1. Open {out_path} in any text editor and edit `[N:k F:k]` quantities ({verb}).")
    else:
        typer.echo(f"  1. Open {out_path} in Excel/Numbers — {verb}.")
    typer.echo(f"  2. When done: mm set ingest {name_or_code!r}  (auto-detects mode={mode} from _meta)")


@set_app.command("jumpstart-list")
def set_jumpstart_list(
    set_code: str = typer.Argument(
        ...,
        help="Jumpstart set code: tle, j25, j22, jmp, etc. The set itself "
             "must publish Jumpstart variants in MTGJSON (run "
             "`mm mtgjson decks --set <CODE>` to confirm).",
    ),
    out: Path = typer.Option(
        None, "--out",
        help="Override output path. When set, collision detection is skipped.",
    ),
    force: bool = typer.Option(
        False, "--force",
        help="Overwrite an existing Jumpstart checklist without prompting.",
    ),
    fmt: str = typer.Option(
        "xlsx", "--format",
        help="Output format: 'xlsx' (default) or 'md' (markdown).",
    ),
):
    """Build a pack-level checklist of every Jumpstart variant for a set.

    One row per sealed-pack variant (e.g. ~66 rows for TLE). Fill
    ``keep_qty`` (0 or 1 — copies kept *constructed*: one ``pack:*`` recipe
    is created and one physical copy is auto-composed) and
    ``deconstructed_qty`` (copies torn into free cards; no pledge). The two
    sum to total packs opened and that total determines how many cards land in
    inventory.

    Complements ``mm set master-list``: that one is per-card across the whole
    family, this one is per-pack inside one Jumpstart set. Use this when
    you've opened sealed product and want to ingest whole packs at once.
    """
    fmt = fmt.lower()
    if fmt not in ("xlsx", "md"):
        typer.echo(f"error: --format must be 'xlsx' or 'md', got {fmt!r}", err=True)
        raise typer.Exit(2)

    code = set_code.lower()
    # The slug embeds 'jumpstart' so the filename is self-describing on disk
    # (Finder/cmux don't show _meta) and never collides with a master-list
    # checklist for the same code.
    slug = f"{code}-jumpstart"
    out_path = out or (CHECKLISTS_DIR / f"{slug}-checklist.{fmt}")

    if out is None and out_path.exists() and not force:
        typer.echo(f"refusing to overwrite existing Jumpstart checklist: {out_path}", err=True)
        typer.echo("", err=True)
        typer.echo("To proceed, either:", err=True)
        typer.echo(f"  - Finish editing the existing file, then: mm set ingest --path {out_path}", err=True)
        typer.echo(f"  - Discard partial edits and regenerate: mm set jumpstart-list {set_code} --force", err=True)
        raise typer.Exit(EXIT_UNPROCESSED_INTAKE)

    # Sync the set's cards so usd_total roll-ups have prices to pull from
    # AND the eventual ingest can resolve every scryfall_id locally. Jumpstart
    # variants for a set like TLE include prints from BOTH the set itself AND
    # its parent expansion (TLA), so sync the family.
    try:
        r, codes = _resolve_codes(set_code, include_kinds=[], only=[])
    except (LookupError, typer.BadParameter) as e:
        typer.echo(f"error: {e}", err=True)
        raise typer.Exit(2)
    typer.echo(f"Syncing {len(codes)} set(s) (Jumpstart contents may span the family): {' '.join(codes)}")
    n_synced = sets_mod.sync(codes)
    typer.echo(f"  → {n_synced} cards upserted")

    if force and out_path.exists():
        typer.echo(f"  ! --force: overwriting {out_path}", err=True)

    writer = (
        sets_mod.write_jumpstart_list_md if fmt == "md"
        else sets_mod.write_jumpstart_list_xlsx
    )
    try:
        n_rows = writer(code, out_path, slug=slug)
    except ValueError as e:
        typer.echo(f"error: {e}", err=True)
        raise typer.Exit(2)
    typer.echo(f"Wrote {n_rows} Jumpstart variant rows to {out_path}")
    typer.echo()
    typer.echo("Next steps:")
    if fmt == "md":
        typer.echo(f"  1. Open {out_path} in any text editor and edit the `[K:k D:d]` brackets (K=0/1 kept constructed, D=copies to deconstruct).")
    else:
        typer.echo(f"  1. Open {out_path} in Excel/Numbers — fill keep_qty (0 or 1) and deconstructed_qty per pack.")
    typer.echo(f"  2. When done: mm set ingest --path {out_path}")


def _jumpstart_pack_rows(
    code: str, matched: dict, *, include_front: bool = True
) -> tuple[list, int]:
    """Materialize one Jumpstart pack's cards into export-ready rows.

    Walks the matched MTGJSON variant's commander/main/side boards, resolves
    each scryfall_id against the local ``cards`` table, and (optionally)
    appends the pack's front/title card from the quarantined ``front_cards``
    table. Returns ``(rows, n_skipped)`` where ``n_skipped`` counts gameplay
    printings absent from ``cards`` (caller decides whether to warn).

    Shared by ``jumpstart-pack`` (one pack) and ``query missing-jumpstart``
    (every un-owned pack in a set). Assumes the family is already synced.
    """
    deck_data = mtgjson_mod.deck(matched["fileName"])

    entries: list[tuple[str, int, bool]] = []  # (scryfall_id, count, is_foil)
    for board_key in ("commander", "mainBoard", "sideBoard"):
        for entry in deck_data.get(board_key) or []:
            sid = (entry.get("identifiers") or {}).get("scryfallId")
            if not sid:
                continue
            count = int(entry.get("count", 1) or 1)
            entries.append((sid, count, bool(entry.get("isFoil"))))

    rows: list[sel_mod.MaterializedRow] = []
    n_skipped = 0
    if entries:
        with db.connect() as conn:
            placeholders = ",".join("?" for _ in entries)
            card_rows = {
                cr["scryfall_id"]: cr
                for cr in conn.execute(
                    f"SELECT {sel_mod._CARD_COLS} FROM cards c "
                    f"WHERE c.scryfall_id IN ({placeholders})",
                    [e[0] for e in entries],
                ).fetchall()
            }
        for sid, count, is_foil in entries:
            cr = card_rows.get(sid)
            if cr is None:
                n_skipped += 1
                continue
            rows.append(sel_mod.MaterializedRow(
                scryfall_id=sid,
                quantity=count,
                finish="foil" if is_foil else "nonfoil",
                card=sel_mod._card_dict(cr),
            ))

    if include_front:
        fc = front_cards_mod.front_card_for_theme(code, matched.get("name") or "")
        if fc is not None:
            rows.append(front_cards_mod.front_card_row(fc))

    return rows, n_skipped


@set_app.command("jumpstart-pack")
def set_jumpstart_pack(
    set_code: str = typer.Argument(..., help="Jumpstart set code, e.g. msh"),
    theme: str = typer.Argument(..., help="Pack theme or MTGJSON fileName, e.g. Scarlet or Scarlet_MSH"),
    fmt: str = typer.Option("manapool", "--format", help="manapool|moxfield|tcgplayer|plain"),
    missing: bool = typer.Option(False, "--missing", help="Only cards not in inventory"),
    out: Path = typer.Option(None, "--out", help="Write to file instead of stdout"),
):
    """Emit a paste-ready export block for one Jumpstart pack — every
    gameplay single plus the pack's front/title card (from the quarantined
    ``front_cards`` table, e.g. FMSC for MSH), which never appears anywhere
    else in the app.
    """
    code = set_code.lower()

    # Sync the family so gameplay scryfall_ids resolve locally, mirroring
    # `jumpstart-list`'s sync sequence. Front-card sync is separate (and
    # NEVER touches the `cards` table) — best-effort.
    try:
        r, codes = _resolve_codes(set_code, include_kinds=[], only=[])
    except (LookupError, typer.BadParameter) as e:
        typer.echo(f"error: {e}", err=True)
        raise typer.Exit(2)
    sets_mod.sync(codes)
    front_cards_mod.sync_front_cards(code)

    variants = mtgjson_mod.jumpstart_variants(code)
    if not variants:
        typer.echo(f"error: no Jumpstart variants found for set {set_code!r}", err=True)
        raise typer.Exit(2)

    normalized_theme = front_cards_mod.normalize_theme(theme)
    exact_file = [v for v in variants if v.get("fileName") == theme]
    matches = exact_file or [v for v in variants if (v.get("name") or "").lower() == theme.lower()]
    if not matches:
        matches = [v for v in variants
                   if front_cards_mod.normalize_theme(v.get("name") or "") == normalized_theme]

    if not matches:
        available = ", ".join(sorted(v.get("name") or v.get("fileName") or "" for v in variants))
        typer.echo(f"error: no Jumpstart variant matches theme {theme!r} in {set_code!r}", err=True)
        typer.echo(f"available themes: {available}", err=True)
        raise typer.Exit(2)
    if len(matches) > 1:
        ambiguous = ", ".join(sorted(v.get("name") or v.get("fileName") or "" for v in matches))
        typer.echo(f"error: theme {theme!r} matches multiple variants: {ambiguous}", err=True)
        raise typer.Exit(2)

    matched = matches[0]
    rows, n_skipped = _jumpstart_pack_rows(code, matched)
    if n_skipped:
        typer.echo(f"warning: {n_skipped} card(s) not found locally, skipped", err=True)

    if missing:
        with db.connect() as conn:
            owned = {
                (row["scryfall_id"], row["finish"])
                for row in conn.execute("SELECT scryfall_id, finish FROM inventory").fetchall()
            }
        rows = [r for r in rows if (r.scryfall_id, r.finish) not in owned]

    try:
        text = exports.build(fmt, rows)
    except ValueError as e:
        typer.echo(f"error: {e}", err=True)
        raise typer.Exit(2)

    if out:
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(text, encoding="utf-8")
        typer.echo(f"wrote {out}", err=True)
    else:
        typer.echo(text, nl=False)


@set_app.command("precon-list")
def set_precon_list(
    only_type: str = typer.Option(
        None, "--type",
        help="Narrow to one exact product type (e.g. 'Commander Deck', "
             "'Box Set', 'Duel Deck'). Default: the modern-constructed precon "
             "types (Commander, Box Set, Planeswalker, Duel, Starter Kit, …).",
    ),
    all_physical: bool = typer.Option(
        False, "--all-physical",
        help="Widen scope to EVERY physical product MTGJSON lists (~1500 rows, "
             "incl. old Theme Decks / Intro Packs / Welcome Decks). Default is "
             "the modern-constructed subset. Ignored when --type is given.",
    ),
    include_collector: bool = typer.Option(
        False, "--include-collector",
        help="Include the '… Collector's Edition' twins (excluded by default — "
             "they're a premium variant the collection doesn't track).",
    ),
    mode: str = typer.Option(
        "add", "--mode",
        help="'add' (default): blank constructed_qty/deconstructed_qty cells; "
             "ingest ADDS the counts you enter (each copy becomes a deck row). "
             "'modify': both columns prefilled from your current deck collection "
             "(counts derived from the decks table); ingest applies the signed "
             "delta. Use 'modify' to see what you already have and avoid "
             "double-adding.",
    ),
    sync_all: bool = typer.Option(
        False, "--sync-all",
        help="Sync every set referenced by the catalog from Scryfall before "
             "rollup so ALL usd_total values populate. Slow (pulls ~180 sets) "
             "and grows the local cards table with sets you may not own. "
             "Default: best-effort — totals blank for un-synced sets.",
    ),
    out: Path = typer.Option(
        None, "--out",
        help="Override output path. When set, collision detection is skipped.",
    ),
    force: bool = typer.Option(
        False, "--force",
        help="Overwrite an existing precon catalog without prompting.",
    ),
    fmt: str = typer.Option(
        "xlsx", "--format",
        help="Output format: 'xlsx' (default) or 'md' (markdown).",
    ),
):
    """Build a global catalog of preconstructed products across ALL sets.

    There are only a handful of precons per set, so this is one master list you
    populate once — not a per-set file. One row per product (Commander Deck,
    Box Set, Planeswalker Deck, …), carrying its set, deck name, product type,
    release date, commander(s), card count, and best-effort market value.

    Track precon decks AS UNITS via two fill columns: ``constructed_qty`` (built
    copies — each creates a deck + adds its cards to inventory) and
    ``deconstructed_qty`` (copies torn down for parts — a deck row is recorded,
    cards go loose). In ``--mode add`` (default) the cells are blank and ingest
    ADDS the counts; in ``--mode modify`` they're prefilled from your current
    deck collection (counts derived from the decks table) and ingest applies the
    signed delta — so you SEE what you already have and don't double-add.

    Scope defaults to the modern-constructed precon types; ``--type`` narrows to
    one, ``--all-physical`` opens it to everything. Collector's Edition variants
    are excluded unless ``--include-collector``. Digital (``MTGO …``), Jumpstart
    (own command), and Secret Lair (own bulk-add flow) products are never
    listed.

    Generation does NOT sync by default (an all-sets catalog can't sync all of
    Magic), so ``usd_total`` is blank for sets not yet in the local cards table;
    ingest of a filled row self-syncs that precon's sets, so values fill in over
    time. Pass ``--sync-all`` to sync every referenced set up front and populate
    all totals in one go (slower).
    """
    fmt = fmt.lower()
    if fmt not in ("xlsx", "md"):
        typer.echo(f"error: --format must be 'xlsx' or 'md', got {fmt!r}", err=True)
        raise typer.Exit(2)
    mode = mode.lower()
    if mode not in ("add", "modify"):
        typer.echo(f"error: --mode must be 'add' or 'modify', got {mode!r}", err=True)
        raise typer.Exit(2)

    # Global catalog — one file, not per-set. The slug embeds 'precons' + the
    # mode so it's self-describing on disk and add/modify files can coexist
    # (mirrors master-list's <slug>-<mode>-checklist convention).
    slug = "precons"
    out_path = out or (CHECKLISTS_DIR / f"{slug}-{mode}-checklist.{fmt}")

    if out is None and out_path.exists() and not force:
        typer.echo(f"refusing to overwrite existing precon catalog: {out_path}", err=True)
        typer.echo("", err=True)
        typer.echo("To proceed, either:", err=True)
        typer.echo(f"  - Finish editing the existing file, then: mm set ingest --path {out_path}", err=True)
        typer.echo(f"  - Discard partial edits and regenerate: mm set precon-list --mode {mode} --force", err=True)
        raise typer.Exit(EXIT_UNPROCESSED_INTAKE)

    if force and out_path.exists():
        typer.echo(f"  ! --force: overwriting {out_path}", err=True)

    typer.echo("Cataloging precons across all sets (reading MTGJSON per-deck files; first run may take a moment)…")
    if mode == "modify":
        typer.echo("  --mode modify: constructed_qty/deconstructed_qty prefilled from the precon ledger.")
    if sync_all:
        typer.echo("  --sync-all: will sync every referenced set from Scryfall first (this is the slow part)…")
    writer = (
        sets_mod.write_precon_list_md if fmt == "md"
        else sets_mod.write_precon_list_xlsx
    )
    try:
        n_rows = writer(out_path, slug=slug, mode=mode, only_type=only_type,
                        all_physical=all_physical,
                        include_collector=include_collector,
                        sync_all=sync_all,
                        progress=lambda m: typer.echo(f"  {m}"))
    except ValueError as e:
        typer.echo(f"error: {e}", err=True)
        raise typer.Exit(2)
    typer.echo(f"Wrote {n_rows} precon product rows to {out_path} (mode={mode})")
    typer.echo()
    typer.echo("Next steps:")
    if fmt == "md":
        typer.echo(f"  1. Open {out_path} in any text editor and edit the `[C:c D:d]` brackets (C=constructed copies, D=deconstructed copies).")
    else:
        typer.echo(f"  1. Open {out_path} in Excel/Numbers — fill constructed_qty and deconstructed_qty per deck.")
    typer.echo(f"  2. When done: mm set ingest --path {out_path}")


def _ingest_deck_checklist(src: Path, *, kind: str, sha: str, force: bool,
                           json_out: bool) -> None:
    """Apply a deck checklist (kind ∈ {'jumpstart', 'precon'}). Branches off
    ``mm set ingest`` when the file's _meta declares one of those kinds.

    A precon is the base concept; Jumpstart is a species of it — so both share
    this one consumer, parameterized by ``kind`` for the log label and the
    human-facing noun. Logs an ingest_log row with label ``<kind>:<setcode>``,
    archives the file under processed/. Mirrors the duplicate-detection +
    archive shape of the inventory ingest path so all checklists look the same
    on disk.
    """
    noun = "Jumpstart" if kind == "jumpstart" else "Precon"
    with db.connect() as conn:
        prior = db.find_ingest_log_by_hash(conn, sha)
    prior_success = next((p for p in prior if p["status"] == "success"), None)
    if prior_success and not force:
        msg = (
            f"this file's SHA-256 matches a previous successful ingest "
            f"(log id {prior_success['id']}, "
            f"mode {prior_success['mode']}, at {prior_success['at']})."
        )
        if json_out:
            json.dump({
                "status": "duplicate",
                "file": str(src),
                "sha256": sha,
                "prior_log": prior_success,
                "message": msg,
            }, sys.stdout, indent=2)
            sys.stdout.write("\n")
        else:
            typer.echo(f"refusing to re-ingest: {msg}", err=True)
            typer.echo("  pass --force to proceed.", err=True)
        raise typer.Exit(EXIT_DUPLICATE_INGEST)

    error: str | None = None
    summary: dict | None = None
    try:
        summary = sets_mod.ingest_deck_checklist_from_path(src, kind=kind)
    except Exception as e:
        error = repr(e)

    archived: Path | None = None
    if summary is not None:
        # Deck checklists use a flat naming pattern: <setcode>-<kind>-checklist.<ext>
        # Archive with timestamp under processed/ so re-runs don't collide.
        ext = src.suffix.lstrip(".") or "xlsx"
        when = datetime.now()
        archived = PROCESSED_DIR / f"{src.stem}-{when:%Y-%m-%d-%H%M%S}.{ext}"
        archived.parent.mkdir(parents=True, exist_ok=True)
        src.rename(archived)

    meta = sets_mod.read_master_list_meta(archived or src) or {}
    # The global precon catalog carries no anchor_code/set_codes (it spans every
    # set); its rows each name their own set. Fall back to a bare kind label.
    set_code = meta.get("anchor_code") or meta.get("set_codes") or ""
    log_label = f"{kind}:{set_code}" if set_code else kind
    # Both summaries expose ``constructed``; precon uses ``deconstructed`` where
    # jumpstart uses ``loose_copies`` for the torn-down count.
    rows_added = (summary or {}).get("constructed", 0)
    rows_updated = ((summary or {}).get("loose_copies")
                    or (summary or {}).get("deconstructed", 0))
    with db.connect() as conn:
        db.record_ingest_log(
            conn,
            label=log_label,
            # Deck-checklist ingest is semantically additive (only adds decks +
            # inventory; never zeroes). The ``label`` already encodes the kind
            # so this row is distinguishable from inventory ingests sharing the
            # same mode.
            mode="additive",
            source_path=str(src),
            archived_path=str(archived) if archived else None,
            file_sha256=sha,
            rows_added=rows_added,
            rows_updated=rows_updated,
            rows_zeroed=0,
            status="success" if error is None else "failed",
            error=error,
        )

    if json_out:
        out = {
            "status": "success" if error is None else "failed",
            "file": str(src),
            "archived_path": str(archived) if archived else None,
            "kind": kind,
            "set_code": set_code,
            "sha256": sha,
            "summary": summary,
            "error": error,
        }
        json.dump(out, sys.stdout, indent=2, default=str)
        sys.stdout.write("\n")
        if error is not None:
            raise typer.Exit(2)
        return

    if error is not None:
        typer.echo(f"error: {kind} ingest failed: {error}", err=True)
        raise typer.Exit(2)

    scope_seg = f" ({set_code})" if set_code else ""
    if kind == "precon":
        # Precon summary: signed transaction against the decks table (counts
        # are derived; no ledger). built/torn_down/pooled + count_before/after
        # (3-tuples: built, deconstructed, pool) per row.
        typer.echo(
            f"{noun}: {summary['rows_acted']}/{summary['rows_total']} rows acted on, "
            f"{summary['built']} built, "
            f"{summary['deconstructed']} deconstructed, "
            f"{summary['pool']} pooled, "
            f"{summary['inv_qty_total']} card-qty added to inventory."
        )
        for row in summary["per_row"]:
            if row["error"]:
                typer.echo(f"  ! {row['file_name']}: {row['error']}", err=True)
                continue
            bc, bd, bp = row["count_before"]
            ac, ad, ap = row["count_after"]
            typer.echo(
                f"  {row['file_name']} ({row['label']}): "
                f"built {bc}→{ac}, deconstructed {bd}→{ad}, pool {bp}→{ap}"
                + (f"  [+{row['built']} built]" if row["built"] else "")
                + (f"  [+{row['torn_down']} torn down]" if row["torn_down"] else "")
                + (f"  [+{row['pooled']} pooled]" if row["pooled"] else "")
            )
            if row.get("warning"):
                typer.echo(f"    note: {row['warning']}", err=True)
            if row["missing_sids"]:
                typer.echo(
                    f"    warning: {len(row['missing_sids'])} entries had no scryfallId",
                    err=True,
                )
    else:
        typer.echo(
            f"{noun}{scope_seg}: "
            f"{summary['rows_acted']}/{summary['rows_total']} rows acted on, "
            f"{summary['constructed']} constructed, "
            f"{summary['loose_copies']} loose copies, "
            f"{summary['inv_qty_total']} card-qty added to inventory."
        )
        for row in summary["per_row"]:
            if row["error"]:
                typer.echo(f"  ! {row['file_name']}: {row['error']}", err=True)
                continue
            if row["keep_qty"] == 1:
                if row["deconstructed_qty"] > 0:
                    bits = f"constructed 1 + {row['deconstructed_qty']} loose → {row['slug']}"
                else:
                    bits = f"constructed 1 → {row['slug']}"
            else:
                bits = f"deconstructed {row['deconstructed_qty']} → loose inventory"
            typer.echo(f"  {row['file_name']} ({row['label']}): {bits}")
            if row["missing_sids"]:
                typer.echo(
                    f"    warning: {len(row['missing_sids'])} entries had no scryfallId",
                    err=True,
                )
    for w in summary["warnings"]:
        typer.echo(f"  warning: {w}", err=True)
    if archived:
        typer.echo(f"Archived {noun} checklist → {archived}")


@set_app.command("ingest")
def set_ingest(
    name_or_code: str = typer.Argument(
        None,
        help="Set name or code. Optional when --path is given AND the file has a _meta sheet.",
    ),
    path: Path = typer.Option(
        None, "--path",
        help="Override path to the inventory checklist. Default: input/<slug>-master.xlsx.",
    ),
    mode: str = typer.Option(
        None, "--mode",
        help="OPTIONAL OVERRIDE. 'replace' (each in-partition row is SET to its "
             "cell value — a signed change vs current) or 'additive' (only qty>0 "
             "cells add to existing). Default: auto-detect from the checklist's "
             "_meta.mode ('modify' → replace, 'add' → additive). Pass --mode "
             "explicitly only to override the file's declared intent (logs a "
             "stderr warning) OR for legacy files with no _meta.mode.",
    ),
    zero_untouched: bool = typer.Option(
        None, "--zero-untouched/--no-zero-untouched",
        help="Replace/modify ingest ONLY. When set, in-partition rows absent "
             "from the file are also zeroed (full-audit: file is authoritative). "
             "Default: not set — you're asked interactively, or with --json it "
             "defaults to NOT zeroing (safe; the file is a delta, not a wipe).",
    ),
    force: bool = typer.Option(
        False, "--force",
        help="Re-ingest even if this exact file (by SHA-256) succeeded previously.",
    ),
    json_out: bool = typer.Option(
        False, "--json",
        help="Emit a single JSON document on stdout summarizing the run "
             "(for the /ingest-new-inventory-list slash command).",
    ),
):
    """Ingest a filled-in inventory checklist, then archive it under input/processed/.

    Imports the qty_normal / qty_foil cells into ``set:<anchor>`` honoring
    the file's partition (set codes + rarity from ``_meta`` or inferred from
    rows), then atomically renames the file with a timestamp so the next
    ``mm set master-list`` run will produce a fresh inventory checklist
    pre-populated from the now-saved DB state.

    The ingest mode is read from the checklist's _meta sheet by default
    (``modify`` checklists → replace semantics, ``add`` checklists → additive
    semantics). Pass ``--mode`` explicitly to override; the override is honored
    with a stderr warning when it disagrees with the file's declared mode.
    """
    if mode is not None and mode not in ("replace", "additive"):
        typer.echo(f"error: --mode must be 'replace' or 'additive', got {mode!r}", err=True)
        raise typer.Exit(2)

    # ---- kind dispatch (early): deck checklists (precon / jumpstart, its
    # species) route to the shared deck-checklist engine BEFORE the
    # inventory-specific anchor resolution below. The global precon catalog has
    # no anchor_code/set_codes in its _meta (it spans every set), so it must
    # short-circuit here or the anchor resolution would reject it. Only applies
    # with an explicit --path (deck checklists are always ingested by path).
    if path is not None and path.exists():
        early_meta = sets_mod.read_master_list_meta(path) or {}
        early_kind = early_meta.get("kind")
        if early_kind in ("jumpstart", "precon"):
            _ingest_deck_checklist(path, kind=early_kind, sha=_file_sha256(path),
                                   force=force, json_out=json_out)
            return

    # Resolve path + anchor. Either name_or_code or --path must give us enough.
    if path is not None:
        src = path
        if not src.exists():
            typer.echo(f"error: no inventory checklist found at {src}", err=True)
            raise typer.Exit(2)
        # Try the file's _meta first; fall back to name_or_code arg.
        meta = sets_mod.read_master_list_meta(src)
        if meta and meta.get("anchor_code"):
            anchor = meta["anchor_code"]
            slug = meta.get("slug") or _slug(anchor)
        elif name_or_code:
            try:
                r = sets_mod.resolve(name_or_code)
            except LookupError as e:
                typer.echo(f"error: {e}", err=True); raise typer.Exit(2)
            anchor = r.code
            slug = _slug(r.name)
        else:
            typer.echo(
                "error: --path file has no _meta sheet; pass NAME_OR_CODE to disambiguate",
                err=True,
            )
            raise typer.Exit(2)
    else:
        if not name_or_code:
            typer.echo("error: provide either NAME_OR_CODE or --path", err=True)
            raise typer.Exit(2)
        try:
            r = sets_mod.resolve(name_or_code)
        except LookupError as e:
            typer.echo(f"error: {e}", err=True); raise typer.Exit(2)
        anchor = r.code
        slug = _slug(r.name)
        # Look for intake docs matching this family's slug, in either format.
        # There can be more than one (master + rarity slices, or both xlsx and
        # md side-by-side); if exactly one matches use it, otherwise force the
        # user to disambiguate via --path.
        candidates = []
        if INPUT_DIR.exists():
            for ext in ("xlsx", "md"):
                candidates.extend(INPUT_DIR.glob(f"{slug}-*.{ext}"))
            candidates = sorted(candidates)
        if not candidates:
            typer.echo(f"error: no inventory checklist (.xlsx/.md) found in {INPUT_DIR}/ for slug {slug!r}", err=True)
            typer.echo(f"  run `mm set master-list {name_or_code!r}` first", err=True)
            raise typer.Exit(2)
        if len(candidates) > 1:
            typer.echo(
                f"error: multiple inventory checklists match slug {slug!r}; "
                "pass --path to choose:",
                err=True,
            )
            for c in candidates:
                typer.echo(f"  - {c}", err=True)
            raise typer.Exit(2)
        src = candidates[0]

    # ---- Mode resolution: auto-detect from _meta.mode, reconcile with --mode ----
    # Read _meta unconditionally now (the path-resolution branches above may
    # or may not have already done so). Source of truth is the file we're
    # about to ingest, regardless of how we found it.
    file_meta = sets_mod.read_master_list_meta(src) or {}
    declared_meta_mode = file_meta.get("mode")  # 'modify', 'add', or None (legacy)
    declared_to_op = {"modify": "replace", "add": "additive"}
    declared_op = declared_to_op.get(declared_meta_mode)  # 'replace', 'additive', or None
    if mode is None:
        # No explicit override → use the file's declared mode.
        if declared_op is None:
            typer.echo(
                f"error: this checklist has no _meta.mode (likely generated before "
                f"mode-aware tagging). Pass --mode replace or --mode additive "
                f"explicitly to ingest it.",
                err=True,
            )
            raise typer.Exit(2)
        mode = declared_op
    elif declared_op is not None and declared_op != mode:
        # User passed --mode AND it disagrees with the file's declaration.
        # Honor the override but warn loudly — getting this wrong on a
        # 'modify' checklist run as 'additive' (or vice versa) silently
        # corrupts the inventory state.
        typer.echo(
            f"warning: file's _meta.mode is {declared_meta_mode!r} (would ingest "
            f"as {declared_op!r}); --mode override is {mode!r} — applying "
            f"{mode!r} as requested. If this is wrong, ctrl-C now.",
            err=True,
        )
    # else: --mode passed and either agrees with declared OR file is legacy
    # (declared_op is None and user provided explicit --mode, which is fine).

    # Resolve zero_untouched (replace/modify only). Tri-state --flag:
    #   explicit --zero-untouched / --no-zero-untouched → honor it.
    #   unset + --json (non-interactive) → False (safe: file is a delta).
    #   unset + interactive → ask; default No.
    do_zero = False
    if mode == "replace":
        if zero_untouched is not None:
            do_zero = zero_untouched
        elif json_out:
            do_zero = False
        else:
            do_zero = typer.confirm(
                "This is a MODIFY (replace) ingest. Each row in the file is set "
                "to its cell value. Also ZERO in-partition rows that are ABSENT "
                "from the file? (only if the file is a full-inventory audit)",
                default=False,
            )
    sha = _file_sha256(src)
    with db.connect() as conn:
        prior = db.find_ingest_log_by_hash(conn, sha)
    prior_success = next((p for p in prior if p["status"] == "success"), None)
    if prior_success and not force:
        msg = (
            f"this file's SHA-256 matches a previous successful ingest "
            f"(log id {prior_success['id']}, "
            f"mode {prior_success['mode']}, at {prior_success['at']})."
        )
        if json_out:
            json.dump({
                "status": "duplicate",
                "file": str(src),
                "sha256": sha,
                "prior_log": prior_success,
                "message": msg,
            }, sys.stdout, indent=2)
            sys.stdout.write("\n")
        else:
            typer.echo(f"refusing to re-ingest: {msg}", err=True)
            typer.echo("  pass --force to proceed (the file will be re-applied "
                       "in the chosen --mode and a new log row will be added).", err=True)
        raise typer.Exit(EXIT_DUPLICATE_INGEST)

    # Run the actual import — V2 path writes directly to the inventory table,
    # honoring the file's partition (set codes + rarity from _meta or rows).
    error: str | None = None
    result: dict | None = None
    try:
        result = sets_mod.ingest_inventory_from_xlsx(src, mode=mode, zero_untouched=do_zero)
    except Exception as e:
        error = repr(e)

    archived: Path | None = None
    if result is not None:
        # Compute slice suffix from the file's stem for the archive name.
        # Possible stems:
        #   <slug>-checklist           → unsliced (V1.6+)
        #   <slug>-<slice>-checklist   → sliced (V1.6+)
        #   <slug>-<slice>             → pre-V1.6 file the user hand-renamed
        #   <slug>                     → pre-V1.6 unsliced (rare)
        # Strip ``-checklist`` first so the rest is just ``<slug>[-<slice>]``.
        stem = src.stem
        if stem.endswith("-checklist"):
            stem = stem[: -len("-checklist")]
        if stem == slug:
            slice_suffix = ""
        elif stem.startswith(f"{slug}-"):
            slice_suffix = stem[len(slug) + 1:]
        else:
            slice_suffix = stem
        ext = src.suffix.lstrip(".") or "xlsx"
        archived = _processed_path(slug, slice_suffix, ext=ext)
        archived.parent.mkdir(parents=True, exist_ok=True)
        src.rename(archived)

    # Persist the log entry. The label column now records the set anchor as
    # a 'set:<code>' string for backwards compatibility with the ingest_log
    # schema; the row no longer means a list_rows row exists.
    log_label = f"set:{anchor}"
    with db.connect() as conn:
        db.record_ingest_log(
            conn,
            label=log_label,
            mode=mode,
            source_path=str(src),
            archived_path=str(archived) if archived else None,
            file_sha256=sha,
            rows_added=(result or {}).get("added", 0),
            rows_updated=(result or {}).get("updated", 0),
            rows_zeroed=(result or {}).get("zeroed", 0),
            status="success" if error is None else "failed",
            error=error,
        )

    # Snapshot inventory in this set's family post-ingest.
    inv_summary = None
    if error is None:
        try:
            family_codes = set(sets_mod.resolve(anchor).all_codes)
        except LookupError:
            family_codes = {anchor}
        inv_rows = [r for r in inv_mod.inventory_show() if r.set_code in family_codes]
        inv_summary = {
            "distinct_rows": len(inv_rows),
            "total_qty": sum(r.quantity for r in inv_rows),
            "total_value": sum((r.line_value or 0.0) for r in inv_rows),
        }

    if json_out:
        out = {
            "status": "success" if error is None else "failed",
            "file": str(src),
            "archived_path": str(archived) if archived else None,
            "anchor_code": anchor,
            "mode": mode,
            "zero_untouched": do_zero,
            "sha256": sha,
            "rows_added": (result or {}).get("added", 0),
            "rows_updated": (result or {}).get("updated", 0),
            "rows_zeroed": (result or {}).get("zeroed", 0),
            "warnings": (result or {}).get("warnings", []),
            "not_found": (result or {}).get("not_found", []),
            "extras": (result or {}).get("extras", []),
            "inventory_summary": inv_summary,
            "error": error,
        }
        json.dump(out, sys.stdout, indent=2, default=str)
        sys.stdout.write("\n")
        if error is not None:
            raise typer.Exit(2)
        return

    if error is not None:
        typer.echo(f"error: ingest failed: {error}", err=True)
        raise typer.Exit(2)

    zero_note = "" if mode != "replace" else (
        " [zeroed absent rows]" if do_zero else " [absent rows left alone]"
    )
    typer.echo(
        f"Inventory ({anchor}): {result['updated']} updated, "
        f"{result['added']} added, {result['zeroed']} zeroed (mode={mode}{zero_note})"
    )
    for w in result["warnings"]:
        typer.echo(f"  warning: {w}", err=True)
    for nf in result["not_found"]:
        if "raw" in nf:
            typer.echo(f"  not found: {nf['raw']} ({nf.get('reason','')})", err=True)
        else:
            typer.echo(f"  not found: {nf}", err=True)
    for ex in result["extras"]:
        typer.echo(f"  extra (outside file's partition): {ex['raw']}", err=True)
    typer.echo(f"Archived inventory checklist → {archived}")
    if inv_summary:
        typer.echo(
            f"Inventory in {anchor} family: {inv_summary['distinct_rows']} rows, "
            f"qty {inv_summary['total_qty']}, value ${inv_summary['total_value']:.2f}"
        )


# ---------- inventory (V2) ----------

def _read_text_or_path(source: str | None) -> tuple[str | None, Path | None]:
    """Resolve the (text, path) tuple for an import source argument.

    None or '-' means stdin. A non-existent path errors out.
    """
    if source is None or source == "-":
        return sys.stdin.read(), None
    p = Path(source)
    if not p.exists():
        typer.echo(f"error: file not found: {p}", err=True); raise typer.Exit(2)
    return None, p


def _resolve_block(text: str | None, path: Path | None):
    """Parse a text block / file with parsers.parse_text + resolve.

    Returns the ParseResult. Caller decides how to route entries into the
    target table (inventory / wishlist / deck_cards).
    """
    from . import parsers as _parsers
    if path is not None:
        fmt = _parsers.detect_format(path)
        if fmt == "xlsx":
            result = _parsers.parse_master_list_xlsx(path)
        elif fmt == "md":
            result = _parsers.parse_master_list_md(path)
        else:
            result = _parsers.parse_text(path.read_text(encoding="utf-8"))
    else:
        result = _parsers.parse_text(text)
    _parsers.resolve(result)
    return result


@inventory_app.command("show")
def inventory_show_cmd():
    """Show every printing in inventory with quantities and current value."""
    rows = inv_mod.inventory_show()
    if not rows:
        typer.echo("(inventory empty)"); return
    typer.echo(f"{'qty':>4} {'finish':>7} {'set':>6} {'cn':>6}  name (rarity, usd)")
    for r in rows:
        usd = f"${r.unit_price:.2f}" if r.unit_price is not None else "—"
        typer.echo(f"{r.quantity:>4} {r.finish:>7} {r.set_code:>6} "
                   f"{r.collector_number:>6}  {r.display_name} ({r.rarity}, {usd})")


@inventory_app.command("value")
def inventory_value_cmd():
    """Total inventory value in USD."""
    v = inv_mod.inventory_value()
    typer.echo(f"Inventory: ${v['total']:.2f} across {v['rows']} rows")
    if v["missing_price"]:
        typer.echo(f"Cards without USD price ({len(v['missing_price'])}):")
        for name, set_code, cn, finish in v["missing_price"]:
            typer.echo(f"  {name} ({set_code.upper()}) {cn} [{finish}]")


@inventory_app.command("add")
def inventory_add_cmd(
    scryfall_id: str = typer.Argument(...),
    finish: str = typer.Argument(..., help="nonfoil | foil"),
    qty: int = typer.Argument(1),
    replace: bool = typer.Option(False, "--replace", help="Set quantity outright instead of summing."),
):
    """Add (or replace) a single printing in inventory."""
    try:
        result = inv_mod.inventory_add(scryfall_id, finish, qty, replace=replace)
    except ValueError as e:
        typer.echo(f"error: {e}", err=True); raise typer.Exit(2)
    typer.echo(f"{result['action']}: {scryfall_id} {finish} qty={result['new_qty']}")


def _parse_card_spec(spec: str) -> tuple[str, str, str, int]:
    """Parse a single card spec into (set_code, collector_number, finish, qty).

    Accepts a space form (``"SET CN [finish] [qty]"``) or a colon form
    (``"SET:CN[:finish][:qty]"``). finish defaults to "nonfoil", qty defaults
    to 1.
    """
    if ":" in spec:
        tokens = spec.split(":")
    else:
        tokens = spec.split()
    if len(tokens) < 2:
        raise ValueError(f"malformed card spec {spec!r}: expected at least 'SET CN'")
    set_code = tokens[0].lower()
    collector_number = tokens[1]
    finish = "nonfoil"
    qty = 1
    for token in tokens[2:]:
        if token == "":
            continue
        if token.lower() in ("foil", "nonfoil"):
            finish = token.lower()
        elif token.isdigit():
            n = int(token)
            if n <= 0:
                raise ValueError(f"malformed card spec {spec!r}: qty must be a positive integer, got {token!r}")
            qty = n
        else:
            raise ValueError(f"malformed card spec {spec!r}: unrecognized token {token!r}")
    return set_code, collector_number, finish, qty


@inventory_app.command("add-card")
def inventory_add_card_cmd(
    specs: list[str] = typer.Argument(..., help="One or more card specs: 'SET CN [finish] [qty]' or 'SET:CN[:finish][:qty]'. finish=nonfoil|foil (default nonfoil), qty default 1."),
    replace: bool = typer.Option(False, "--replace", help="Set quantity outright instead of summing."),
    json_out: bool = typer.Option(False, "--json", help="Emit result as JSON."),
):
    """Add one or more cards to inventory by set + collector number."""
    from . import parsers as _parsers
    try:
        parsed = [_parse_card_spec(spec) for spec in specs]
    except ValueError as e:
        typer.echo(f"error: {e}", err=True); raise typer.Exit(2)

    result = _parsers.ParseResult(entries=[
        _parsers.Entry(qty=qty, raw=spec, name="", set=set_code, collector_number=cn, foil=(finish == "foil"), section="mainboard")
        for spec, (set_code, cn, finish, qty) in zip(specs, parsed)
    ])
    _parsers.resolve(result)

    added = updated = 0
    cards_out = []
    with db.connect() as conn:
        for entry in result.entries:
            if entry.card is None:
                continue
            db.upsert_card(conn, entry.card)
    for entry in result.entries:
        if entry.card is None:
            continue
        finish = "foil" if entry.foil else "nonfoil"
        r = inv_mod.inventory_add(entry.card["id"], finish, entry.qty, replace=replace)
        if r["action"] == "inserted":
            added += 1
        else:
            updated += 1
        cards_out.append({
            "set": entry.set, "cn": entry.collector_number, "finish": finish,
            "qty": r["new_qty"], "name": entry.card.get("name"), "action": r["action"],
        })
        if not json_out:
            typer.echo(f"  {r['action']}: {entry.card.get('name')} ({entry.set} {entry.collector_number}) [{finish}] qty={r['new_qty']}")

    if json_out:
        json.dump({
            "added": added, "updated": updated, "cards": cards_out,
            "warnings": result.warnings, "not_found": result.not_found,
        }, sys.stdout, indent=2)
        sys.stdout.write("\n")
        return

    typer.echo(f"Inventory: {added} added, {updated} updated")
    for w in result.warnings:
        typer.echo(f"  warning: {w}", err=True)
    for nf in result.not_found:
        if isinstance(nf, dict) and "raw" in nf:
            typer.echo(f"  not found: {nf['raw']} ({nf.get('reason','')})", err=True)
        else:
            typer.echo(f"  not found: {nf}", err=True)


@inventory_app.command("remove")
def inventory_remove_cmd(
    scryfall_id: str = typer.Argument(...),
    finish: str = typer.Argument(..., help="nonfoil | foil"),
    qty: int = typer.Option(None, "--qty", help="Subtract this many (omit to delete the row)."),
):
    """Remove (or decrement) a single printing in inventory."""
    result = inv_mod.inventory_remove(scryfall_id, finish, qty)
    typer.echo(f"{result['action']}: {scryfall_id} {finish} new_qty={result['new_qty']}")


@inventory_app.command("import")
def inventory_import_cmd(
    source: str = typer.Argument(None, help="Path to file (XLSX/text) or '-' for stdin."),
):
    """Read a Moxfield-style block (stdin or file) and add to inventory.

    Insert-or-sum semantics: re-importing the same block doubles the qty.
    Use ``mm inventory remove`` or ``mm inventory add --replace`` to undo.
    """
    text, path = _read_text_or_path(source)
    result = _resolve_block(text, path)
    added = updated = 0
    with db.connect() as conn:
        for entry in result.entries:
            if entry.card is None:
                continue
            db.upsert_card(conn, entry.card)
    for entry in result.entries:
        if entry.card is None:
            continue
        finish = "foil" if entry.foil else "nonfoil"
        r = inv_mod.inventory_add(entry.card["id"], finish, entry.qty)
        if r["action"] == "inserted":
            added += 1
        else:
            updated += 1
    typer.echo(f"Inventory: {added} added, {updated} updated")
    for w in result.warnings:
        typer.echo(f"  warning: {w}", err=True)
    for nf in result.not_found:
        if isinstance(nf, dict) and "raw" in nf:
            typer.echo(f"  not found: {nf['raw']} ({nf.get('reason','')})", err=True)
        else:
            typer.echo(f"  not found: {nf}", err=True)


# ---------- wishlist (V2) ----------

@wishlist_app.command("show")
def wishlist_show_cmd(
    category: str = typer.Option(None, "--category", "-c", help="Filter to one category."),
):
    """Show wishlist entries (optionally filtered to one category)."""
    rows = wishlist_mod.wishlist_show(category=category)
    if not rows:
        scope = f"category={category!r}" if category else "all categories"
        typer.echo(f"(wishlist empty for {scope})"); return
    typer.echo(f"{'qty':>4} {'finish':>7} {'set':>6} {'cn':>6}  category   name (rarity, usd)")
    for r in rows:
        usd = f"${r.unit_price:.2f}" if r.unit_price is not None else "—"
        typer.echo(f"{r.qty_wanted:>4} {r.finish:>7} {r.set_code:>6} "
                   f"{r.collector_number:>6}  {r.category:10} {r.display_name} ({r.rarity}, {usd})")


@wishlist_app.command("categories")
def wishlist_categories_cmd():
    """List distinct wishlist categories with row/qty counts."""
    cats = wishlist_mod.wishlist_categories()
    if not cats:
        typer.echo("(no wishlist entries)"); return
    typer.echo(f"{'category':30} {'rows':>6} {'qty':>6}")
    for c in cats:
        typer.echo(f"{c['category']:30} {c['rows']:>6} {c['total_qty']:>6}")


@wishlist_app.command("value")
def wishlist_value_cmd(
    category: str = typer.Option(None, "--category", "-c"),
):
    """Total wishlist value in USD (acquisition floor)."""
    v = wishlist_mod.wishlist_value(category=category)
    scope = f"({category})" if category else "(all)"
    typer.echo(f"Wishlist {scope}: ${v['total']:.2f} across {v['rows']} rows")
    if v["missing_price"]:
        typer.echo(f"Cards without USD price ({len(v['missing_price'])}):")
        for name, set_code, cn, finish in v["missing_price"]:
            typer.echo(f"  {name} ({set_code.upper()}) {cn} [{finish}]")


@wishlist_app.command("add")
def wishlist_add_cmd(
    scryfall_id: str = typer.Argument(...),
    finish: str = typer.Argument(..., help="nonfoil | foil | either"),
    category: str = typer.Argument("default"),
    qty: int = typer.Argument(1),
    priority: int = typer.Option(None, "--priority"),
    notes: str = typer.Option(None, "--notes"),
):
    """Add a single printing to a wishlist category."""
    try:
        result = wishlist_mod.wishlist_add(scryfall_id, finish, category, qty,
                                           priority=priority, notes=notes)
    except ValueError as e:
        typer.echo(f"error: {e}", err=True); raise typer.Exit(2)
    typer.echo(f"{result['action']}: {scryfall_id} {finish} {category} qty={result['new_qty']}")


@wishlist_app.command("remove")
def wishlist_remove_cmd(
    scryfall_id: str = typer.Argument(...),
    finish: str = typer.Argument(..., help="nonfoil | foil | either"),
    category: str = typer.Argument("default"),
    qty: int = typer.Option(None, "--qty"),
):
    """Remove a wishlist entry (or decrement its qty)."""
    result = wishlist_mod.wishlist_remove(scryfall_id, finish, category, qty)
    typer.echo(f"{result['action']}: {scryfall_id} {finish} {category} new_qty={result['new_qty']}")


@wishlist_app.command("import")
def wishlist_import_cmd(
    category: str = typer.Argument(...),
    source: str = typer.Argument(None, help="Path to file or '-' for stdin."),
    finish: str = typer.Option("either", "--finish", help="Default finish for imported lines."),
):
    """Read a Moxfield-style block and add to a wishlist category."""
    if finish not in ("nonfoil", "foil", "either"):
        typer.echo(f"error: --finish must be nonfoil|foil|either, got {finish!r}", err=True)
        raise typer.Exit(2)
    text, path = _read_text_or_path(source)
    result = _resolve_block(text, path)
    added = updated = 0
    with db.connect() as conn:
        for entry in result.entries:
            if entry.card is None:
                continue
            db.upsert_card(conn, entry.card)
    for entry in result.entries:
        if entry.card is None:
            continue
        # If the entry's parsed finish is foil, use that; otherwise the --finish default.
        eff_finish = "foil" if entry.foil else finish
        r = wishlist_mod.wishlist_add(entry.card["id"], eff_finish, category, entry.qty)
        if r["action"] == "inserted":
            added += 1
        else:
            updated += 1
    typer.echo(f"Wishlist {category!r}: {added} added, {updated} updated")
    for w in result.warnings:
        typer.echo(f"  warning: {w}", err=True)
    for nf in result.not_found:
        if isinstance(nf, dict) and "raw" in nf:
            typer.echo(f"  not found: {nf['raw']} ({nf.get('reason','')})", err=True)
        else:
            typer.echo(f"  not found: {nf}", err=True)


# ---------- deck (V2) ----------

@deck_app.command("ls")
def deck_ls_cmd():
    """List every deck. The ``state`` column flags precon units that aren't
    built playable decks: ``decon`` (torn down for parts) or ``pool`` (a card
    pool — Starter Collection / Scene Box — that was never a deck)."""
    ds = decks_mod.deck_list()
    if not ds:
        typer.echo("(no decks)"); return
    _state_flag = {"built": "", "deconstructed": "decon", "pool": "pool"}
    typer.echo(f"{'slug':30} {'name':40} {'format':12} {'state':6} {'updated_at'}")
    for d in ds:
        flags = _state_flag.get(getattr(d, "precon_state", "built"), "")
        typer.echo(f"{d.slug:30} {d.name:40} {(d.format or '—'):12} {flags:6} {d.updated_at}")


@deck_app.command("show")
def deck_show_cmd(slug: str = typer.Argument(...)):
    """Show every card in a deck (all boards)."""
    try:
        rows = decks_mod.deck_show(slug)
    except LookupError as e:
        typer.echo(f"error: {e}", err=True); raise typer.Exit(2)
    if not rows:
        typer.echo(f"(deck {slug!r} is empty)"); return
    typer.echo(f"{'cnt':>4} {'finish':>7} {'board':>10} {'set':>6} {'cn':>6}  name (rarity, usd)")
    for r in rows:
        usd = f"${r.unit_price:.2f}" if r.unit_price is not None else "—"
        typer.echo(f"{r.count:>4} {r.finish:>7} {r.board:>10} {r.set_code:>6} "
                   f"{r.collector_number:>6}  {r.display_name} ({r.rarity}, {usd})")


@deck_app.command("create")
def deck_create_cmd(
    slug: str = typer.Argument(...),
    name: str = typer.Option(..., "--name"),
    format: str = typer.Option(None, "--format"),
    archetype: str = typer.Option(None, "--archetype"),
    notes: str = typer.Option(None, "--notes"),
):
    """Create a new (empty) deck."""
    try:
        d = decks_mod.deck_create(slug, name, format=format, archetype=archetype, notes=notes)
    except ValueError as e:
        typer.echo(f"error: {e}", err=True); raise typer.Exit(2)
    typer.echo(f"Created deck #{d.deck_id}: {d.slug} ({d.name})")


@deck_app.command("delete")
def deck_delete_cmd(
    slug: str = typer.Argument(...),
    yes: bool = typer.Option(False, "--yes", "-y"),
):
    """Delete a deck (cascades to deck_cards)."""
    if not yes:
        typer.echo("refusing without --yes; this deletes the deck and all its cards.", err=True)
        raise typer.Exit(2)
    n = decks_mod.deck_delete(slug)
    typer.echo(f"Deleted {n} deck(s)")


@deck_app.command("find")
def deck_find_cmd(
    query: str = typer.Argument(
        ...,
        help="A scryfall_id, a 'set cn' pair (QUOTED, e.g. 'fin 248'), or an exact card name (QUOTED if it has spaces). Tries each form in order.",
    ),
    json_out: bool = typer.Option(False, "--json"),
):
    """List every deck that contains a given printing.

    Resolution order: scryfall_id (UUID-shaped) → 'set cn' pair (two
    whitespace-separated tokens) → exact case-insensitive name match against
    cards.name OR cards.flavor_name. Reports per-deck commitments plus the
    inventory↔committed↔available math for the resolved scryfall_id.
    """
    import re as _re

    q = query.strip()
    candidates: list[str] = []  # scryfall_ids matching the query
    with db.connect() as conn:
        # Form 1: UUID-shaped → exact scryfall_id lookup.
        if _re.fullmatch(r"[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}", q.lower()):
            r = conn.execute("SELECT scryfall_id FROM cards WHERE scryfall_id = ?", (q.lower(),)).fetchone()
            if r is not None:
                candidates.append(r["scryfall_id"])

        # Form 2: 'set cn' two-token form.
        if not candidates:
            parts = q.split()
            if len(parts) == 2:
                setc, cn = parts[0].lower(), parts[1]
                rows = conn.execute(
                    "SELECT scryfall_id FROM cards WHERE set_code = ? AND collector_number = ?",
                    (setc, cn),
                ).fetchall()
                candidates.extend(r["scryfall_id"] for r in rows)

        # Form 3: exact name (or flavor_name) match, case-insensitive.
        if not candidates:
            rows = conn.execute(
                "SELECT scryfall_id FROM cards "
                "WHERE LOWER(name) = ? OR LOWER(flavor_name) = ? "
                "ORDER BY set_code, collector_number",
                (q.lower(), q.lower()),
            ).fetchall()
            candidates.extend(r["scryfall_id"] for r in rows)

    if not candidates:
        typer.echo(f"no card found matching {query!r} (tried scryfall_id, 'set cn', and exact name)", err=True)
        raise typer.Exit(1)

    # For each candidate scryfall_id, gather (deck slug, board, finish, count)
    # plus inventory and computed available.
    results: list[dict] = []
    with db.connect() as conn:
        for sid in candidates:
            card = conn.execute(
                "SELECT name, flavor_name, set_code, collector_number, rarity FROM cards WHERE scryfall_id=?",
                (sid,),
            ).fetchone()
            deck_rows = conn.execute(
                "SELECT d.slug, dc.board, dc.finish, dc.count "
                "FROM deck_cards dc JOIN decks d ON d.deck_id = dc.deck_id "
                "WHERE dc.scryfall_id = ? "
                "ORDER BY d.slug, dc.board, dc.finish",
                (sid,),
            ).fetchall()
            inv_rows = conn.execute(
                "SELECT finish, quantity FROM inventory WHERE scryfall_id = ?",
                (sid,),
            ).fetchall()
            committed_by_finish: dict[str, int] = {}
            for r in deck_rows:
                committed_by_finish[r["finish"]] = committed_by_finish.get(r["finish"], 0) + r["count"]
            owned_by_finish = {r["finish"]: r["quantity"] for r in inv_rows}
            available_by_finish = {
                fin: max(0, owned_by_finish.get(fin, 0) - committed_by_finish.get(fin, 0))
                for fin in set(owned_by_finish) | set(committed_by_finish)
            }
            results.append({
                "scryfall_id": sid,
                "name": card["name"],
                "flavor_name": card["flavor_name"],
                "set": card["set_code"],
                "collector_number": card["collector_number"],
                "rarity": card["rarity"],
                "decks": [
                    {"slug": r["slug"], "board": r["board"], "finish": r["finish"], "count": r["count"]}
                    for r in deck_rows
                ],
                "owned": owned_by_finish,
                "committed": committed_by_finish,
                "available": available_by_finish,
            })

    if json_out:
        json.dump(results, sys.stdout, indent=2); sys.stdout.write("\n")
        return

    for res in results:
        flavor = res["flavor_name"]
        display = f"{flavor} / {res['name']}" if flavor else res["name"]
        setc = (res["set"] or "?").upper()
        cn = res["collector_number"] or "?"
        typer.echo(f"\n{display} ({setc} {cn}, {res['rarity']})  scryfall_id={res['scryfall_id']}")
        if not res["decks"]:
            typer.echo("  (not in any deck)")
        else:
            for dk in res["decks"]:
                typer.echo(f"  deck={dk['slug']:<30} board={dk['board']:<10} finish={dk['finish']:<8} count={dk['count']}")
        # Math line
        finish_keys = sorted(set(res["owned"]) | set(res["committed"]))
        for fin in finish_keys:
            o = res["owned"].get(fin, 0)
            c = res["committed"].get(fin, 0)
            a = res["available"].get(fin, 0)
            typer.echo(f"  {fin:>7}: owned={o}  committed={c}  available={a}")


@deck_app.command("value")
def deck_value_cmd(slug: str = typer.Argument(...)):
    """Total deck value in USD."""
    try:
        v = decks_mod.deck_value(slug)
    except LookupError as e:
        typer.echo(f"error: {e}", err=True); raise typer.Exit(2)
    typer.echo(f"Deck {slug!r}: ${v['total']:.2f} across {v['rows']} rows")
    if v["missing_price"]:
        typer.echo(f"Cards without USD price ({len(v['missing_price'])}):")
        for name, set_code, cn, finish in v["missing_price"]:
            typer.echo(f"  {name} ({set_code.upper()}) {cn} [{finish}]")


@deck_app.command("import-precon")
def deck_import_precon_cmd(
    file_name: str = typer.Argument(
        ...,
        help="MTGJSON deck fileName (e.g. CounterBlitzFinalFantasyX_FIC). See `mm mtgjson decks` to list available precons.",
    ),
    slug: str = typer.Option(
        None, "--slug",
        help="Override the auto-derived slug. Defaults to slugified MTGJSON deck name.",
    ),
    name: str = typer.Option(
        None, "--name",
        help="Override the deck's display name. Defaults to MTGJSON 'name' field.",
    ),
    copies: int = typer.Option(
        1, "--copies", min=1,
        help="Multiply inventory-add by N (you opened N physical copies). Deck composition is still created ONCE — a recipe is a recipe. Use `mm deck compose <slug>` afterwards to pledge the physical copies to the deck.",
    ),
    add_inventory: bool = typer.Option(
        True, "--add-inventory/--no-add-inventory",
        help="Also add the precon's cards to inventory (default: yes). --no-add-inventory builds the deck composition without claiming physical ownership.",
    ),
    state: str = typer.Option(
        None, "--state",
        help="Precon unit state: 'built' (assembled deck, cards pledged), "
             "'deconstructed' (torn down for parts — recipe kept, cards loose), "
             "or 'pool' (a card pool like the Starter Collection / a Scene Box "
             "— never a deck; cards loose, marker row). Default: auto-detected "
             "(pool for pool-like products, else built).",
    ),
    deconstruct: bool = typer.Option(
        False, "--deconstruct", hidden=True,
        help="Deprecated alias for --state deconstructed.",
    ),
    merge_inventory: bool = typer.Option(
        False, "--merge-inventory",
        help="Skip deck creation; the composition already exists at this slug. Just add another copy's worth of cards to inventory. Requires the deck to already exist.",
    ),
):
    """Import an MTGJSON precon into the local DB.

    Creates one deck row (carrying the MTGJSON fileName, so precon unit counts
    derive straight from the decks table) AND adds the cards to inventory. The
    row's ``--state`` decides handling: ``built`` pledges a physical copy;
    ``deconstructed`` / ``pool`` leave cards loose. When ``--state`` is omitted
    it's auto-detected — pool-like products (Starter Collection, Scene Box)
    default to ``pool``, everything else to ``built`` — and the choice is
    printed. This and the precon checklist feed one source of truth (no ledger).

    The MTGJSON Card(Deck) entries carry `identifiers.scryfallId` which maps
    directly to our cards table. No Scryfall API calls; the precon JSON is
    cached after first fetch.
    """
    # Resolve the state: explicit --state > --deconstruct alias > auto-detect.
    if state is not None:
        if state not in ("built", "deconstructed", "pool"):
            typer.echo(f"error: --state must be built|deconstructed|pool, got {state!r}", err=True)
            raise typer.Exit(2)
        resolved_state = state
    elif deconstruct:
        resolved_state = "deconstructed"
    elif merge_inventory:
        resolved_state = "built"  # unused (no row created), keep valid
    else:
        resolved_state = mtgjson_mod.default_precon_state(file_name, name=name)
        if resolved_state == "pool":
            typer.echo(
                f"ℹ auto-detected a card POOL (not a playable deck) → recording as "
                f"pool: cards go loose in inventory, no pledged deck. "
                f"Override with --state built if you really want a deck.",
            )
    try:
        result = decks_mod.import_precon(
            file_name,
            slug=slug,
            name=name,
            copies=copies,
            add_inventory=add_inventory,
            deconstruct=(resolved_state != "built"),
            precon_state=(None if merge_inventory else resolved_state),
            merge_inventory=merge_inventory,
        )
    except mtgjson_mod.MtgJsonError as e:
        typer.echo(f"error: could not fetch MTGJSON precon {file_name!r}: {e}", err=True)
        raise typer.Exit(2)
    except ValueError as e:
        typer.echo(f"error: {e}", err=True)
        raise typer.Exit(2)

    deck_name = result["deck_name"]
    effective_slugs = result["effective_slugs"]
    if not merge_inventory and resolved_state != "built":
        slug_list = ", ".join(effective_slugs) or "(none)"
        label = "torn-down copy" if resolved_state == "deconstructed" else "card pool"
        typer.echo(
            f"Imported precon {deck_name!r} as {resolved_state} ({label}): "
            f"marker deck {slug_list}, cards loose in inventory — "
            f"{result['inv_added']} new rows, {result['inv_updated']} bumped, "
            f"{result['inv_qty_total']} total card-qty across "
            f"{result['inv_distinct']} distinct (printing, finish) entries × {copies} copies."
        )
    elif merge_inventory:
        typer.echo(
            f"Merged {copies} physical copies of {deck_name!r} into inventory: "
            f"{result['inv_added']} new rows, {result['inv_updated']} bumped "
            f"({result['inv_qty_total']} total card-qty). Deck composition untouched."
        )
    else:
        slug_list = ", ".join(effective_slugs) or "(none)"
        copies_note = f" × {copies} physical copies" if copies > 1 else ""
        typer.echo(
            f"Imported precon {deck_name!r} as deck: {slug_list}. "
            f"Deck rows: {result['deck_added']} added, {result['deck_updated']} updated "
            f"({result['deck_card_qty']} total card-qty)."
        )
        if add_inventory:
            typer.echo(
                f"Inventory{copies_note}: {result['inv_added']} new rows, "
                f"{result['inv_updated']} bumped ({result['inv_qty_total']} total card-qty)."
            )
            if copies > 1:
                typer.echo(
                    f"  Physical copies are LOOSE. Pledge one to the deck with "
                    f"`mm deck compose {effective_slugs[0]}`."
                )
        else:
            typer.echo("Inventory: skipped (--no-add-inventory).")

    missing_sids = result["missing_sids"]
    if missing_sids:
        typer.echo(
            f"warning: {len(missing_sids)} entries had no scryfallId and were skipped:",
            err=True,
        )
        for m in missing_sids[:5]:
            typer.echo(f"  - {m['name']} ({m['set']} {m['cn']}, board={m['board']})", err=True)
        if len(missing_sids) > 5:
            typer.echo(f"  ...and {len(missing_sids) - 5} more", err=True)


def _resolve_precon_filenames(
    target: str,
    name_query: str | None,
    *,
    want_all: bool,
    only_type: str | None,
    include_collector: bool,
) -> list[dict]:
    """Resolve the add-precon selector to a list of MTGJSON deck dicts.

    Two forms:
      - ``target`` is an exact MTGJSON fileName (``Name_CODE``) → that one deck.
      - ``target`` is a set code → every physical precon in that set
        (``--all`` / no name query), or the fuzzy-name match of ``name_query``.

    Raises ``LookupError`` (→ exit 2) when nothing matches or a name query is
    ambiguous; the message lists the candidates so the caller can disambiguate.
    """
    # Exact fileName? (deck() succeeds only for a real fileName; cheap, cached.)
    if name_query is None and not want_all:
        try:
            d = mtgjson_mod.deck(target)
            return [{
                "fileName": target,
                "name": d.get("name") or target,
                "type": d.get("type") or "",
                "code": (d.get("code") or "").upper(),
            }]
        except mtgjson_mod.MtgJsonError:
            pass  # not a fileName — fall through to set-code resolution

    variants = mtgjson_mod.precon_variants(
        target.lower(), only_type=only_type, include_collector=include_collector,
    )
    if not variants:
        type_note = f" of type {only_type!r}" if only_type else ""
        raise LookupError(
            f"no physical precons{type_note} found for set {target!r} "
            f"(try `mm mtgjson decks --set {target.lower()}` to list them)."
        )

    if want_all or name_query is None:
        return variants

    # Fuzzy name match: case-insensitive substring first, then difflib.
    q = name_query.strip().lower()
    subs = [v for v in variants if q in (v["name"] or "").lower()]
    if len(subs) == 1:
        return subs
    pool = subs or variants
    if len(subs) > 1:
        names = ", ".join(f"{v['name']!r} ({v['fileName']})" for v in subs)
        raise LookupError(
            f"ambiguous name {name_query!r} for set {target!r} — matches: {names}. "
            f"Refine the name or pass --all."
        )
    # No substring hit — try difflib against the full variant list.
    import difflib
    best = difflib.get_close_matches(
        q, [(v["name"] or "").lower() for v in pool], n=3, cutoff=0.5,
    )
    if len(best) == 1:
        return [v for v in pool if (v["name"] or "").lower() == best[0]]
    names = ", ".join(f"{v['name']!r} ({v['fileName']})" for v in pool)
    raise LookupError(
        f"no precon in set {target!r} matched name {name_query!r}. "
        f"Available: {names}. Pass --all to add every deck."
    )


@deck_app.command("add-precon")
def deck_add_precon_cmd(
    target: str = typer.Argument(
        ...,
        help="A set code (e.g. blc) OR an exact MTGJSON fileName (e.g. FamilyMatters_BLC).",
    ),
    name_query: str = typer.Argument(
        None,
        help="Fuzzy deck-name filter within the set (e.g. 'Family Matters'). Omit with a set code to require --all.",
    ),
    constructed: int = typer.Option(
        None, "--constructed", "-c", min=0,
        help="Built copies to add per deck (each creates a deck + adds its cards to inventory).",
    ),
    deconstructed: int = typer.Option(
        None, "--deconstructed", "-d", min=0,
        help="Torn-down copies to add per deck (loose cards, marker deck row).",
    ),
    pool: int = typer.Option(
        None, "--pool", "-p", min=0,
        help="Card-pool copies to add per deck (Starter Collection / Scene Box — cards loose, marker row).",
    ),
    want_all: bool = typer.Option(
        False, "--all",
        help="With a set code: add EVERY physical precon in the set (no name query).",
    ),
    only_type: str = typer.Option(
        None, "--type",
        help="Filter set-code resolution to one exact product type (e.g. 'Commander Deck').",
    ),
    include_collector: bool = typer.Option(
        False, "--include-collector",
        help="Include '… Collector's Edition' variants (excluded by default).",
    ),
    json_out: bool = typer.Option(False, "--json", help="Emit the summary as JSON."),
):
    """Add constructed/deconstructed precon copies AS TRACKED UNITS.

    The one-liner form of filling a precon checklist: resolve the deck(s), then
    run the same deck+inventory transaction the checklist ingest uses. Precon
    unit counts are DERIVED from the ``decks`` table (each copy is a deck row
    carrying the MTGJSON fileName + a precon_state of built/deconstructed/pool),
    so both this and ``mm deck import-precon`` feed the same single source of
    truth — there's no separate ledger to keep in sync.

    Selection:
      - ``mm deck add-precon blc --all``            → all BLC precons
      - ``mm deck add-precon blc "Family Matters"``  → one, by fuzzy name
      - ``mm deck add-precon FamilyMatters_BLC``     → one, by exact fileName
      - ``--type "Commander Deck"``                  → narrow set resolution

    Additive: re-running adds ANOTHER copy (built 1→2). Remove a copy with
    ``mm deck delete <slug>``. If you pass NO count flags, the state is
    auto-detected per deck — pool-like products (Starter Collection, Scene Box)
    default to one ``pool`` copy, everything else to one ``built`` copy.
    """
    explicit = any(v is not None for v in (constructed, deconstructed, pool))
    if explicit and not (constructed or deconstructed or pool):
        typer.echo("error: nothing to add — pass --constructed / --deconstructed / --pool > 0.", err=True)
        raise typer.Exit(2)

    try:
        decks = _resolve_precon_filenames(
            target, name_query, want_all=want_all,
            only_type=only_type, include_collector=include_collector,
        )
    except LookupError as e:
        typer.echo(f"error: {e}", err=True)
        raise typer.Exit(2)

    # Build the in-memory precon checklist (add mode) and run the shared engine —
    # the deck+inventory transaction the checklist ingest uses. No XLSX round-trip.
    # When the user gave explicit counts, use them verbatim for every deck.
    # Otherwise auto-detect per deck: pool-like → pool 1, else built 1.
    from . import parsers as _parsers
    rows = []
    auto_pooled = []
    for d in decks:
        if explicit:
            c, dq, p = (constructed or 0), (deconstructed or 0), (pool or 0)
        else:
            st = mtgjson_mod.default_precon_state(d["fileName"], name=d.get("name"))
            c, dq, p = (0, 0, 1) if st == "pool" else (1, 0, 0)
            if st == "pool":
                auto_pooled.append(d.get("name") or d["fileName"])
        rows.append(_parsers.JumpstartRow(
            file_name=d["fileName"], theme=d.get("name") or "",
            keep_qty=c, deconstructed_qty=dq, pool_qty=p,
        ))
    parsed = _parsers.JumpstartParseResult(
        rows=rows, warnings=[], meta={"kind": "precon", "mode": "add"},
    )
    summary = sets_mod._apply_precon_checklist(parsed, mode="add")

    if json_out:
        json.dump(summary, sys.stdout, indent=2)
        sys.stdout.write("\n")
        return

    if auto_pooled:
        typer.echo(
            f"ℹ auto-detected {len(auto_pooled)} card pool(s) (not playable decks) → "
            f"recorded as pool (cards loose): {', '.join(auto_pooled)}. "
            f"Override with --constructed/-c."
        )
    typer.echo(
        f"Added precons: {summary['rows_acted']} deck(s) changed — "
        f"{summary['built']} built, {summary['deconstructed']} torn down, "
        f"{summary['pool']} pooled, {summary['inv_qty_total']} cards added."
    )
    for pr in summary["per_row"]:
        bc, bd, bp = pr["count_before"]
        ac, ad, ap = pr["count_after"]
        typer.echo(
            f"  {pr['label']} ({pr['file_name']}): "
            f"built {bc}→{ac}, deconstructed {bd}→{ad}, pool {bp}→{ap}"
        )
        if pr.get("warning"):
            typer.echo(f"    warning: {pr['warning']}", err=True)
        if pr.get("error"):
            typer.echo(f"    error: {pr['error']}", err=True)
        if pr.get("missing_sids"):
            typer.echo(f"    note: {len(pr['missing_sids'])} entries had no scryfallId (skipped).", err=True)
    for w in summary.get("warnings", []):
        typer.echo(f"  warning: {w}", err=True)


@deck_app.command("import")
def deck_import_cmd(
    slug: str = typer.Argument(...),
    source: str = typer.Argument(None, help="Path to file or '-' for stdin."),
    board: str = typer.Option("main", "--board", help="main | side | commander | companion | maybe"),
):
    """Import a Moxfield-style block into a deck/board.

    The deck must already exist (use ``mm deck create``). All entries land
    on the given --board; for sideboards, run a second import with
    --board side.
    """
    if board not in ("main", "side", "commander", "companion", "maybe"):
        typer.echo(f"error: --board must be one of main/side/commander/companion/maybe, got {board!r}", err=True)
        raise typer.Exit(2)
    if decks_mod.deck_get(slug) is None:
        typer.echo(f"error: no deck {slug!r}; create with `mm deck create`", err=True)
        raise typer.Exit(2)
    text, path = _read_text_or_path(source)
    result = _resolve_block(text, path)
    with db.connect() as conn:
        for entry in result.entries:
            if entry.card is None:
                continue
            db.upsert_card(conn, entry.card)
    added = updated = 0
    for entry in result.entries:
        if entry.card is None:
            continue
        finish = "foil" if entry.foil else "nonfoil"
        r = decks_mod.deck_add_card(slug, entry.card["id"], board, finish, entry.qty)
        if r["action"] == "inserted":
            added += 1
        else:
            updated += 1
    typer.echo(f"Deck {slug!r} (board={board}): {added} added, {updated} updated")
    for w in result.warnings:
        typer.echo(f"  warning: {w}", err=True)
    for nf in result.not_found:
        if isinstance(nf, dict) and "raw" in nf:
            typer.echo(f"  not found: {nf['raw']} ({nf.get('reason','')})", err=True)
        else:
            typer.echo(f"  not found: {nf}", err=True)


# ---------- V5: physical composition (deck_assignments) ----------

@deck_app.command("compose")
def deck_compose_cmd(
    slug: str = typer.Argument(..., help="Deck slug to physically assemble from loose inventory."),
    foil_first: bool = typer.Option(
        False, "--foil-first",
        help="For 'either'-finish recipe slots, prefer foil over nonfoil (default: nonfoil first).",
    ),
    allow_shortfall: bool = typer.Option(
        False, "--allow-shortfall",
        help="Assign whatever inventory covers, skip rows that would overflow. Default: refuse to write anything if ANY row would overflow.",
    ),
    dry_run: bool = typer.Option(
        False, "--dry-run",
        help="Preview the assignment (rows, shortfalls, 'either' choices) without writing.",
    ),
):
    """Pledge loose inventory to a deck to physically assemble it.

    The deck composition (deck_cards) is unchanged. Assignment happens in
    ``deck_assignments`` — a separate join table so inventory qty is preserved
    and the recipe survives a later decompose.

    Overflow protection: refuses to write if the total pledged (including
    prior assignments across all decks) would exceed either the deck's own
    recipe or the free inventory pool.
    """
    try:
        plan = decks_mod.deck_compose_plan(slug, foil_first=foil_first)
    except LookupError as e:
        typer.echo(f"error: {e}", err=True); raise typer.Exit(2)

    if dry_run:
        typer.echo(f"# selector: deck:{slug} (compose plan, foil_first={foil_first})")
        typer.echo(f"# rows: {len(plan['rows'])}  shortfalls: {len(plan['shortfalls'])}  either_choices: {len(plan['either_choices'])}")
        if plan["shortfalls"]:
            typer.echo("shortfalls:")
            for s in plan["shortfalls"][:20]:
                typer.echo(f"  {s['scryfall_id'][:8]}/{s['finish']}: need {s['need']}, free {s['free']}")
            if len(plan["shortfalls"]) > 20:
                typer.echo(f"  ...and {len(plan['shortfalls']) - 20} more")
        if plan["either_choices"]:
            typer.echo(f"either-slot resolutions: {len(plan['either_choices'])} (first 5 shown)")
            for c in plan["either_choices"][:5]:
                typer.echo(f"  {c['scryfall_id'][:8]}: chose {c['chose_finish']} ({c['reason']})")
        return

    try:
        result = decks_mod.deck_assign_from_composition(
            slug, foil_first=foil_first, allow_shortfall=allow_shortfall,
        )
    except LookupError as e:
        typer.echo(f"error: {e}", err=True); raise typer.Exit(2)
    except decks_mod.AssignmentOverflow as e:
        typer.echo(f"error: {e}", err=True)
        typer.echo(
            "Retry with --allow-shortfall to skip overflowing rows, or "
            "add the missing cards to inventory first.",
            err=True,
        )
        raise typer.Exit(3)

    typer.echo(
        f"Deck {slug!r} composed: {result['assigned_rows']} rows pledged "
        f"({result['assigned_qty']} total card-qty)."
    )
    if result.get("either_choices"):
        typer.echo(f"  Resolved {len(result['either_choices'])} 'either'-finish slot(s).")
    if result.get("shortfalls"):
        typer.echo(
            f"warning: skipped {len(result['shortfalls'])} row(s) that would overflow "
            f"(--allow-shortfall). Run `mm deck free {slug} --dry-run` to see them.",
            err=True,
        )


@deck_app.command("decompose")
def deck_decompose_cmd(
    slug: str = typer.Argument(..., help="Deck slug to physically disassemble; recipe survives."),
):
    """Unpledge every card assigned to a deck — the recipe stays intact.

    Under V5 this is a pure ``deck_assignments`` delete: inventory qty is
    unchanged, ``deck_cards`` (the recipe) is unchanged, only the
    "which physical copies are currently pledged" join table shrinks. Use
    ``mm deck delete <slug>`` if you also want to drop the composition.
    """
    try:
        result = decks_mod.deck_unassign_batch(slug, "all")
    except LookupError as e:
        typer.echo(f"error: {e}", err=True); raise typer.Exit(2)
    if result["unassigned_rows"] == 0:
        typer.echo(f"Deck {slug!r} had no assignments to unpledge.")
    else:
        typer.echo(
            f"Deck {slug!r} decomposed: {result['unassigned_rows']} rows "
            f"({result['unassigned_qty']} card-qty) unpledged. Recipe preserved."
        )


@deck_app.command("assign")
def deck_assign_cmd(
    slug: str = typer.Argument(..., help="Deck slug to assign inventory to."),
    from_composition: bool = typer.Option(
        True, "--from-composition/--no-from-composition",
        help="Assign the entire recipe in one shot (default). --no-from-composition reserved for future partial-assign paths.",
    ),
    foil_first: bool = typer.Option(False, "--foil-first"),
    allow_shortfall: bool = typer.Option(False, "--allow-shortfall"),
):
    """Alias for ``mm deck compose`` today. Reserved for a future partial-
    assign flow (e.g. accepting a piped card block for one-by-one binding).
    """
    if not from_composition:
        typer.echo(
            "error: --no-from-composition is reserved for a future partial-assign flow; "
            "for now use `mm deck compose <slug>` (or omit the flag).",
            err=True,
        )
        raise typer.Exit(2)
    # Delegate to compose (same underlying primitive).
    ctx = typer.Context.current if hasattr(typer.Context, "current") else None
    del ctx
    return deck_compose_cmd(
        slug=slug, foil_first=foil_first, allow_shortfall=allow_shortfall, dry_run=False,
    )


@deck_app.command("unassign")
def deck_unassign_cmd(
    slug: str = typer.Argument(..., help="Deck slug to unassign from."),
    all_flag: bool = typer.Option(
        True, "--all/--no-all",
        help="Unassign every row (default). --no-all reserved for future partial paths.",
    ),
):
    """Alias for ``mm deck decompose`` today; reserved for future partial paths."""
    if not all_flag:
        typer.echo(
            "error: --no-all reserved for a future partial-unassign flow. "
            "For now use `mm deck decompose <slug>`.",
            err=True,
        )
        raise typer.Exit(2)
    return deck_decompose_cmd(slug=slug)


@deck_app.command("free")
def deck_free_cmd(
    slug: str = typer.Argument(..., help="Deck slug to plan a compose against."),
    foil_first: bool = typer.Option(False, "--foil-first"),
):
    """Preview what ``mm deck compose <slug>`` would do — shortfalls, 'either'
    resolutions — without writing. Alias for ``mm deck compose <slug> --dry-run``.
    """
    return deck_compose_cmd(
        slug=slug, foil_first=foil_first, allow_shortfall=False, dry_run=True,
    )


# ---------- query (V2 selectors) ----------

QUERIES_DIR = Path("queries")


# _selector_slug merged into _slug (identical implementation) — use _slug.


def _materialize_or_die(selector: str):
    try:
        return sel_mod.materialize(selector)
    except sel_mod.SelectorParseError as e:
        typer.echo(f"error: invalid selector: {e}", err=True); raise typer.Exit(2)
    except LookupError as e:
        typer.echo(f"error: {e}", err=True); raise typer.Exit(2)


def _row_unit_price(r: sel_mod.MaterializedRow) -> float | None:
    if r.finish == "foil":
        return r.card.get("prices_usd_foil")
    return r.card.get("prices_usd")


def _row_line_value(r: sel_mod.MaterializedRow) -> float | None:
    p = _row_unit_price(r)
    return p * r.quantity if p is not None else None


def _row_display_name(r: sel_mod.MaterializedRow) -> str:
    flavor = r.card.get("flavor_name")
    name = r.card.get("name") or ""
    return f"{flavor} / {name}" if flavor else name


_RARITY_RANK = {"mythic": 0, "rare": 1, "uncommon": 2, "common": 3,
                "special": 4, "bonus": 5}
_VALID_SORTS = ("default", "value-desc", "value-asc", "rarity")


def _apply_sort(rows: list[sel_mod.MaterializedRow], sort_key: str) -> list[sel_mod.MaterializedRow]:
    """Apply a named sort to materialized rows, with deterministic tie-breaking.

    'default' is a no-op — the materializer already sorts by (set, cn, finish).
    'value-desc' / 'value-asc' use line value. Unpriced rows always sink to
    the BOTTOM (regardless of direction) — None is informationally similar to
    'unknown' rather than 'cheapest', so surfacing it at the top of value-asc
    would mislead someone scanning for low-hanging fruit. 'rarity' orders
    mythic > rare > uncommon > common > special > bonus. All sorts break ties
    on (set, cn, finish).
    """
    if sort_key == "default":
        return rows
    if sort_key not in _VALID_SORTS:
        typer.echo(f"error: --sort must be one of {_VALID_SORTS}, got {sort_key!r}", err=True)
        raise typer.Exit(2)

    def tie(r: sel_mod.MaterializedRow):
        return (r.card.get("set") or "", r.card.get("collector_number") or "", r.finish)

    if sort_key == "value-desc":
        # Unpriced rows go to the bottom: priority bit 1 for unpriced, 0 for priced.
        return sorted(rows, key=lambda r: (
            0 if _row_line_value(r) is not None else 1,
            -(_row_line_value(r) or 0.0),
            tie(r),
        ))
    if sort_key == "value-asc":
        # Same priority bit so unpriced rows stay at the bottom in asc too.
        return sorted(rows, key=lambda r: (
            0 if _row_line_value(r) is not None else 1,
            _row_line_value(r) if _row_line_value(r) is not None else 0.0,
            tie(r),
        ))
    if sort_key == "rarity":
        return sorted(rows, key=lambda r: (_RARITY_RANK.get((r.card.get("rarity") or "").lower(), 99), tie(r)))
    return rows


@query_app.command("show")
def query_show_cmd(
    selector: str = typer.Argument(..., help="V2 selector, e.g. 'inventory' or 'set:sld missing rarity=mythic'"),
    first: int = typer.Option(None, "--first", help="Cap displayed rows (total count still printed)."),
    sort: str = typer.Option("default", "--sort",
        help="Sort order: default (set,cn,finish) | value-desc | value-asc | rarity."),
    json_out: bool = typer.Option(False, "--json"),
):
    """Show rows matching a selector."""
    rows = _materialize_or_die(selector)
    rows = _apply_sort(rows, sort)
    if json_out:
        out = [
            {
                "scryfall_id": r.scryfall_id,
                "set": r.card.get("set"),
                "collector_number": r.card.get("collector_number"),
                "name": r.card.get("name"),
                "flavor_name": r.card.get("flavor_name"),
                "rarity": r.card.get("rarity"),
                "finish": r.finish,
                "quantity": r.quantity,
                "unit_price": _row_unit_price(r),
                "line_value": _row_line_value(r),
            }
            for r in rows
        ]
        json.dump(out, sys.stdout, indent=2); sys.stdout.write("\n")
        return
    typer.echo(f"# selector: {selector}", err=True)
    typer.echo(f"# rows: {len(rows)}", err=True)
    if not rows:
        raise typer.Exit(1)
    capped = rows[:first] if first else rows
    typer.echo(f"{'qty':>4} {'finish':>7} {'set':>6} {'cn':>6} {'rarity':>9}  name (usd / line)")
    for r in capped:
        unit = _row_unit_price(r); line = _row_line_value(r)
        usd = f"${unit:.2f}" if unit is not None else "—"
        line_s = f"${line:.2f}" if line is not None else "—"
        typer.echo(f"{r.quantity:>4} {r.finish:>7} {r.card.get('set',''):>6} "
                   f"{r.card.get('collector_number',''):>6} {(r.card.get('rarity') or ''):>9}  "
                   f"{_row_display_name(r)} ({usd} / {line_s})")
    if first and len(rows) > first:
        typer.echo(f"# truncated to first {first}; total {len(rows)}", err=True)


@query_app.command("value")
def query_value_cmd(
    selector: str = typer.Argument(...),
    json_out: bool = typer.Option(False, "--json"),
):
    """Total USD value of a selector's rows."""
    rows = _materialize_or_die(selector)
    total = 0.0
    missing = []
    priced_rows: list[tuple[float, sel_mod.MaterializedRow]] = []
    for r in rows:
        line = _row_line_value(r)
        if line is None and r.quantity > 0:
            missing.append((_row_display_name(r), r.card.get("set"), r.card.get("collector_number"), r.finish))
        else:
            total += line or 0.0
            if line is not None:
                priced_rows.append((line, r))
    priced_rows.sort(key=lambda t: t[0], reverse=True)
    top_5 = [
        {"name": _row_display_name(r), "set": r.card.get("set"),
         "cn": r.card.get("collector_number"), "finish": r.finish, "line_value": v}
        for v, r in priced_rows[:5]
    ]
    if json_out:
        json.dump({"selector": selector, "total": total, "rows": len(rows),
                   "missing_price": [{"name": n, "set": s, "cn": c, "finish": f}
                                     for n, s, c, f in missing],
                   "top_5": top_5}, sys.stdout, indent=2)
        sys.stdout.write("\n")
        return
    typer.echo(f"Selector {selector!r}: ${total:.2f} across {len(rows)} rows")
    if missing:
        typer.echo(f"  ({len(missing)} rows have no USD price)")
    if top_5:
        typer.echo("Top 5 by line value:")
        for t in top_5:
            typer.echo(f"  ${t['line_value']:.2f}  {t['name']} ({(t['set'] or '').upper()}) {t['cn']} [{t['finish']}]")


@query_app.command("top")
def query_top_cmd(
    n: int = typer.Argument(10, help="Show top-N rows by line value."),
):
    """Top-N inventory rows by line value (shorthand for `mm query show inventory` sorted)."""
    rows = _materialize_or_die("inventory")
    priced = [(r, _row_line_value(r)) for r in rows]
    priced = [(r, v) for r, v in priced if v is not None]
    priced.sort(key=lambda t: t[1], reverse=True)
    capped = priced[:n]
    typer.echo(f"Top {len(capped)} inventory rows by line value:")
    for r, line in capped:
        typer.echo(f"  ${line:.2f}  {_row_display_name(r)} "
                   f"({(r.card.get('set') or '').upper()}) {r.card.get('collector_number')} "
                   f"[{r.finish}] qty={r.quantity}")


@query_app.command("total")
def query_total_cmd():
    """Shorthand for `mm query value inventory`."""
    query_value_cmd(selector="inventory", json_out=False)


@query_app.command("multiples")
def query_multiples_cmd():
    """Inventory rows with quantity > 1, ordered by qty desc."""
    rows = _materialize_or_die("inventory qty>=2")
    rows.sort(key=lambda r: r.quantity, reverse=True)
    if not rows:
        typer.echo("(no multiples in inventory)"); return
    typer.echo(f"{'qty':>4} {'finish':>7} {'set':>6} {'cn':>6}  name")
    for r in rows:
        typer.echo(f"{r.quantity:>4} {r.finish:>7} {(r.card.get('set') or ''):>6} "
                   f"{r.card.get('collector_number',''):>6}  {_row_display_name(r)}")


@query_app.command("stats")
def query_stats_cmd(
    json_out: bool = typer.Option(False, "--json"),
):
    """Inventory rollup: totals, by-rarity, by-set, by-finish."""
    rows = _materialize_or_die("inventory")
    total = 0.0
    by_rarity: dict[str, dict] = {}
    by_set: dict[str, dict] = {}
    by_finish: dict[str, dict] = {}
    for r in rows:
        line = _row_line_value(r) or 0.0
        total += line
        rar = (r.card.get("rarity") or "unknown").lower()
        st = (r.card.get("set") or "unknown").lower()
        bucket = by_rarity.setdefault(rar, {"rows": 0, "qty": 0, "value": 0.0})
        bucket["rows"] += 1; bucket["qty"] += r.quantity; bucket["value"] += line
        bucket = by_set.setdefault(st, {"rows": 0, "qty": 0, "value": 0.0})
        bucket["rows"] += 1; bucket["qty"] += r.quantity; bucket["value"] += line
        bucket = by_finish.setdefault(r.finish, {"rows": 0, "qty": 0, "value": 0.0})
        bucket["rows"] += 1; bucket["qty"] += r.quantity; bucket["value"] += line
    out = {"total": total, "rows": len(rows),
           "qty": sum(r.quantity for r in rows),
           "by_rarity": by_rarity, "by_set": by_set, "by_finish": by_finish}
    if json_out:
        json.dump(out, sys.stdout, indent=2); sys.stdout.write("\n"); return
    typer.echo(f"Total: ${total:.2f} / {out['qty']} cards / {len(rows)} rows")
    typer.echo("\nBy rarity:")
    for rar, b in sorted(by_rarity.items()):
        typer.echo(f"  {rar:10} rows={b['rows']:>4} qty={b['qty']:>4} value=${b['value']:.2f}")
    typer.echo("\nBy set:")
    for st, b in sorted(by_set.items()):
        typer.echo(f"  {st.upper():10} rows={b['rows']:>4} qty={b['qty']:>4} value=${b['value']:.2f}")
    typer.echo("\nBy finish:")
    for fin, b in sorted(by_finish.items()):
        typer.echo(f"  {fin:10} rows={b['rows']:>4} qty={b['qty']:>4} value=${b['value']:.2f}")


@query_app.command("url")
def query_url_cmd(
    selector: str = typer.Argument(...),
    chunk_size: int = typer.Option(20, "--chunk-size",
                                   help="Cards per Scryfall search URL chunk (default 20; Scryfall web UI caps at 20 nested OR conditions)."),
    mode: str = typer.Option("oracle", "--mode",
                             help="'oracle' = !\"<name>\" form, dedupe by oracle name (good for shopping by name). "
                                  "'prints' = (set:CODE cn:\"CN\") form with unique=prints, one entry per printing "
                                  "(good for set-completion / 'which exact printing am I missing')."),
    sort: str = typer.Option("default", "--sort",
                             help="Sort order applied before chunking. default (set,cn,finish) | value-desc | value-asc | rarity. "
                                  "Use value-asc with --mode prints for cheapest-first set-completion URLs."),
):
    """Synthesize Scryfall search URLs for the result of a selector.

    Two modes:

    \b
    - oracle (default): emits `!"<oracle name>"` ORed terms, deduped by oracle
      name. Multiple finishes / printings of the same card collapse to one URL
      term. Best for shopping by name (let Scryfall show every printing so you
      can pick the cheapest).
    - prints: emits `(set:CODE cn:"CN")` ORed terms, one per distinct printing
      from the selector results, with `unique=prints&order=usd&dir=asc` appended
      to the URL so Scryfall returns each printing as a separate result sorted
      cheapest-first. Best for set-completion / "which exact printing am I
      missing" workflows. Honors `cn:"..."` quoting so hyphenated CNs (PMEI
      2025-13) and A-prefix variants (FIN A-248) work correctly.

    URLs are chunked at --chunk-size (default 20). Scryfall's web UI caps OR'd
    queries at 20 nested conditions; chunks larger than 20 will fail in the
    browser even if the API accepts them.
    """
    from urllib.parse import quote_plus
    rows = _materialize_or_die(selector)
    rows = _apply_sort(rows, sort)
    if not rows:
        typer.echo("(selector matched 0 rows)"); raise typer.Exit(1)

    if mode not in ("oracle", "prints"):
        typer.echo(f"error: --mode must be 'oracle' or 'prints', got {mode!r}", err=True)
        raise typer.Exit(2)

    if mode == "oracle":
        # Dedupe by oracle name — multiple finishes / printings of the same card
        # collapse to one URL term.
        names: list[str] = []
        seen: set[str] = set()
        for r in rows:
            nm = r.card.get("name")
            if nm and nm not in seen:
                seen.add(nm); names.append(nm)
        chunks = [names[i:i+chunk_size] for i in range(0, len(names), chunk_size)]
        typer.echo(f"{len(names)} distinct cards → {len(chunks)} URL(s) (mode=oracle)")
        for i, chunk in enumerate(chunks, start=1):
            terms = " or ".join(f'!"{nm}"' for nm in chunk)
            url = f"https://scryfall.com/search?q={quote_plus(terms)}"
            typer.echo(f"Chunk {i}/{len(chunks)} ({len(chunk)} cards): {url}")
        return

    # mode == "prints"
    # Collapse to one entry per (set, cn) — within a printing, multiple finishes
    # are the same Scryfall card. Preserve sort order from _apply_sort.
    seen_printings: set[tuple[str, str]] = set()
    printings: list[tuple[str, str]] = []  # [(set_code, cn), ...]
    for r in rows:
        setc = r.card.get("set") or ""
        cn = r.card.get("collector_number") or ""
        if not setc or not cn:
            continue
        key = (setc, cn)
        if key in seen_printings:
            continue
        seen_printings.add(key)
        printings.append(key)

    chunks = [printings[i:i+chunk_size] for i in range(0, len(printings), chunk_size)]
    typer.echo(f"{len(printings)} distinct printings → {len(chunks)} URL(s) (mode=prints)")
    for i, chunk in enumerate(chunks, start=1):
        terms = " or ".join(f'(set:{s} cn:"{cn}")' for s, cn in chunk)
        url = f"https://scryfall.com/search?q={quote_plus(terms)}&unique=prints&order=usd&dir=asc"
        typer.echo(f"Chunk {i}/{len(chunks)} ({len(chunk)} printings): {url}")


def _write_query_xlsx(
    rows: list[sel_mod.MaterializedRow],
    target: Path,
    selector: str,
    slug: str,
    kind: str = "query",
) -> None:
    """Write a list of materialized rows to an XLSX checklist artifact.

    Shared by `mm query xlsx` (kind="query", ad-hoc selector results) and
    `mm query missing-set` (kind="missing", canonical missing-checklist
    output). The `kind` field is recorded in the hidden `_meta` sheet so
    consumers can distinguish missing checklists from ad-hoc query results
    or from `mm set master-list`'s inventory checklists (which are written
    by a separate function in sets.py with `kind: "inventory"`).

    Columns: set, collector_number, name, rarity, finish, qty, unit_usd,
    line_value. The `name` cell is hyperlinked to the card's Scryfall page
    (same convention as the inventory checklist) so the user can click
    through to the printing without copying a UUID. Hidden _meta sheet
    records the originating selector and timestamp.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    target.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "results"
    headers = ["set", "collector_number", "name", "rarity", "finish",
               "qty", "unit_usd", "line_value"]
    ws.append(headers)
    for col, _ in enumerate(headers, start=1):
        ws.cell(row=1, column=col).font = Font(bold=True)
    # Match master-list's hyperlink styling: blue + underline mimics web links.
    link_font = Font(color="0563C1", underline="single")
    for r in rows:
        unit = _row_unit_price(r); line = _row_line_value(r)
        ws.append([
            r.card.get("set"), r.card.get("collector_number"),
            _row_display_name(r), r.card.get("rarity"), r.finish,
            r.quantity, unit, line,
        ])
        # Force CN to text to avoid Excel's "Number Stored as Text" warning.
        ws.cell(row=ws.max_row, column=2).number_format = "@"
        uri = r.card.get("scryfall_uri")
        if uri:
            name_cell = ws.cell(row=ws.max_row, column=3)
            name_cell.hyperlink = uri
            name_cell.font = link_font
    last = ws.max_row
    for col_idx in (7, 8):
        for row_idx in range(2, last + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = '"$"#,##0.00'
    widths = {1: 6, 2: 8, 3: 48, 4: 10, 5: 8, 6: 5, 7: 9, 8: 11}
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A2"

    meta_ws = wb.create_sheet("_meta")
    meta_ws.sheet_state = "hidden"
    meta_ws.append(["key", "value"])
    meta_ws.append(["kind", kind])
    meta_ws.append(["selector", selector])
    meta_ws.append(["slug", slug])
    meta_ws.append(["generated_at", datetime.now().isoformat(timespec="seconds")])
    meta_ws.append(["row_count", str(len(rows))])

    for _ws in wb.worksheets:
        util.apply_base_font_size(_ws)
    wb.save(target)


@query_app.command("xlsx")
def query_xlsx_cmd(
    selector: str = typer.Argument(...),
    name: str = typer.Option(None, "--name", help="Override the slug for the filename."),
    out: Path = typer.Option(None, "--out", help="Override the full output path."),
    sort: str = typer.Option("default", "--sort",
        help="Sort order: default (set,cn,finish) | value-desc | value-asc | rarity."),
):
    """Write the selector's rows to a queries/<slug>-<timestamp>.xlsx artifact.

    The XLSX has columns: set, collector_number, name, rarity, finish, qty,
    unit_usd, line_value. A hidden _meta sheet records the selector verbatim.
    Empty result still writes a file (with headers + empty body) and warns.
    """
    rows = _materialize_or_die(selector)
    rows = _apply_sort(rows, sort)
    slug = name or _slug(selector)
    ts = datetime.now().strftime("%Y-%m-%d-%H%M%S")
    target = out if out else QUERIES_DIR / f"{slug}-{ts}.xlsx"
    _write_query_xlsx(rows, target, selector, slug)
    if not rows:
        typer.echo(f"warning: selector matched 0 rows; wrote empty file {target}", err=True)
    typer.echo(f"wrote {target} ({len(rows)} rows)")


@query_app.command("missing-set")
def query_missing_set_cmd(
    code: str = typer.Argument(
        ...,
        help="Anchor set code (e.g. 'fin', 'avatar', 'tmnt'). Resolves +related family automatically.",
    ),
    chunk_size: int = typer.Option(
        20, "--chunk-size",
        help="Cards per Scryfall URL chunk (default 20; matches Scryfall web UI's nested-conditions cap).",
    ),
    treatment_class: str = typer.Option(
        "preferred", "--treatment-class",
        help="Treatment class for the alt sub-selector. Default 'preferred' (collectible-alt minus "
             "datestamped-with-sibling and family-configured fancy-foil dupes). "
             "Pass 'collectible-alt' to skip the dupe filtering, 'alt' to also include pure-ff, "
             "'any-alt' to also include ext.",
    ),
):
    """Canonical "what am I missing from set <CODE>?" workflow.

    Materializes the union of three sub-selectors (rare-regular, mythic-regular,
    treatment-class) printing-level, then emits:

    \b
    1. Scryfall printing-specific URL chunks (markdown table → STDOUT for chat).
       Sorted cheapest-first; uses (set:CODE cn:"CN") form with unique=prints
       and order=usd&dir=asc so each URL renders the EXACT missing printings.
    2. XLSX checklist (set-grouped, sorted by CN within each set) → queries/.
    3. ManaPool bulk-add .txt (flat list, *F* foil markers per line) → queries/.
    4. TCGplayer Mass Entry .txt (flat list, no per-line foil marker — user runs
       TCGplayer's cart optimizer to select foil/nonfoil per row) → queries/.

    The chat output is always just the URL table + file:// link lines so the
    user can click to open the artifacts. The bulk-add files are NEVER rendered
    inline — the user explicitly wants them as files only, and they must be
    paste-ready (no comments, headers, or fences) since portals don't tolerate
    extra characters.

    When --treatment-class=preferred (default), the rare/mythic regular-treatment
    sub-selectors are ALSO post-filtered to drop datestamped reprints that have
    a non-stamped sibling at the same treatment in the family. This catches
    e.g. PFIN's prerelease-stamped versions of FIN cards that are otherwise
    visually identical.

    New families with no FAMILY_DUPE_FOIL_PROMO_TYPES config will fail with a
    clear error from the selector layer. Either configure the family or pass
    --treatment-class collectible-alt to opt into the looser pre-`preferred`
    behavior.

    Set-agnostic: works for FIN today, Avatar/TMNT/etc. tomorrow with the same
    invocation, once each new family is configured.
    """
    import re as _re
    import json as _json
    from urllib.parse import quote_plus as _quote_plus

    from . import missing as missing_mod

    code_l = code.lower()

    # 1. Materialize the printing-level union of the missing-set sub-selectors.
    # Shared with scripts/manapool_cart_check.py via magic_manager.missing.
    try:
        rows_union = missing_mod.missing_printings(code_l, treatment_class)
    except sel_mod.SelectorParseError as e:
        typer.echo(f"error: invalid selector: {e}", err=True); raise typer.Exit(2)
    except LookupError as e:
        typer.echo(f"error: {e}", err=True); raise typer.Exit(2)

    if not rows_union:
        typer.echo(f"# No missing printings found for set:{code_l}+related (full collection? wrong code?).")
        raise typer.Exit(0)

    # 2. Cheapest-first ordering for the Scryfall URL chunks.
    rows_by_value = sorted(rows_union, key=lambda r: (
        0 if _row_line_value(r) is not None else 1,
        _row_line_value(r) or 0.0,
        r.card.get("set") or "", util.cn_sort_key(r.card.get("collector_number")),
    ))
    chunks = [rows_by_value[i:i+chunk_size] for i in range(0, len(rows_by_value), chunk_size)]

    # 3. Build the XLSX checklist artifact (grouped by set, sorted by CN within each).
    rows_for_xlsx = sorted(rows_union, key=lambda r: (
        r.card.get("set") or "",
        util.cn_sort_key(r.card.get("collector_number")),
        r.finish,
    ))
    ts = datetime.now().strftime("%Y-%m-%d-%H%M%S")
    QUERIES_DIR.mkdir(parents=True, exist_ok=True)
    xlsx_path = QUERIES_DIR / f"missing-{code_l}-checklist-{ts}.xlsx"
    _subs = missing_mod.sub_selectors(code_l, treatment_class)
    union_selector_repr = (
        f"({_subs[0][1]}) ∪ ({_subs[1][1]}) ∪ ({_subs[2][1]})  [printing-level union]"
    )
    _write_query_xlsx(
        rows_for_xlsx, xlsx_path, union_selector_repr,
        f"missing-{code_l}-checklist", kind="missing",
    )

    # 4. Build the bulk-add artifacts.
    #    All three are plain text — no headings, no comments, no fences. Pasting
    #    into a portal's mass-entry box must succeed without any pre-edit.
    #    Order rows by (set, cn, finish) within each file for predictability.
    rows_for_bulk = sorted(rows_union, key=lambda r: (
        r.card.get("set") or "",
        util.cn_sort_key(r.card.get("collector_number")),
        r.finish,
    ))
    total_value = sum((_row_line_value(r) or 0.0) for r in rows_union)

    # ManaPool: single flat list, *F* foil markers preserved per-line.
    mp_path = QUERIES_DIR / f"missing-{code_l}-manapool-{ts}.txt"
    mp_path.write_text(exports.build("manapool", rows_for_bulk), encoding="utf-8")

    # TCGplayer: single flat list. Foil/nonfoil isn't marked per-line — the
    # user runs TCGplayer's cart optimizer afterward to pick finish per row.
    tcg_path = QUERIES_DIR / f"missing-{code_l}-tcgplayer-{ts}.txt"
    tcg_path.write_text(exports.build("tcgplayer", rows_for_bulk), encoding="utf-8")

    # 5. Emit chat output: URL table + file:// links. Nothing else.
    typer.echo(f"# Missing from set:{code_l}+related — {len(rows_union)} distinct printings · ${total_value:,.2f}")
    typer.echo(f"")
    typer.echo(f"## Scryfall URLs ({len(chunks)} chunks, cheapest first)")
    typer.echo(f"")
    typer.echo(f"| # | Printings | Price band | URL |")
    typer.echo(f"|---:|---:|---|---|")
    for i, chunk in enumerate(chunks, start=1):
        cheap = _row_line_value(chunk[0])
        most = _row_line_value(chunk[-1])
        cs = f"${cheap:.2f}" if cheap is not None else "—"
        ms = f"${most:.2f}" if most is not None else "—"
        terms = " or ".join(
            f'(set:{r.card.get("set")} cn:"{r.card.get("collector_number")}")' for r in chunk
        )
        url = f"https://scryfall.com/search?q={_quote_plus(terms)}&unique=prints&order=usd&dir=asc"
        typer.echo(f"| {i} | {len(chunk)} | {cs} → {ms} | [chunk {i}]({url}) |")
    typer.echo(f"")
    typer.echo(f"📋 Checklist (xlsx): [{xlsx_path}](file://{xlsx_path.resolve()})")
    typer.echo(f"🛒 ManaPool bulk-add ({len(rows_for_bulk)} rows): [{mp_path}](file://{mp_path.resolve()})")
    typer.echo(f"🛒 TCGplayer Mass Entry ({len(rows_for_bulk)} rows): [{tcg_path}](file://{tcg_path.resolve()})")


@query_app.command("missing-jumpstart")
def query_missing_jumpstart_cmd(
    code: str = typer.Argument(
        ...,
        help="Jumpstart set code (e.g. 'j25', 'msh', 'tle'). Must publish "
             "Jumpstart variants in MTGJSON.",
    ),
):
    """Buy list for the Jumpstart packs you don't own from a set.

    A pack is "missing" when you have no ``pack:<theme>-<code>`` deck for it
    (i.e. you never opened/ingested it). For every such pack this emits the
    full singles list — gameplay cards plus the pack's front/title card — as
    three combined artifacts under ``queries/``: an XLSX checklist, a ManaPool
    bulk-add ``.txt``, and a TCGplayer Mass Entry ``.txt``.

    Contents are NOT deduped across packs and NOT reduced by cards you already
    own — it lists each un-owned pack's full contents. (A future workflow will
    shrink the list by building packs from free inventory.)
    """
    code_l = code.lower()

    # Sync the family (gameplay scryfall_ids) + front cards, mirroring
    # jumpstart-list / jumpstart-pack.
    try:
        _r, codes = _resolve_codes(code, include_kinds=[], only=[])
    except (LookupError, typer.BadParameter) as e:
        typer.echo(f"error: {e}", err=True)
        raise typer.Exit(2)
    sets_mod.sync(codes)
    front_cards_mod.sync_front_cards(code_l)

    variants = mtgjson_mod.jumpstart_variants(code_l)
    if not variants:
        typer.echo(
            f"error: no Jumpstart variants found for set {code!r}. "
            f"Check `mm mtgjson decks --set {code_l}` for available decks.",
            err=True,
        )
        raise typer.Exit(2)

    # Diff against owned pack:* decks. Copy suffixes (-2/-3) share the base
    # slug, so an owned base slug counts the theme as owned.
    with db.connect() as conn:
        owned_slugs = {
            row["slug"]
            for row in conn.execute(
                "SELECT slug FROM decks WHERE slug LIKE ?", (f"pack:%-{code_l}",)
            ).fetchall()
        }
    missing_variants = [
        v for v in variants
        if sets_mod._slug_theme(v.get("name") or v.get("fileName") or "", code_l)
        not in owned_slugs
    ]
    missing_variants.sort(key=lambda v: (v.get("name") or v.get("fileName") or ""))

    if not missing_variants:
        typer.echo(
            f"# You own all {len(variants)} Jumpstart pack(s) for set:{code_l}. Nothing missing."
        )
        raise typer.Exit(0)

    # Build combined rows across every missing pack (no cross-pack dedup —
    # each pack's buy list is its full contents). Track a per-pack summary.
    all_rows: list[sel_mod.MaterializedRow] = []
    pack_summaries: list[tuple[str, int, float]] = []  # (theme, card_count, usd_total)
    total_skipped = 0
    for v in missing_variants:
        rows, n_skipped = _jumpstart_pack_rows(code_l, v)
        total_skipped += n_skipped
        pack_usd = sum((_row_line_value(r) or 0.0) for r in rows)
        pack_summaries.append((v.get("name") or v.get("fileName") or "?", len(rows), pack_usd))
        all_rows.extend(rows)

    if not all_rows:
        typer.echo(
            f"# {len(missing_variants)} missing pack(s) for set:{code_l}, but no cards "
            f"resolved locally (sync issue?).",
            err=True,
        )
        raise typer.Exit(2)

    ts = datetime.now().strftime("%Y-%m-%d-%H%M%S")
    QUERIES_DIR.mkdir(parents=True, exist_ok=True)

    # Artifacts are named `missing-jumpstart-<code>-*` so they never collide
    # with `query missing-set`'s `missing-<code>-*` files.
    rows_for_xlsx = sorted(all_rows, key=lambda r: (
        r.card.get("set") or "",
        util.cn_sort_key(r.card.get("collector_number")),
        r.finish,
    ))
    xlsx_path = QUERIES_DIR / f"missing-jumpstart-{code_l}-checklist-{ts}.xlsx"
    _write_query_xlsx(
        rows_for_xlsx, xlsx_path,
        f"jumpstart-missing:{code_l} ({len(missing_variants)} un-owned packs)",
        f"missing-jumpstart-{code_l}-checklist", kind="missing",
    )

    mp_path = QUERIES_DIR / f"missing-jumpstart-{code_l}-manapool-{ts}.txt"
    mp_path.write_text(exports.build("manapool", rows_for_xlsx), encoding="utf-8")

    tcg_path = QUERIES_DIR / f"missing-jumpstart-{code_l}-tcgplayer-{ts}.txt"
    tcg_path.write_text(exports.build("tcgplayer", rows_for_xlsx), encoding="utf-8")

    total_value = sum((_row_line_value(r) or 0.0) for r in all_rows)
    typer.echo(
        f"# Missing Jumpstart packs for set:{code_l} — "
        f"{len(missing_variants)} of {len(variants)} packs · "
        f"{len(all_rows)} card-rows · ${total_value:,.2f}"
    )
    typer.echo("")
    typer.echo("| Pack | Cards | Value |")
    typer.echo("|---|---:|---:|")
    for theme, n_cards, usd in pack_summaries:
        typer.echo(f"| {theme} | {n_cards} | ${usd:,.2f} |")
    typer.echo("")
    if total_skipped:
        typer.echo(f"> ⚠️ {total_skipped} card(s) not found locally, skipped.")
        typer.echo("")
    typer.echo(f"📋 Checklist (xlsx): [{xlsx_path}](file://{xlsx_path.resolve()})")
    typer.echo(f"🛒 ManaPool bulk-add ({len(rows_for_xlsx)} rows): [{mp_path}](file://{mp_path.resolve()})")
    typer.echo(f"🛒 TCGplayer Mass Entry ({len(rows_for_xlsx)} rows): [{tcg_path}](file://{tcg_path.resolve()})")


# ---------- ad-hoc scryfall query ----------

@app.command("scryfall")
def scryfall_cmd(
    query: str = typer.Argument(..., help="Scryfall search query (any syntax the API accepts)."),
    first: int = typer.Option(20, "--first", help="Show at most N results."),
    json_out: bool = typer.Option(False, "--json", help="Emit raw Scryfall JSON instead of the table."),
    fields: str = typer.Option(
        "set,collector_number,name,treatment,rarity",
        "--fields",
        help="Comma-separated columns. Available: set,collector_number,name,rarity,"
             "treatment,full_art,border_color,frame_effects,promo_types,security_stamp,"
             "prices_usd,prices_usd_foil,scryfall_uri",
    ),
):
    """Run an ad-hoc Scryfall search and pretty-print the results.

    Avoids the shell-quoting trap of writing one-shot Python at the prompt.
    Uses the rate-limited wrapper (``scryfall.sh``) under the hood, so
    multi-page queries paginate cleanly. Card-name apostrophes are passed
    through without escaping issues.

    Each row's computed treatment string (per ``treatments.compute_treatment``)
    is included by default, so distinct printings of the same card
    (e.g. Cloud, Ex-SOLDIER variants) are immediately visually distinguishable.
    """
    from .scryfall import search as sf_search
    from .treatments import compute_treatment

    cols = [c.strip().lower() for c in fields.split(",") if c.strip()]
    rows: list[dict] = []
    try:
        for i, c in enumerate(sf_search(query, unique="prints")):
            if i >= first:
                break
            rows.append(c)
    except Exception as e:
        typer.echo(f"error: {e}", err=True)
        raise typer.Exit(2)

    if not rows:
        typer.echo(f"(no results for {query!r})", err=True)
        raise typer.Exit(1)

    if json_out:
        json.dump(rows, sys.stdout, indent=2, default=str)
        sys.stdout.write("\n")
        return

    def get(c: dict, col: str):
        if col == "treatment":
            return compute_treatment(c) or "—"
        if col == "frame_effects":
            return ",".join(c.get("frame_effects") or []) or "—"
        if col == "promo_types":
            return ",".join(c.get("promo_types") or []) or "—"
        if col == "prices_usd":
            return (c.get("prices") or {}).get("usd") or "—"
        if col == "prices_usd_foil":
            return (c.get("prices") or {}).get("usd_foil") or "—"
        if col == "scryfall_uri":
            return c.get("scryfall_uri") or "—"
        if col == "full_art":
            return "yes" if c.get("full_art") else "no"
        v = c.get(col, "—")
        return v if v not in (None, "") else "—"

    # Compute column widths for a tight table.
    widths = {col: max(len(col), max(len(str(get(r, col))) for r in rows)) for col in cols}
    header = "  ".join(col.ljust(widths[col]) for col in cols)
    typer.echo(header)
    typer.echo("  ".join("-" * widths[col] for col in cols))
    for r in rows:
        typer.echo("  ".join(str(get(r, col)).ljust(widths[col]) for col in cols))
    typer.echo("", err=True)
    typer.echo(f"# {len(rows)} result(s) for {query!r}", err=True)


# ---------- mtgjson (precon decks + set data) ----------

@mtgjson_app.command("meta")
def mtgjson_meta(
    json_out: bool = typer.Option(False, "--json", help="Emit raw JSON."),
):
    """Show MTGJSON's current build date and version."""
    from . import mtgjson as mj
    m = mj.meta()
    if json_out:
        json.dump(m, sys.stdout, indent=2)
        sys.stdout.write("\n")
        return
    typer.echo(f"date:    {m.get('date')}")
    typer.echo(f"version: {m.get('version')}")


@mtgjson_app.command("set")
def mtgjson_set(
    set_code: str = typer.Argument(..., help="Set code, e.g. fic, FIC."),
    json_out: bool = typer.Option(False, "--json", help="Emit raw set JSON."),
):
    """Pretty-print MTGJSON's per-set summary."""
    from . import mtgjson as mj
    s = mj.set_file(set_code)
    if json_out:
        json.dump(s, sys.stdout, indent=2, default=str)
        sys.stdout.write("\n")
        return
    typer.echo(f"name:           {s.get('name')}")
    typer.echo(f"code:           {s.get('code')}")
    typer.echo(f"type:           {s.get('type')}")
    typer.echo(f"releaseDate:    {s.get('releaseDate')}")
    typer.echo(f"totalSetSize:   {s.get('totalSetSize')}")
    typer.echo(f"baseSetSize:    {s.get('baseSetSize')}")
    cards = s.get("cards") or []
    tokens = s.get("tokens") or []
    decks = s.get("decks") or []
    typer.echo(f"cards:          {len(cards)}")
    typer.echo(f"tokens:         {len(tokens)}")
    typer.echo(f"decks (inline): {len(decks)}")


@mtgjson_app.command("decks")
def mtgjson_decks(
    set_code: str = typer.Option(None, "--set", help="Filter to one set code."),
    first: int = typer.Option(50, "--first", help="Show at most N rows."),
    json_out: bool = typer.Option(False, "--json", help="Emit raw DeckList JSON."),
):
    """List decks (filterable by set code) from MTGJSON's DeckList.json."""
    from . import mtgjson as mj
    rows = mj.deck_list(set_code=set_code)
    if json_out:
        json.dump(rows, sys.stdout, indent=2, default=str)
        sys.stdout.write("\n")
        return
    if not rows:
        typer.echo(f"(no decks{f' for set {set_code.upper()}' if set_code else ''})", err=True)
        raise typer.Exit(1)
    rows = rows[:first]
    cols = ("code", "fileName", "name", "type", "releaseDate")
    widths = {c: max(len(c), max(len(str(r.get(c) or "")) for r in rows)) for c in cols}
    typer.echo("  ".join(c.ljust(widths[c]) for c in cols))
    typer.echo("  ".join("-" * widths[c] for c in cols))
    for r in rows:
        typer.echo("  ".join(str(r.get(c) or "—").ljust(widths[c]) for c in cols))


@mtgjson_app.command("deck")
def mtgjson_deck(
    file_name: str = typer.Argument(..., help="MTGJSON deck fileName, e.g. CounterBlitzFinalFantasyX_FIC."),
    json_out: bool = typer.Option(False, "--json", help="Emit raw deck JSON."),
    show: int = typer.Option(10, "--show", help="Show at most N cards per board in summary view."),
):
    """Pretty-print one MTGJSON deck file (commander, mainBoard, sideBoard)."""
    from . import mtgjson as mj
    d = mj.deck(file_name)
    if json_out:
        json.dump(d, sys.stdout, indent=2, default=str)
        sys.stdout.write("\n")
        return
    typer.echo(f"name:        {d.get('name')}")
    typer.echo(f"code:        {d.get('code')}")
    typer.echo(f"type:        {d.get('type')}")
    typer.echo(f"releaseDate: {d.get('releaseDate')}")

    def emit(board: str):
        cards = d.get(board) or []
        if not cards:
            return
        typer.echo(f"\n{board} ({len(cards)} {'entry' if len(cards) == 1 else 'entries'}):")
        for c in cards[:show]:
            finish = "foil" if c.get("isFoil") else "nonfoil"
            sid = (c.get("identifiers") or {}).get("scryfallId", "—")
            typer.echo(
                f"  {c.get('count', 1):>2}x  ({c.get('setCode')}) {str(c.get('number')):>5}  "
                f"{(c.get('name') or '')[:40]:40}  {finish:7}  scryfall:{sid}"
            )
        if len(cards) > show:
            typer.echo(f"  … {len(cards) - show} more")

    emit("commander")
    emit("mainBoard")
    emit("sideBoard")
    emit("tokens")


@mtgjson_app.command("refresh")
def mtgjson_refresh(
    resource_path: str = typer.Argument(..., help="Resource path, e.g. FIC.json or DeckList.json."),
):
    """Delete the cached copy of ``resource_path`` so the next fetch re-downloads."""
    from . import mtgjson as mj
    mj.refresh(resource_path)
    typer.echo(f"refreshed: {resource_path}")


@mtgjson_app.command("check-stale")
def mtgjson_check_stale(
    resource_path: str = typer.Argument(..., help="Resource path, e.g. FIC.json or DeckList.json."),
):
    """Compare cached SHA-256 to MTGJSON's published .sha256 sidecar.

    Exits 0 if fresh, 1 if stale, 2 if not cached.
    """
    from . import mtgjson as mj
    try:
        stale = mj.is_stale(resource_path)
    except mj.MtgJsonError as e:
        typer.echo(f"absent ({e})", err=True)
        raise typer.Exit(2)
    if stale:
        typer.echo("stale")
        raise typer.Exit(1)
    typer.echo("fresh")


# ---------- db (snapshots, restore, integrity) ----------

@db_app.command("snapshot")
def db_snapshot_cmd(
    label: str = typer.Option(None, "--label", help="Suffix appended to the backup filename."),
):
    """Take a timestamped snapshot of the active DB next to the live file."""
    backup = db.snapshot(label=label)
    typer.echo(str(backup))


@db_app.command("snapshots")
def db_snapshots_cmd():
    """List local DB snapshots (newest first)."""
    snaps = db.list_snapshots()
    if not snaps:
        typer.echo("(no snapshots)", err=True)
        raise typer.Exit(0)
    for p in snaps:
        st = p.stat()
        size_mb = st.st_size / (1024 * 1024)
        when = datetime.fromtimestamp(st.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
        typer.echo(f"{when}  {size_mb:>6.2f} MB  {p}")


@db_app.command("restore")
def db_restore_cmd(
    backup_path: Path = typer.Argument(..., help="Path to a snapshot file."),
):
    """Restore the active DB from a snapshot. Renames the current live DB to <live>.replaced-<ts>."""
    replaced = db.restore(backup_path)
    if replaced is not None:
        typer.echo(f"prior live DB moved to: {replaced}", err=True)
    typer.echo(f"restored from: {backup_path}")


@db_app.command("integrity")
def db_integrity_cmd():
    """Run PRAGMA integrity_check on the live DB. Exits non-zero if not 'ok'."""
    result = db._check_integrity(db.db_path())
    typer.echo(result)
    if result != "ok":
        raise typer.Exit(1)


@db_app.command("unlock")
def db_unlock_cmd(
    yes: bool = typer.Option(False, "--yes", "-y", help="Skip the confirmation prompt."),
):
    """Clear stale -wal/-shm sidecar files left by a run that died mid-write.

    Automates the ``rm -rf db/*.db-wal`` hand-fix. GUARDED: refuses if the DB is
    currently locked by a live process (a busy PRAGMA probe), so it can't corrupt
    an active write. Only removes the sidecars when the DB opens cleanly.
    """
    p = db.db_path()
    wal = p.with_name(p.name + "-wal")
    shm = p.with_name(p.name + "-shm")
    present = [f for f in (wal, shm) if f.exists()]
    if not present:
        typer.echo("No -wal/-shm sidecars present; nothing to unlock.")
        return

    # Guard: if another process holds the lock, a quick write probe raises
    # "database is locked" — refuse rather than risk clobbering a live write.
    import sqlite3
    try:
        probe = sqlite3.connect(str(p), timeout=0.5)
        probe.execute("BEGIN IMMEDIATE")
        probe.rollback()
        probe.close()
    except sqlite3.OperationalError as e:
        typer.echo(f"error: DB appears locked by another process ({e}); "
                   f"not touching sidecars.", err=True)
        raise typer.Exit(1)

    typer.echo("Stale sidecars to remove:")
    for f in present:
        typer.echo(f"  {f.name}")
    if not yes:
        typer.confirm("Remove them?", abort=True)
    # Checkpoint first so any durable WAL content folds into the main DB, then
    # the sidecars are safe to drop.
    with db.connect() as conn:
        conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
    for f in present:
        if f.exists():
            f.unlink()
    typer.echo(f"Removed {len(present)} sidecar file(s).")


@audit_app.command("deck-inventory")
def audit_deck_inventory_cmd(
    fix: bool = typer.Option(False, "--fix", help="Delete orphan (empty) decks found."),
):
    """Report DB consistency issues between decks and inventory; ``--fix`` repairs
    the safe ones.

    Detects:
      - **orphan decks** — deck rows with zero ``deck_cards`` (e.g. an
        interrupted import before the atomicity fix). ``--fix`` deletes these
        (cascades to any stray rows) — automates the manual ``deck delete``
        recovery.
      - **over-assignment** — printings whose ``deck_assignments`` total exceeds
        the owned inventory quantity (report-only; never auto-changed).
    """
    with db.connect() as conn:
        orphans = conn.execute(
            """
            SELECT d.slug, d.name
            FROM decks d
            LEFT JOIN deck_cards dc ON dc.deck_id = d.deck_id
            WHERE dc.deck_id IS NULL
            ORDER BY d.slug
            """
        ).fetchall()
        over = conn.execute(
            """
            SELECT da.scryfall_id, da.finish,
                   SUM(da.count) AS assigned,
                   COALESCE((SELECT quantity FROM inventory i
                             WHERE i.scryfall_id = da.scryfall_id
                               AND i.finish = da.finish), 0) AS owned
            FROM deck_assignments da
            GROUP BY da.scryfall_id, da.finish
            HAVING assigned > owned
            """
        ).fetchall()

    typer.echo(f"Orphan decks (0 cards): {len(orphans)}")
    for o in orphans:
        typer.echo(f"  {o['slug']}  ({o['name']})")
    typer.echo(f"Over-assigned printings (assigned > owned): {len(over)}")
    for o in over:
        typer.echo(f"  {o['scryfall_id']} [{o['finish']}]  assigned {o['assigned']} > owned {o['owned']}")

    if fix and orphans:
        for o in orphans:
            decks_mod.deck_delete(o["slug"])
        typer.echo(f"Fixed: deleted {len(orphans)} orphan deck(s).")
    elif orphans and not fix:
        typer.echo("Re-run with --fix to delete the orphan deck(s).")

    if over:
        # Over-assignment is a data-integrity smell but auto-fixing it (which
        # copies to keep?) needs human judgment — report only.
        raise typer.Exit(1)


# ---------- intake (scan-loop REPL) ----------

@app.command("intake")
def intake_cmd(
    name_or_code: str = typer.Argument(...),
    only: list[str] = typer.Option(None, "--only"),
    include: list[str] = typer.Option(None, "--include"),
):
    """Scan-loop REPL: type ``<set>? <cn> [+N|=N] [f]`` per card, qty updates live.

    Bound to a resolved set family. Writes directly to the V2 ``inventory``
    table — no master-list seeding required (any card synced via
    ``mm set sync`` is fair game). The first set code you type becomes
    sticky; subsequent lines without a set use it. Each entry is a separate
    DB transaction — Ctrl-C is safe.

    Modes per line:
      - bare              → +1 (default)
      - +N                → increment by N
      - =N                → overwrite to exactly N (requires N >= 0)
      - trailing f / foil → this card is foil

    Other commands: u/undo, s <code>/set <code>, ?/help, q/quit.
    """
    try:
        r, codes = _resolve_codes(
            name_or_code, include_kinds=list(include or []), only=list(only or []),
        )
    except (LookupError, typer.BadParameter) as e:
        typer.echo(f"error: {e}", err=True); raise typer.Exit(2)
    intake_mod.run_repl(r)


# ---------- export ----------

@app.command("export")
def export_cmd(
    target: str = typer.Argument(..., help="moxfield | manapool | tcgplayer | archidekt | plain | scryfall-json"),
    selector: str = typer.Argument(..., help="V2 selector, e.g. 'inventory', 'set:fca missing', 'wishlist:edh-staples'"),
    out: Path = typer.Option(None, "--out", help="Optional output path; otherwise prints to stdout."),
):
    """Materialize a V2 selector and emit a paste-ready block for the target service."""
    try:
        rows = sel_mod.materialize(selector)
    except sel_mod.SelectorParseError as e:
        typer.echo(f"error: invalid selector: {e}", err=True); raise typer.Exit(2)
    except LookupError as e:
        typer.echo(f"error: {e}", err=True); raise typer.Exit(2)
    if not rows:
        typer.echo(f"(selector matched 0 rows: {selector})", err=True)
        raise typer.Exit(1)
    text = exports.build(target, rows)

    typer.echo(f"# selector: {selector}", err=True)
    typer.echo(f"# target: {target}", err=True)
    typer.echo(f"# rows: {len(rows)}", err=True)
    if target == "tcgplayer":
        typer.echo("# NOTE: TCGplayer Mass Entry format is '1 Card Name [SETCODE] CN'.", err=True)
        typer.echo("#       Foil is set per-batch via the cart UI toggle, not per-line —", err=True)
        typer.echo("#       run twice with finish=nonfoil and finish=foil for a mixed cart.", err=True)

    if out:
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(text, encoding="utf-8")
        typer.echo(f"wrote {out}", err=True)
    else:
        typer.echo(text, nl=False)


# ---------- input/ inspection (used by the slash command) ----------

@checklists_app.command("list")
def input_list(
    json_out: bool = typer.Option(False, "--json", help="Emit machine-readable JSON."),
):
    """List every active inventory checklist in ``input/`` with a summary and
    duplicate-vs-prior-ingest flag.

    The ``/ingest-new-inventory-list`` slash command reads the JSON form of
    this output to drive its per-file conversation. Every file shows up
    even if its hash matches a prior successful ingest — those are flagged
    for the user to triage.
    """
    if not INPUT_DIR.exists():
        if json_out:
            json.dump({"input_dir": str(INPUT_DIR), "files": []}, sys.stdout, indent=2)
            sys.stdout.write("\n")
        else:
            typer.echo(f"(input dir {INPUT_DIR}/ does not exist)")
        return

    # Walk both supported formats. Don't recurse into processed/ — those are
    # immutable archives, not active intake docs. Skip Excel/LibreOffice
    # temp-lock sidecars (``~$<name>.xlsx``, ``.~lock.<name>#``) — they aren't
    # real workbooks; the user just has the file open in Excel.
    files: list[Path] = []
    skipped_lock_files: list[str] = []
    for pattern in ("*.xlsx", "*.md"):
        for p in INPUT_DIR.glob(pattern):
            if not p.is_file():
                continue
            if p.name.startswith("~$") or p.name.startswith(".~lock."):
                skipped_lock_files.append(p.name)
                continue
            files.append(p)
    files = sorted(files)
    if skipped_lock_files and not json_out:
        typer.echo(
            f"(skipped {len(skipped_lock_files)} Excel/LibreOffice lock file(s); "
            f"close the workbook in your editor if you intended to ingest it: "
            f"{', '.join(skipped_lock_files)})",
            err=True,
        )
    out_files: list[dict] = []
    for f in files:
        sha = _file_sha256(f)
        with db.connect() as conn:
            prior = db.find_ingest_log_by_hash(conn, sha)
        prior_success = next((p for p in prior if p["status"] == "success"), None)
        prior_failed = next((p for p in prior if p["status"] == "failed"), None)
        try:
            summary = sets_mod.summarize_intake_file(f)
        except Exception as e:  # malformed XLSX shouldn't crash the listing
            summary = {"error": repr(e)}
        out_files.append({
            "path": str(f),
            "name": f.name,
            "sha256": sha,
            "size_bytes": f.stat().st_size,
            "summary": summary,
            "duplicate_of_log_id": prior_success["id"] if prior_success else None,
            "prior_success": prior_success,
            "prior_failed": prior_failed,
        })

    if json_out:
        json.dump({"input_dir": str(INPUT_DIR), "files": out_files}, sys.stdout, indent=2, default=str)
        sys.stdout.write("\n")
        return

    if not out_files:
        typer.echo(f"(no XLSX files in {INPUT_DIR}/)"); return
    for f in out_files:
        typer.echo(f"- {f['name']} ({f['size_bytes']} bytes, sha={f['sha256'][:12]}…)")
        s = f["summary"]
        if "error" in s:
            typer.echo(f"    parse error: {s['error']}")
            continue
        if s.get("kind") in ("precon", "jumpstart"):
            # Deck checklist: qty_normal/qty_foil don't apply — report the
            # keep/deconstruct shape instead.
            noun = "packs" if s["kind"] == "jumpstart" else "decks"
            typer.echo(f"    kind={s['kind']} (deck checklist)")
            typer.echo(
                f"    {s['rows_with_qty']}/{s['rows_total']} {noun} filled — "
                f"{s.get('decks_to_construct', 0)} to construct, "
                f"{s.get('loose_copies', 0)} loose copies, "
                f"value=${s['estimated_value']:.2f}"
            )
        else:
            rarity = ",".join(s.get("rarity_filter") or []) or "(none)"
            codes = ",".join(s.get("set_codes") or []) or "(none)"
            typer.echo(
                f"    anchor={s.get('anchor_code') or '?'} "
                f"set_codes={codes} rarity_filter={rarity}"
            )
            typer.echo(
                f"    rows_total={s['rows_total']} "
                f"rows_with_qty={s['rows_with_qty']} "
                f"total_qty={s['total_qty']} "
                f"value=${s['estimated_value']:.2f}"
            )
        if f["duplicate_of_log_id"]:
            ps = f["prior_success"]
            typer.echo(
                f"    ⚠ DUPLICATE: SHA matches log id {ps['id']} ingested at "
                f"{ps['at']} (likely a failed cleanup; pass --force to re-apply)"
            )
        if f["prior_failed"]:
            pf = f["prior_failed"]
            typer.echo(f"    ⚠ previously failed at {pf['at']}: {pf['error']}")


# ---------- entry point ----------

if __name__ == "__main__":
    app()

"""Parse pasted card-list blocks (Moxfield, Archidekt, MTGA, MTGO, deckstats)
and filled-in master-list spreadsheets (XLSX) into resolved Scryfall cards.

The text parser handles the syntax that all the major deck-builders converge
on: ``<qty> <Card Name> [(<SET>) <CN>][★|*F*]`` plus section headers like
``SIDEBOARD:``. Section info is preserved as ``entry["section"]`` but most
callers will ignore it (a wishlist or set list has only a single section).
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from . import scryfall


SECTION_ALIASES = {
    "deck": "mainboard",
    "main": "mainboard",
    "main deck": "mainboard",
    "maindeck": "mainboard",
    "mainboard": "mainboard",
    "sideboard": "sideboard",
    "side board": "sideboard",
    "sb": "sideboard",
    "commander": "commander",
    "commanders": "commander",
    "companion": "companion",
    "maybeboard": "maybeboard",
    "maybe": "maybeboard",
}

SECTION_RE = re.compile(
    r"^\s*(?://\s*)?(" + "|".join(re.escape(k) for k in SECTION_ALIASES) + r")\s*:?\s*$",
    re.IGNORECASE,
)

# qty (with optional 'x'), name (non-greedy), optional (SET) CN, optional foil marker.
CARD_RE = re.compile(
    r"""
    ^\s*
    (?P<qty>\d+)x?
    \s+
    (?P<name>.+?)
    (?:\s+\((?P<set>[A-Za-z0-9]{2,6})\)\s+(?P<cn>[^\s★*]+))?
    (?:\s*(?P<foil>★|\*F\*))?
    \s*$
    """,
    re.VERBOSE,
)

IGNORE_RE = re.compile(
    r"""^\s*(?:
          \#.*
        | //(?!\s*(?:Main|Sideboard|Maybeboard|Commander|Companion))
        | About\b.*
        | Name\s+.*
        | Layout\s+.*
        )\s*$""",
    re.IGNORECASE | re.VERBOSE,
)


@dataclass
class Entry:
    qty: int
    raw: str
    name: str
    set: str | None
    collector_number: str | None
    foil: bool
    section: str = "mainboard"
    card: dict | None = None  # populated by resolve()


@dataclass
class ParseResult:
    entries: list[Entry] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    not_found: list[dict] = field(default_factory=list)
    # Populated by parse_master_list_xlsx when the file has a _meta sheet.
    # Keys: anchor_code, set_codes, rarity_filter, slug, include_tokens,
    # generated_at, magic_manager_version. Each value is a string; comma-lists
    # are NOT pre-split (callers split as needed). ``None`` if no _meta sheet.
    meta: dict | None = None

    @property
    def found_entries(self) -> list[Entry]:
        return [e for e in self.entries if e.card is not None]


# ---------- text parsing ----------

def parse_text(text: str) -> ParseResult:
    """Parse a Moxfield-style text block. Caller must run ``resolve()`` to fill
    in Scryfall data."""
    res = ParseResult()
    current = "mainboard"
    for raw_line in text.splitlines():
        line = raw_line.rstrip()
        if not line.strip():
            continue
        if IGNORE_RE.match(line):
            continue
        m = SECTION_RE.match(line)
        if m:
            current = SECTION_ALIASES[m.group(1).lower()]
            continue
        m = CARD_RE.match(line)
        if not m:
            res.warnings.append(f"unparseable line: {line!r}")
            continue
        res.entries.append(Entry(
            qty=int(m.group("qty")),
            raw=line,
            name=m.group("name").strip(),
            set=m.group("set").lower() if m.group("set") else None,
            collector_number=m.group("cn") if m.group("cn") else None,
            foil=bool(m.group("foil")),
            section=current,
        ))
    return res


# ---------- XLSX parsing (master-list round-trip) ----------

MASTER_LIST_COLUMNS = (
    "set", "collector_number", "name", "rarity", "mana_value",
    "usd", "usd_foil", "qty_normal", "qty_foil",
)


def parse_master_list_xlsx(path: Path) -> ParseResult:
    """Read a filled-in master-list spreadsheet emitted by ``mm set master-list``.

    Yields one Entry per (row, finish) where the qty column is a positive
    integer. Both ``qty_normal`` and ``qty_foil`` may produce entries from the
    same row.

    If the workbook has a ``_meta`` sheet (written by recent
    ``write_master_list_xlsx`` runs) its key/value pairs are attached to
    ``ParseResult.meta``. Older XLSX files without ``_meta`` parse fine but
    leave ``meta=None`` — the caller can then infer scope from the rows.
    """
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), data_only=True)

    # Active sheet is the visible master-list sheet; the meta sheet (if any)
    # is hidden and named "_meta".
    ws = None
    for name in wb.sheetnames:
        if name != "_meta":
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    res = ParseResult()

    if "_meta" in wb.sheetnames:
        meta_ws = wb["_meta"]
        meta: dict[str, str] = {}
        meta_rows = meta_ws.iter_rows(values_only=True)
        next(meta_rows, None)  # skip header
        for mrow in meta_rows:
            if not mrow or mrow[0] is None:
                continue
            k = str(mrow[0]).strip()
            v = "" if (len(mrow) < 2 or mrow[1] is None) else str(mrow[1]).strip()
            meta[k] = v
        res.meta = meta

    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        res.warnings.append("XLSX has no header row")
        return res

    header_lower = [str(h).strip().lower() if h is not None else "" for h in header]
    idx = {c: header_lower.index(c) for c in MASTER_LIST_COLUMNS if c in header_lower}
    missing = [c for c in MASTER_LIST_COLUMNS if c not in idx]
    if missing:
        res.warnings.append(f"XLSX missing expected columns: {missing!r}")

    for row_num, row in enumerate(rows, start=2):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        try:
            set_code = (row[idx["set"]] or "").strip().lower()
            cn = str(row[idx["collector_number"]]).strip() if row[idx["collector_number"]] is not None else ""
            name = (row[idx["name"]] or "").strip() if "name" in idx else ""
            qn_raw = row[idx["qty_normal"]] if "qty_normal" in idx else None
            qf_raw = row[idx["qty_foil"]] if "qty_foil" in idx else None
        except KeyError:
            continue

        qn = _coerce_qty(qn_raw, row_num, "qty_normal", res)
        qf = _coerce_qty(qf_raw, row_num, "qty_foil", res)

        if qn > 0:
            res.entries.append(Entry(
                qty=qn, raw=f"row {row_num}: {qn} {name} ({set_code}) {cn}",
                name=name, set=set_code, collector_number=cn,
                foil=False, section="mainboard",
            ))
        if qf > 0:
            res.entries.append(Entry(
                qty=qf, raw=f"row {row_num}: {qf} {name} ({set_code}) {cn} ★",
                name=name, set=set_code, collector_number=cn,
                foil=True, section="mainboard",
            ))

    return res


# Markdown intake-doc line: `- (SET) CN [N:k F:k] — ...`
# The trailing chunk after the brackets (link, prices, anything else) is ignored.
MD_LINE_RE = re.compile(
    r"""
    ^\s*-\s+
    \((?P<set>[A-Za-z0-9]{2,6})\)\s+
    (?P<cn>\S+)\s+
    \[\s*N:\s*(?P<n>\d+)\s+F:\s*(?P<f>\d+)\s*\]
    """,
    re.VERBOSE,
)


def parse_master_list_md(path: Path) -> ParseResult:
    """Parse the markdown intake format emitted by ``write_master_list_md()``.

    Reads YAML-frontmatter ``_meta`` into ``ParseResult.meta``. Body lines
    matching ``- (SET) CN [N:k F:k] ...`` are emitted as one or two ``Entry``
    records (one per nonzero finish). Lines that don't match are ignored.

    Like the XLSX parser, this keys cards by ``(set, collector_number)`` —
    the displayed name (whether reskin-merged or plain) is informational
    only, so users can edit it without breaking ingest.
    """
    res = ParseResult()
    text = path.read_text(encoding="utf-8")
    body = text

    # Parse the YAML frontmatter (we don't need a real YAML parser — the
    # writer only emits ``key: value`` pairs of strings).
    if text.startswith("---\n") or text.startswith("---\r\n"):
        end = text.find("\n---\n", 4)
        if end == -1:
            end = text.find("\n---\r\n", 4)
        if end != -1:
            fm = text[4:end]
            meta: dict[str, str] = {}
            for line in fm.splitlines():
                if ":" not in line:
                    continue
                k, _, v = line.partition(":")
                meta[k.strip()] = v.strip()
            if meta:
                res.meta = meta
            body = text[end:].split("\n", 1)[1] if "\n" in text[end:] else ""

    line_num = 0
    for raw_line in body.splitlines():
        line_num += 1
        m = MD_LINE_RE.match(raw_line)
        if not m:
            continue
        set_code = m.group("set").lower()
        cn = m.group("cn")
        try:
            qn = int(m.group("n"))
            qf = int(m.group("f"))
        except ValueError:
            res.warnings.append(f"line {line_num}: bad integer in {raw_line!r}")
            continue

        if qn > 0:
            res.entries.append(Entry(
                qty=qn, raw=f"line {line_num}: {qn} (SET) {cn} [N:{qn}]",
                name="", set=set_code, collector_number=cn,
                foil=False, section="mainboard",
            ))
        if qf > 0:
            res.entries.append(Entry(
                qty=qf, raw=f"line {line_num}: {qf} (SET) {cn} [F:{qf}]",
                name="", set=set_code, collector_number=cn,
                foil=True, section="mainboard",
            ))

    return res


def _coerce_qty(raw, row_num: int, col: str, res: ParseResult) -> int:
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return 0
    try:
        v = int(raw)
    except (TypeError, ValueError):
        res.warnings.append(f"row {row_num}: {col!r} value {raw!r} is not an integer; treating as 0")
        return 0
    if v < 0:
        res.warnings.append(f"row {row_num}: {col!r} value {v} is negative; treating as 0")
        return 0
    return v


# ---------- format detection ----------

def detect_format(text_or_path: str | Path) -> str:
    """Return one of: ``xlsx``, ``md``, ``moxfield`` (covers
    Archidekt/MTGA/MTGO/deckstats — they all share the parser), ``unknown``."""
    if isinstance(text_or_path, Path):
        if text_or_path.suffix.lower() in (".xlsx", ".xlsm"):
            return "xlsx"
        if text_or_path.suffix.lower() in (".md", ".markdown"):
            return "md"
        if text_or_path.suffix.lower() in (".csv",):
            return "csv"
        text = text_or_path.read_text(encoding="utf-8")
    else:
        text = text_or_path
    for line in text.splitlines():
        s = line.strip()
        if not s or IGNORE_RE.match(line) or SECTION_RE.match(line):
            continue
        if CARD_RE.match(line):
            return "moxfield"
        return "unknown"
    return "unknown"


# ---------- resolution against Scryfall ----------

def resolve(result: ParseResult) -> ParseResult:
    """Fill ``Entry.card`` for every parsed entry by hitting /cards/collection
    in batches of 75 through the rate-limited wrapper.

    Multiple entries can share the same identifier (e.g. a master-list row with
    both qty_normal and qty_foil > 0 produces two entries with the same
    ``(set, collector_number)``). We deduplicate the identifiers we send to
    Scryfall but apply the resolved card to every entry that wanted it.
    """
    if not result.entries:
        return result

    def ident_key(d: dict) -> tuple:
        return tuple(sorted((k, str(v).lower()) for k, v in d.items()))

    # Map identifier-key → list of entry indices that need that card.
    pending: dict[tuple, list[int]] = {}
    unique_idents: list[dict] = []
    for i, entry in enumerate(result.entries):
        ident = _identifier_for(entry)
        key = ident_key(ident)
        if key not in pending:
            pending[key] = []
            unique_idents.append(ident)
        pending[key].append(i)

    found, not_found = scryfall.collection(unique_idents)
    for card in found:
        ck_set = (card.get("set") or "").lower()
        ck_cn = card.get("collector_number") or ""
        ck_name = (card.get("name") or "").lower()
        keys_to_try = [
            ident_key({"set": ck_set, "collector_number": ck_cn}),
            ident_key({"name": ck_name}),
            ident_key({"name": ck_name.split(" // ")[0]}),
        ]
        for k in keys_to_try:
            if k in pending:
                for idx in pending.pop(k):
                    result.entries[idx].card = card
                break

    for entry in result.entries:
        if entry.card is None:
            result.not_found.append({
                "qty": entry.qty, "raw": entry.raw, "section": entry.section,
                "reason": (
                    f"no Scryfall match for ({entry.set}) {entry.collector_number}"
                    if entry.set else f"name '{entry.name}'"
                ),
            })
            continue
        if entry.set and entry.collector_number:
            typed = entry.name.lower()
            resolved_name = (entry.card.get("name") or "").lower()
            flavor_name = (entry.card.get("flavor_name") or "").lower()
            # Accepted forms (any of these is a non-mismatch):
            # - oracle name ("Counterspell")
            # - front face of a DFC ("Pegasus Guardian")
            # - flavor name alone ("Wild Rose Rebellion")
            # - the merged "<flavor> / <oracle>" form our XLSX renders
            front = resolved_name.split(" // ")[0]
            accepted = {resolved_name, front}
            if flavor_name:
                accepted.add(flavor_name)
                accepted.add(f"{flavor_name} / {resolved_name}")
                accepted.add(f"{flavor_name} / {front}")
            if typed and typed not in accepted:
                result.warnings.append(
                    f"name/printing mismatch [{entry.section}]: {entry.raw!r} "
                    f"resolved to {entry.card.get('name')!r} via "
                    f"({entry.set}) {entry.collector_number} — likely a typo "
                    f"in the set code or collector number"
                )

    # Attach raw not_found for completeness.
    result.not_found.extend(not_found)
    return result


def _identifier_for(entry: Entry) -> dict:
    if entry.set and entry.collector_number:
        return {"set": entry.set, "collector_number": entry.collector_number}
    return {"name": entry.name}

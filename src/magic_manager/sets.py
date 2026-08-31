"""Set resolution, syncing, and the fillable master-list XLSX builder.

A "set" in Magic isn't always one Scryfall set code. "Final Fantasy" is the
parent expansion ``fin`` plus 8 sibling/child sets (commander, masterpiece,
promos, art series, etc.). The resolver returns the parent + every set whose
``parent_set_code`` traces back to it.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from . import db, scryfall, util


RARITY_ORDER = {
    "mythic":   0,
    "rare":     1,
    "uncommon": 2,
    "common":   3,
    "bonus":    4,
    "special":  5,
}


# Set types that count as "things players actively collect" for the default
# inventory bundle. Tokens and memorabilia (art series, scene boxes) are off
# by default; explicitly opt them in via include_kinds.
#
# `eternal` is the set_type for the "Jumpstart-analog" Eternal products that
# ship alongside a modern UB release (e.g. `tle` for Avatar, `tmc` for TMNT).
# Their cards appear in Collector/Play Boosters for the family, so they are
# part of the collectable family by user direction — always included.
DEFAULT_INVENTORY_SET_TYPES = frozenset({"expansion", "commander", "masterpiece", "promo", "eternal"})


# Per V1.5 user direction: prerelease promos, store-stamped promos,
# japanshowcase variants, serialized cards, and weird-border variants are
# excluded from default master-list output so the user only sees printings
# they actually catalog. Toggled by --include-variants on master-list.
EXCLUDED_BORDERS = frozenset({"white", "yellow"})
EXCLUDED_PROMO_TYPES = frozenset({
    "prerelease", "datestamped", "stamped", "promopack",
    "japanshowcase", "serialized",
    # Arena/Alchemy rebalanced cards exist only as digital re-tunings — they
    # have no physical counterpart, no foil finish, no secondary-market price,
    # and a literal "arena" security_stamp. Always filtered from physical
    # collection workflows. Mirrors selectors.DIGITAL_ONLY_PROMO_TYPES on the
    # missing-set side; both signals are universal across MTG (not set-specific).
    "rebalanced", "alchemy",
})


def is_excluded_variant(card_row) -> bool:
    """Return True if a card row should be filtered from default master-list
    output. Operates on either a sqlite Row or a Scryfall API dict.

    The filter intentionally errs on the side of exclusion — if a card has
    ANY of the excluded promo_types or borders, it's out. The user-facing
    effect is "the master list shows the printings I'd want to catalog,
    nothing else."
    """
    bc = card_row["border_color"] if hasattr(card_row, "keys") else card_row.get("border_color")
    if bc and str(bc).lower() in EXCLUDED_BORDERS:
        return True
    raw_pt = card_row["promo_types"] if hasattr(card_row, "keys") else card_row.get("promo_types")
    if raw_pt is None:
        return False
    # promo_types is a JSON-encoded list in our DB rows but a real list in
    # the Scryfall response — handle both.
    if isinstance(raw_pt, str):
        import json as _json
        try:
            pts = _json.loads(raw_pt)
        except _json.JSONDecodeError:
            return False
    else:
        pts = raw_pt
    return any(p in EXCLUDED_PROMO_TYPES for p in (pts or []))


@dataclass
class ResolvedSet:
    code: str           # anchor set code, e.g. "fin"
    name: str           # display name, e.g. "Final Fantasy"
    related: list[dict] # all sets in the family, anchor first

    @property
    def all_codes(self) -> list[str]:
        return [s["code"] for s in self.related]

    def filtered_codes(self, *, include_kinds: Iterable[str] = ()) -> list[str]:
        """Codes in the family whose ``set_type`` is in the default inventory
        bundle (expansion / commander / masterpiece / promo / eternal),
        expanded by ``include_kinds`` (e.g. ``{"token", "memorabilia"}``).

        The anchor is always included regardless — naming a token set
        explicitly should still produce that set in the output.
        """
        allowed = set(DEFAULT_INVENTORY_SET_TYPES) | set(include_kinds)
        out: list[str] = []
        for s in self.related:
            if s["code"] == self.code or s.get("set_type") in allowed:
                out.append(s["code"])
        return out

    @property
    def filtered_related(self) -> list[dict]:
        codes = set(self.filtered_codes())
        return [s for s in self.related if s["code"] in codes]


# ---------- name resolution ----------

def resolve(name_or_code: str) -> ResolvedSet:
    """Resolve to a specific Scryfall set (the "anchor") plus everything in
    its family tree.

    If the user names a specific child set ("Final Fantasy: Through the Ages"
    or ``fca``), the anchor is that set — ``--include-related`` then expands
    to the parent + all siblings. If they name a parent ("Final Fantasy" or
    ``fin``), the anchor is the parent.
    """
    needle = name_or_code.strip().lower()
    all_sets = scryfall.all_sets()
    by_code = {s["code"].lower(): s for s in all_sets}

    if needle in by_code:
        anchor = by_code[needle]
    else:
        candidates = [s for s in all_sets if s["name"].lower() == needle]
        if not candidates:
            candidates = [s for s in all_sets if needle in s["name"].lower()]
        if not candidates:
            raise LookupError(f"no Scryfall set matches {name_or_code!r}")
        # Prefer parents when there's ambiguity, otherwise take the first hit.
        parents = [s for s in candidates if not s.get("parent_set_code")]
        anchor = parents[0] if parents else candidates[0]

    # The "family" is the parent + every set whose ancestry chains back to it.
    parent = _walk_to_parent(by_code, anchor)
    related = [parent] + _descendants_of(all_sets, parent["code"])
    # Move the anchor to the front so callers/UIs can show it first.
    related = [anchor] + [s for s in related if s["code"] != anchor["code"]]
    return ResolvedSet(code=anchor["code"], name=anchor["name"], related=related)


def _walk_to_parent(by_code: dict, start: dict) -> dict:
    cur = start
    while cur.get("parent_set_code"):
        nxt = by_code.get(cur["parent_set_code"])
        if not nxt or nxt["code"] == cur["code"]:
            break
        cur = nxt
    return cur


def _descendants_of(all_sets: list[dict], parent_code: str) -> list[dict]:
    """All sets whose parent_set_code chains back to ``parent_code``."""
    by_code = {s["code"]: s for s in all_sets}
    out: list[dict] = []
    for s in all_sets:
        if s["code"] == parent_code:
            continue
        cur = s
        while cur.get("parent_set_code"):
            if cur["parent_set_code"] == parent_code:
                out.append(s)
                break
            cur = by_code.get(cur["parent_set_code"])
            if not cur:
                break
    return out


# ---------- syncing ----------

def sync(set_codes: Iterable[str]) -> int:
    """Pull every printing in ``set_codes`` into the cards table. Returns rows synced.

    English-only: sets that ship only in non-English (e.g. ``rfin`` regional
    promos which are JP-only) will simply have zero rows imported. The user
    catalogs English copies; non-English-only prints don't belong in the checklist.
    """
    codes = [c.lower() for c in set_codes]
    if not codes:
        return 0
    # Build a single search query using `or` so we paginate once. ``lang:en``
    # filters out the Japanese-only rfin J1/J2 prints (and any future non-English
    # variants Scryfall adds to a release). Cap the codes-per-query so a very
    # large set list (e.g. the all-sets precon catalog's ~180 sets) can't build
    # a Scryfall query string past its length limit — batch and sum instead.
    _MAX_CODES_PER_QUERY = 60
    n = 0
    with db.connect() as conn:
        for i in range(0, len(codes), _MAX_CODES_PER_QUERY):
            batch = codes[i:i + _MAX_CODES_PER_QUERY]
            query = "(" + " or ".join(f"e:{c}" for c in batch) + ") lang:en"
            for card in scryfall.search(query, unique="prints"):
                db.upsert_card(conn, card)
                n += 1
    return n


# ---------- master-list seeding + XLSX emit ----------

def register_set_target(anchor_code: str, related_codes: Iterable[str], *,
                        include_variants: bool = False,
                        rarity_filter: Iterable[str] | None = None) -> dict:
    """Insert (or update) a set_targets row recording user intent to track a set.

    The set's universe of printings lives in the cards table — set_targets
    just records "I'm tracking this anchor + family" for `set:CODE missing`
    queries. Replaces V1's seed-rows-at-qty-0 pattern.

    Returns ``{"action": "inserted"|"updated", "anchor_code": str,
    "related_codes": list[str]}``.
    """
    import json as _json
    anchor = anchor_code.lower()
    codes = sorted({c.lower() for c in related_codes})
    if not codes:
        codes = [anchor]
    rarities = sorted({r.lower() for r in (rarity_filter or []) if r}) or None
    now = datetime.now(timezone.utc).isoformat(timespec="seconds")
    with db.connect() as conn:
        existing = conn.execute(
            "SELECT 1 FROM set_targets WHERE anchor_code = ?", (anchor,)
        ).fetchone()
        if existing:
            conn.execute(
                """
                UPDATE set_targets
                SET related_codes = ?, include_variants = ?, rarity_filter = ?, updated_at = ?
                WHERE anchor_code = ?
                """,
                (_json.dumps(codes), 1 if include_variants else 0,
                 _json.dumps(rarities) if rarities else None, now, anchor),
            )
            action = "updated"
        else:
            conn.execute(
                """
                INSERT INTO set_targets
                  (anchor_code, related_codes, include_variants, rarity_filter, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (anchor, _json.dumps(codes), 1 if include_variants else 0,
                 _json.dumps(rarities) if rarities else None, now, now),
            )
            action = "inserted"
    return {"action": action, "anchor_code": anchor, "related_codes": codes}


def _add_mode_banner_sheet(wb, mode: str) -> None:
    """Prepend a visible, colored ``README`` sheet stating the checklist's
    ingest semantics, so the danger of a ``modify`` file is legible the moment
    it's opened — not just implied by the filename token.

    Kept as a separate sheet (not a banner row) so the ``checklist`` grid,
    its row-2 freeze/validation ranges, and the parser (which reads row 1 as
    the header) are all untouched. The checklist sheet is left active so the
    user still lands on the data; this sheet sits first in the tab order.
    """
    from openpyxl.styles import Alignment, Font, PatternFill

    banner = wb.create_sheet("README", 0)  # index 0 → first tab
    if mode == "modify":
        banner.sheet_properties.tabColor = "C00000"  # red tab
        lines = [
            "⚠  MODIFY CHECKLIST — read before editing",
            "",
            "Ingesting this file SETS each row to the quantity shown.",
            "The pre-filled numbers are your CURRENT collection.",
            "Editing a number changes your collection by the difference",
            "(3 → 5 adds 2;  3 → 1 removes 2;  set to 0 to zero that row).",
            "",
            "Rows you DON'T touch are left alone — this is not a wipe.",
            "At ingest you'll be asked whether to also zero in-partition",
            "rows that are absent from the file (only say yes for a full audit).",
        ]
    else:
        banner.sheet_properties.tabColor = "548235"  # green tab
        lines = [
            "ADD CHECKLIST — new acquisitions",
            "",
            "Cells start blank. Ingesting this file ADDS the quantities you",
            "enter to your existing collection. It can only increase counts —",
            "it never overwrites or removes anything. Safe for booster packs,",
            "precons, trade-ins, and anything you just picked up.",
        ]
    title_font = Font(bold=True, size=13, color="FFFFFF")
    title_fill = PatternFill(start_color=("C00000" if mode == "modify" else "548235"),
                             end_color=("C00000" if mode == "modify" else "548235"),
                             fill_type="solid")
    for i, text in enumerate(lines, start=1):
        cell = banner.cell(row=i, column=1, value=text)
        if i == 1:
            cell.font = title_font
            cell.fill = title_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    banner.column_dimensions["A"].width = 68
    # Inserting at index 0 shifts the active-sheet pointer onto README. Restore
    # the checklist sheet as active so the user still lands on the data grid,
    # with README sitting first in the tab order as an unmissable heads-up.
    if "checklist" in wb.sheetnames:
        wb.active = wb.sheetnames.index("checklist")


def write_master_list_xlsx(set_codes: Iterable[str], out_path: Path,
                           include_tokens: bool = False,
                           prepopulate_from_inventory: bool = True,
                           rarity_filter: Iterable[str] | None = None,
                           anchor_code: str | None = None,
                           slug: str | None = None,
                           include_variants: bool = False,
                           mode: str = "add") -> tuple[int, int]:
    """Emit a fillable XLSX of every printing in ``set_codes``.

    When ``prepopulate_from_inventory`` is True (default), qty cells are
    pre-filled from the ``inventory`` table for printings the user already
    owns, so resuming after an ingest doesn't lose visible progress.

    When ``rarity_filter`` is given (case-insensitive iterable of rarities),
    only printings with one of those rarities are emitted.

    A hidden ``_meta`` sheet is always written so ingest can recover scope
    later: ``anchor_code``, ``set_codes``, ``rarity_filter``, ``slug``,
    ``generated_at``, ``magic_manager_version``.

    Returns ``(rows_written, cells_prefilled)``.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    from . import __version__

    codes = [c.lower() for c in set_codes]
    if not codes:
        raise ValueError("no set codes provided")

    rarity_set: set[str] | None = None
    if rarity_filter is not None:
        rarity_set = {r.lower() for r in rarity_filter if r and str(r).strip()}
        if not rarity_set:
            rarity_set = None  # treat empty list as "no filter"

    placeholders = ",".join("?" for _ in codes)
    with db.connect() as conn:
        rows = conn.execute(
            f"""
            SELECT scryfall_id, set_code, collector_number, name, flavor_name,
                   rarity, cmc, prices_usd, prices_usd_foil, is_token, scryfall_uri,
                   frame_effects, promo_types, border_color, full_art
            FROM cards
            WHERE set_code IN ({placeholders})
            ORDER BY 1, 2
            """,
            codes,
        ).fetchall()

        # (scryfall_id, finish) -> quantity (from inventory; only printings
        # in this set's family will actually be looked up by the loop below).
        prepop: dict[tuple[str, str], int] = {}
        if prepopulate_from_inventory:
            for r in conn.execute(
                "SELECT scryfall_id, finish, quantity FROM inventory"
            ).fetchall():
                prepop[(r["scryfall_id"], r["finish"])] = r["quantity"]

    if not include_tokens:
        rows = [r for r in rows if not r["is_token"]]
    if not include_variants:
        rows = [r for r in rows if not is_excluded_variant(r)]
    if rarity_set is not None:
        rows = [r for r in rows if (r["rarity"] or "").lower() in rarity_set]

    # Sort: set code asc, then collector_number asc (numeric where possible).
    # Inventory checklists are *input* tools — the user fills them in while
    # holding a physically-sorted pile of cards. Set+CN matches how MTG
    # players sort cards on their desk; rarity grouping (the old order) made
    # it harder to find any specific card. Output artifacts (missing-set,
    # query reports) still sort rarity-first because they're for *reading*.
    def cn_sortkey(cn: str) -> tuple:
        m = re.match(r"^(\d+)(.*)$", cn or "")
        if m:
            return (int(m.group(1)), m.group(2))
        return (10**9, cn or "")

    rows = sorted(
        rows,
        key=lambda r: (
            r["set_code"],
            cn_sortkey(r["collector_number"]),
        ),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "checklist"

    # Column order is fixed; treatment is V1.5 between rarity and mana_value.
    # If columns shift, update parse_master_list_xlsx, the qty-tint indices,
    # and the widths dict below.
    headers = ["set", "collector_number", "name", "rarity", "treatment",
               "mana_value", "usd", "usd_foil", "qty_normal", "qty_foil"]
    ws.append(headers)
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

    qty_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    int_validator = DataValidation(type="whole", operator="greaterThanOrEqual",
                                   formula1=0, allow_blank=True)
    int_validator.error = "Enter a non-negative integer (or leave blank for 0)."
    int_validator.errorTitle = "Invalid quantity"
    ws.add_data_validation(int_validator)

    cells_prefilled = 0
    # Hyperlink-styled font for the name cell — blue + underline mimics how
    # most apps render web links. The cell's value is unchanged (just the
    # displayed name); the hyperlink is a separate property openpyxl supports.
    link_font = Font(color="0563C1", underline="single")
    from .treatments import compute_treatment
    for r in rows:
        qn = prepop.get((r["scryfall_id"], "nonfoil"))
        qf = prepop.get((r["scryfall_id"], "foil"))
        if qn is not None:
            cells_prefilled += 1
        if qf is not None:
            cells_prefilled += 1
        # Render the displayed name as "<flavor_name> / <oracle_name>" when the
        # printing has a Universes Beyond reskin name (e.g. FCA Counterspell →
        # "Wild Rose Rebellion / Counterspell"); otherwise just the oracle name.
        # Round-trip-safe: parse_master_list_xlsx() keys on (set_code, cn).
        flavor = r["flavor_name"]
        display_name = f"{flavor} / {r['name']}" if flavor else r["name"]
        treatment = compute_treatment(r)
        ws.append([
            r["set_code"],
            r["collector_number"],
            display_name,
            r["rarity"],
            treatment,
            r["cmc"],
            r["prices_usd"],
            r["prices_usd_foil"],
            qn,
            qf,
        ])
        # Force collector_number to render as text. Many CNs are pure
        # digits ('4', '210') and Excel auto-coerces them to numbers,
        # then complains with the green-triangle "Number Stored as Text"
        # warning when other CNs in the same column have letter suffixes
        # (like '212s' or '551f'). Setting the cell's number_format to '@'
        # tells Excel "this is intentional text" and the warning disappears.
        cn_cell = ws.cell(row=ws.max_row, column=2)
        cn_cell.number_format = "@"
        # Attach a clickable hyperlink to the name cell, pointing at the card's
        # Scryfall page. Falls through silently if scryfall_uri is missing for
        # this row (older DB rows from V1.2 might not have one — re-sync fixes).
        uri = r["scryfall_uri"]
        if uri:
            name_cell = ws.cell(row=ws.max_row, column=3)
            name_cell.hyperlink = uri
            name_cell.font = link_font
    last_row = ws.max_row

    # Tint qty columns and apply integer validation. With treatment inserted
    # at column 5, qty_normal/qty_foil are now columns 9/10.
    for col_idx in (9, 10):
        col_letter = get_column_letter(col_idx)
        rng = f"{col_letter}2:{col_letter}{last_row}"
        int_validator.add(rng)
        for r in range(2, last_row + 1):
            ws.cell(row=r, column=col_idx).fill = qty_fill

    # Sensible widths. Column 5 is treatment — sized for "b|shw|ext|sm|ff"
    # worst case. Column 3 (name) holds long reskin pairs like
    # "Knights of San d'Oria / Ranger-Captain of Eos" so it gets generous
    # room.
    widths = {1: 6, 2: 8, 3: 48, 4: 10, 5: 14, 6: 6, 7: 9, 8: 9, 9: 11, 10: 9}
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Format USD columns as currency with two decimals so prices line up
    # ($3.00 / $0.43 instead of $3.0 / $0.43).
    for col_idx in (7, 8):
        for row_idx in range(2, last_row + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = '"$"#,##0.00'

    ws.freeze_panes = "A2"

    # Hidden _meta sheet: lets `mm set ingest` recover scope without trusting
    # the filename. Two columns (key, value) so the format stays human-readable
    # in case someone unhides the sheet for debugging.
    meta_ws = wb.create_sheet("_meta")
    meta_ws.sheet_state = "hidden"
    meta_ws.append(["key", "value"])
    meta_ws["A1"].font = Font(bold=True)
    meta_ws["B1"].font = Font(bold=True)

    rarity_value = ",".join(sorted(rarity_set)) if rarity_set else ""
    meta = {
        # `kind` distinguishes this artifact from `mm query missing-set`'s output
        # (which writes `kind: "missing"` to its own _meta sheet). Inventory
        # checklists round-trip through `mm set ingest`; missing checklists
        # never do. See feedback_checklist_artifacts memory for the full split.
        "kind": "inventory",
        # `mode` declares the intended ingest semantics — read by `mm set
        # ingest` and applied automatically. 'modify' → replace ingest (each
        # in-partition row is SET to its cell value — a signed change vs
        # current; absent rows are left alone unless the user opts into zeroing
        # at ingest time); 'add' → additive ingest (qty>0 cells sum into
        # existing inventory). Also encoded in the filename for on-disk clarity,
        # and surfaced in the visible README banner sheet.
        "mode": mode,
        "anchor_code": (anchor_code or codes[0]).lower(),
        "set_codes": ",".join(codes),
        "rarity_filter": rarity_value,
        "slug": slug or out_path.stem,
        "include_tokens": "1" if include_tokens else "0",
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "magic_manager_version": __version__,
    }
    for k, v in meta.items():
        meta_ws.append([k, v])

    # Hidden _legend sheet: documents the treatment-column keyword space so
    # users can unhide it for reference without leaving the workbook.
    from .treatments import LEGEND
    legend_ws = wb.create_sheet("_legend")
    legend_ws.sheet_state = "hidden"
    legend_ws.append(["code", "meaning"])
    legend_ws["A1"].font = Font(bold=True)
    legend_ws["B1"].font = Font(bold=True)
    for code, meaning in LEGEND:
        legend_ws.append([code, meaning])
    legend_ws.column_dimensions["A"].width = 6
    legend_ws.column_dimensions["B"].width = 90

    _add_mode_banner_sheet(wb, mode)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    for _ws in wb.worksheets:
        util.apply_base_font_size(_ws)
    wb.save(out_path)
    return (last_row - 1, cells_prefilled)


def read_master_list_meta(path: Path) -> dict | None:
    """Read the ``_meta`` sheet (XLSX) or YAML frontmatter (MD) from a
    checklist file. Returns the dict of key/value strings, or ``None`` if no
    metadata is present.

    Works for both inventory checklists and Jumpstart checklists — the meta
    shape differs but the read is the same.
    """
    suffix = path.suffix.lower()
    if suffix == ".md":
        text = path.read_text(encoding="utf-8")
        if not (text.startswith("---\n") or text.startswith("---\r\n")):
            return None
        end = text.find("\n---\n", 4)
        if end == -1:
            end = text.find("\n---\r\n", 4)
        if end == -1:
            return None
        out: dict[str, str] = {}
        for line in text[4:end].splitlines():
            if ":" not in line:
                continue
            k, _, v = line.partition(":")
            out[k.strip()] = v.strip()
        return out or None

    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), data_only=True)
    if "_meta" not in wb.sheetnames:
        return None
    ws = wb["_meta"]
    out: dict[str, str] = {}
    rows = ws.iter_rows(values_only=True)
    next(rows, None)  # skip header
    for row in rows:
        if not row or row[0] is None:
            continue
        key = str(row[0]).strip()
        val = "" if (len(row) < 2 or row[1] is None) else str(row[1]).strip()
        out[key] = val
    return out


# ---------- V2 inventory ingest ----------

def ingest_inventory_from_xlsx(path: Path, *, mode: str = "replace",
                               zero_untouched: bool = False) -> dict:
    """Parse a filled-in master-list XLSX/MD and write qty cells to inventory.

    Per-cell semantics by ``mode``:
      - 'replace' (from a ``modify`` checklist): each row's quantity is SET to
        the cell value. Prefilled cells left as-is are no-ops; editing a number
        moves the collection by the difference; a cell explicitly set to 0
        zeroes THAT row. This is a signed transaction, not a history rewrite.
      - 'additive' (from an ``add`` checklist): only cells with qty>0 add to
        existing inventory; nothing is zeroed.

    ``zero_untouched`` (replace mode only): when True, in-partition rows that
    are ABSENT from the file are also zeroed — the full-audit "the file is
    authoritative" behavior. Default False: untouched/absent rows are left
    alone (we don't rewrite history the user didn't record). No effect in
    additive mode.

    Partition is derived from the file's _meta sheet (definitive) or
    inferred from the rows present (fallback). Cards in the file that
    aren't in the partition's set codes are flagged as 'extras' (the user
    pasted unrelated cards into a set's checklist).

    Returns ``{"added": N, "updated": N, "zeroed": N, "warnings": [...],
    "not_found": [...], "extras": [...]}``.
    """
    from . import db, parsers
    if mode not in ("replace", "additive"):
        raise ValueError(f"unknown mode {mode!r}; expected 'replace' or 'additive'")

    fmt = parsers.detect_format(path)
    if fmt == "xlsx":
        result = parsers.parse_master_list_xlsx(path)
    elif fmt == "md":
        result = parsers.parse_master_list_md(path)
    else:
        result = parsers.parse_text(path.read_text(encoding="utf-8"))
    parsers.resolve(result)

    partition = _derive_inventory_partition(result)
    added = 0
    updated = 0
    zeroed = 0
    extras: list[dict] = []
    seen_keys: set[tuple[str, str]] = set()
    now = datetime.now(timezone.utc).isoformat(timespec="seconds")

    with db.connect() as conn:
        for entry in result.entries:
            if entry.card is None:
                continue
            db.upsert_card(conn, entry.card)
            scry_id = entry.card["id"]
            finish = "foil" if entry.foil else "nonfoil"
            card_set = (entry.card.get("set") or "").lower()

            # Out-of-partition cards are 'extras' (file tried to set qty for
            # a card outside the file's declared scope).
            if partition and card_set not in partition.set_codes:
                extras.append({
                    "raw": entry.raw,
                    "reason": (
                        f"card {entry.card['name']} ({card_set}) "
                        f"{entry.card.get('collector_number')} is outside the "
                        f"file's partition (set codes: {partition.set_codes})"
                    ),
                })
                continue

            seen_keys.add((scry_id, finish))

            row = conn.execute(
                "SELECT quantity FROM inventory WHERE scryfall_id = ? AND finish = ?",
                (scry_id, finish),
            ).fetchone()
            current_qty = row["quantity"] if row else 0

            if mode == "replace":
                new_qty = entry.qty
            else:
                if entry.qty <= 0:
                    continue
                new_qty = current_qty + entry.qty

            if new_qty == current_qty:
                continue
            if new_qty == 0:
                if current_qty > 0:
                    conn.execute(
                        "DELETE FROM inventory WHERE scryfall_id = ? AND finish = ?",
                        (scry_id, finish),
                    )
                    zeroed += 1
            elif row:
                conn.execute(
                    "UPDATE inventory SET quantity = ? WHERE scryfall_id = ? AND finish = ?",
                    (new_qty, scry_id, finish),
                )
                updated += 1
            else:
                conn.execute(
                    """
                    INSERT INTO inventory (scryfall_id, finish, quantity, acquired_at)
                    VALUES (?, ?, ?, ?)
                    """,
                    (scry_id, finish, new_qty, now),
                )
                added += 1

        # Replace mode, opt-in only: zero out in-partition inventory rows not
        # seen in the file (full-audit "file is authoritative"). Off by default
        # so a modify ingest is a signed per-row transaction that never wipes
        # rows the user didn't touch.
        if mode == "replace" and zero_untouched and partition is not None:
            in_partition_rows = conn.execute(
                f"""
                SELECT inv.scryfall_id, inv.finish, inv.quantity
                FROM inventory inv
                JOIN cards c ON c.scryfall_id = inv.scryfall_id
                WHERE LOWER(c.set_code) IN ({",".join("?" for _ in partition.set_codes)})
                """ + (
                    f" AND LOWER(c.rarity) IN ({','.join('?' for _ in partition.rarities)})"
                    if partition.rarities else ""
                ),
                partition.set_codes + (partition.rarities or []),
            ).fetchall()
            for r in in_partition_rows:
                key = (r["scryfall_id"], r["finish"])
                if key in seen_keys:
                    continue
                conn.execute(
                    "DELETE FROM inventory WHERE scryfall_id = ? AND finish = ?",
                    (r["scryfall_id"], r["finish"]),
                )
                zeroed += 1

        db.record_import(conn,
                         command=f"ingest_inventory_from_xlsx mode={mode}",
                         source_path=str(path),
                         rows_changed=added + updated + zeroed)

    return {
        "added": added,
        "updated": updated,
        "zeroed": zeroed,
        "warnings": result.warnings,
        "not_found": result.not_found,
        "extras": extras,
    }


@dataclass
class _InventoryPartition:
    set_codes: list[str]
    rarities: list[str] | None


def _derive_inventory_partition(result) -> "_InventoryPartition | None":
    """Same partition logic as lists._derive_partition but parameterized
    against the inventory table (no label scoping)."""
    meta = result.meta or {}
    if meta:
        codes = [c.strip().lower() for c in (meta.get("set_codes") or "").split(",") if c.strip()]
        rar = [r.strip().lower() for r in (meta.get("rarity_filter") or "").split(",") if r.strip()]
        if codes:
            return _InventoryPartition(set_codes=codes, rarities=(rar or None))

    seen_codes: set[str] = set()
    seen_rarities: set[str] = set()
    for entry in result.entries:
        if entry.card:
            seen_codes.add((entry.card.get("set") or "").lower())
            r = (entry.card.get("rarity") or "").lower()
            if r:
                seen_rarities.add(r)
    if not seen_codes:
        return None
    return _InventoryPartition(
        set_codes=sorted(seen_codes),
        rarities=sorted(seen_rarities) if len(seen_rarities) == 1 else None,
    )


def _summarize_deck_checklist(path: Path, meta: dict) -> dict:
    """Pre-ingest preview for a deck checklist (kind=precon | jumpstart).

    The inventory summarizer's qty_normal/qty_foil model doesn't fit these
    files (they use keep_qty/deconstructed_qty), so this reports the
    deck-shaped stats the ``/ingest-new-inventory-list`` command needs:
    which decks are filled, how many will be constructed vs deconstructed.

    Returns a dict sharing the inventory summary's key names where they carry
    over (``rows_total``, ``rows_with_qty``, ``total_qty``, ``estimated_value``,
    ``warnings``) plus a ``kind`` discriminator and deck-specific fields:
    ``decks_to_construct``, ``loose_copies``, and ``filled`` (per-acted-row
    ``{file_name, label, constructed_qty, deconstructed_qty, delta, set,
    usd_total}``). For a precon ``modify`` file the entered numbers are absolute
    targets prefilled from the live deck counts, so a row is "acted on" only
    when it differs from its current count; the reported construct/loose counts
    are the positive deltas (what this ingest would build/tear down).
    ``estimated_value`` sums ``usd_total`` over rows that build a new copy.
    """
    from . import parsers
    kind = meta.get("kind") or "precon"
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        parsed = parsers.parse_jumpstart_list_xlsx(path)
    else:
        parsed = parsers.parse_jumpstart_list_md(path)

    # The deck-checklist parser keeps only file_name/keep_qty/deconstructed_qty.
    # Pull the descriptive set/usd_total columns straight from the XLSX cells so
    # the preview can show per-deck set + value without a network round-trip.
    extra: dict[str, dict] = {}
    if suffix == ".xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(filename=str(path), data_only=True)
        ws = wb["checklist"] if "checklist" in wb.sheetnames else wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None) or ()
        hidx = {str(h).strip().lower(): i for i, h in enumerate(header) if h is not None}
        fn_i = hidx.get("file_name")
        set_i = hidx.get("set")
        usd_i = hidx.get("usd_total")
        if fn_i is not None:
            for row in rows_iter:
                if fn_i >= len(row) or row[fn_i] is None:
                    continue
                fn = str(row[fn_i]).strip()
                usd = row[usd_i] if (usd_i is not None and usd_i < len(row)) else None
                setc = row[set_i] if (set_i is not None and set_i < len(row)) else None
                extra[fn] = {
                    "set": (str(setc).lower() if setc else ""),
                    "usd_total": (float(usd) if isinstance(usd, (int, float)) else None),
                }

    # For precon, decide which rows count as "acted on". In `add` mode any
    # nonzero entered count acts; in `modify` mode the entered numbers are
    # absolute targets prefilled from the real deck counts, so a row acts only
    # if it differs from its current derived count (a nonzero delta).
    file_mode = (meta.get("mode") or "add").lower()
    counts = {}
    if kind == "precon" and file_mode == "modify":
        from . import decks as decks_mod
        counts = decks_mod.precon_unit_counts()

    filled: list[dict] = []
    decks_to_construct = 0
    loose_copies = 0    # deconstructed + pool positive deltas (loose-card copies)
    pool_copies = 0
    total_qty = 0
    estimated_value = 0.0
    for r in parsed.rows:
        entered = (r.keep_qty, r.deconstructed_qty, r.pool_qty)  # (c, d, p)
        if kind == "precon" and file_mode == "modify":
            before = counts.get(r.file_name, (0, 0, 0))
            delta = tuple(e - b for e, b in zip(entered, before))
            acts = any(x != 0 for x in delta)
        else:
            delta = entered
            acts = sum(entered) > 0
        if not acts:
            continue
        delta_c, delta_d, delta_p = delta
        info = extra.get(r.file_name, {})
        usd = info.get("usd_total")
        filled.append({
            "file_name": r.file_name,
            "label": r.theme or r.file_name,
            "keep_qty": entered[0],          # jumpstart flag / precon target constructed
            "constructed_qty": entered[0],
            "deconstructed_qty": entered[1],
            "pool_qty": entered[2],
            "delta": delta,
            "set": info.get("set", ""),
            "usd_total": usd,
        })
        total_qty += max(0, delta_c) + max(0, delta_d) + max(0, delta_p)
        # "to construct" counts positive construct deltas (new builds this ingest).
        decks_to_construct += max(0, delta_c) if kind == "precon" else (1 if entered[0] == 1 else 0)
        loose_copies += max(0, delta_d) + max(0, delta_p)
        pool_copies += max(0, delta_p)
        if usd is not None and delta_c > 0:
            estimated_value += usd

    return {
        "path": str(path),
        "meta": meta,
        "kind": kind,
        # Inventory-shape keys kept for the shared text printer / skill.
        "anchor_code": "",
        "set_codes": [],
        "rarity_filter": [],
        "rows_total": len(parsed.rows),
        "rows_with_qty": len(filled),
        "total_qty": total_qty,
        "estimated_value": estimated_value,
        "top_value": [],
        # Deck-specific fields.
        "mode": file_mode,
        "decks_to_construct": decks_to_construct,
        "loose_copies": loose_copies,
        "pool_copies": pool_copies,
        "filled": filled,
        "warnings": parsed.warnings,
    }


def summarize_intake_file(path: Path) -> dict:
    """Pre-ingest preview for the slash command. Handles both XLSX and md.

    For inventory checklists (``kind=inventory`` or untagged) returns a dict
    with: ``path``, ``meta`` (or ``None``), ``kind``, ``anchor_code``,
    ``set_codes``, ``rarity_filter``, ``rows_total``, ``rows_with_qty``,
    ``total_qty``, ``estimated_value``, ``top_value`` (top 5 rows by line
    value), ``warnings`` (parser warnings). Deck checklists (``kind=precon`` /
    ``kind=jumpstart``) are routed to ``_summarize_deck_checklist``, which
    returns the same key names where they carry over plus deck-specific fields.

    Doesn't hit the network beyond what the parser already does (the
    rate-limited /cards/collection lookup for resolution).
    """
    from . import parsers
    # Kind-dispatch: deck checklists don't fit the inventory qty model.
    file_meta = read_master_list_meta(path) or {}
    if file_meta.get("kind") in ("precon", "jumpstart"):
        return _summarize_deck_checklist(path, file_meta)

    fmt = parsers.detect_format(path)
    if fmt == "md":
        result = parsers.parse_master_list_md(path)
    else:
        result = parsers.parse_master_list_xlsx(path)
    parsers.resolve(result)
    meta = result.meta
    rows_with_qty = 0
    total_qty = 0
    estimated_value = 0.0
    rows_for_top: list[tuple[float, dict]] = []
    for e in result.entries:
        if not e.card or e.qty <= 0:
            continue
        rows_with_qty += 1
        total_qty += e.qty
        prices = e.card.get("prices") or {}
        unit_str = prices.get("usd_foil") if e.foil else prices.get("usd")
        try:
            unit = float(unit_str) if unit_str is not None else None
        except (TypeError, ValueError):
            unit = None
        line_value = (unit or 0.0) * e.qty
        if unit is not None:
            estimated_value += line_value
        oracle_name = e.card.get("name") or ""
        flavor_name = e.card.get("flavor_name") or (
            ((e.card.get("card_faces") or [{}])[0] or {}).get("flavor_name")
        )
        display_name = f"{flavor_name} / {oracle_name}" if flavor_name else oracle_name
        rows_for_top.append((line_value, {
            "qty": e.qty,
            "name": display_name,
            "set": (e.card.get("set") or "").lower(),
            "collector_number": e.card.get("collector_number"),
            "finish": "foil" if e.foil else "nonfoil",
            "unit_usd": unit,
            "line_value": line_value if unit is not None else None,
        }))
    rows_for_top.sort(key=lambda x: x[0], reverse=True)
    top_value = [r[1] for r in rows_for_top[:5]]
    anchor = (meta or {}).get("anchor_code") or ""
    set_codes = (meta or {}).get("set_codes") or ""
    rarity_filter = (meta or {}).get("rarity_filter") or ""
    return {
        "path": str(path),
        "meta": meta,
        "kind": (meta or {}).get("kind") or "inventory",
        "anchor_code": anchor,
        "set_codes": [c for c in set_codes.split(",") if c],
        "rarity_filter": [r for r in rarity_filter.split(",") if r],
        "rows_total": len(result.entries),
        "rows_with_qty": rows_with_qty,
        "total_qty": total_qty,
        "estimated_value": estimated_value,
        "top_value": top_value,
        "warnings": result.warnings,
    }


# ---------- markdown intake format ----------

def write_master_list_md(set_codes: Iterable[str], out_path: Path,
                         include_tokens: bool = False,
                         prepopulate_from_inventory: bool = True,
                         rarity_filter: Iterable[str] | None = None,
                         anchor_code: str | None = None,
                         slug: str | None = None,
                         include_variants: bool = False,
                         mode: str = "add") -> tuple[int, int]:
    """Markdown twin of ``write_master_list_xlsx()``.

    File shape:

        ---
        anchor_code: fca
        set_codes: fin,fic,...
        rarity_filter: rare        # blank when no rarity slice
        slug: final-fantasy-...
        include_tokens: 0
        generated_at: 2026-...
        magic_manager_version: 0.1.0
        ---

        # <set name> — <slice description>

        ## Mythic (15 cards)

        - (FCA) 2 [N:0 F:0] — [<displayed name>](<scryfall_uri>) — $4.66 / $164.18
        - (FCA) 5 [N:0 F:0] — ...

    Returns ``(rows_written, cells_prefilled)`` to mirror the XLSX writer.
    The user edits the ``[N:k F:k]`` brackets to record their inventory; the
    parser keys on ``(SET) CN`` so display changes don't affect ingest.
    """
    codes = [c.lower() for c in set_codes]
    if not codes:
        raise ValueError("no set codes provided")

    rarity_set: set[str] | None = None
    if rarity_filter is not None:
        rarity_set = {r.lower() for r in rarity_filter if r and str(r).strip()}
        if not rarity_set:
            rarity_set = None

    placeholders = ",".join("?" for _ in codes)
    with db.connect() as conn:
        rows = conn.execute(
            f"""
            SELECT scryfall_id, set_code, collector_number, name, flavor_name,
                   rarity, prices_usd, prices_usd_foil, is_token, scryfall_uri,
                   frame_effects, promo_types, border_color, full_art
            FROM cards
            WHERE set_code IN ({placeholders})
            ORDER BY 1, 2
            """,
            codes,
        ).fetchall()

        prepop: dict[tuple[str, str], int] = {}
        if prepopulate_from_inventory:
            for r in conn.execute(
                "SELECT scryfall_id, finish, quantity FROM inventory"
            ).fetchall():
                prepop[(r["scryfall_id"], r["finish"])] = r["quantity"]

    if not include_tokens:
        rows = [r for r in rows if not r["is_token"]]
    if not include_variants:
        rows = [r for r in rows if not is_excluded_variant(r)]
    if rarity_set is not None:
        rows = [r for r in rows if (r["rarity"] or "").lower() in rarity_set]

    # Sort: set code asc, then collector_number asc. See the XLSX writer for
    # why inventory checklists sort this way (input tool, not a report).
    def cn_sortkey(cn: str) -> tuple:
        m = re.match(r"^(\d+)(.*)$", cn or "")
        if m:
            return (int(m.group(1)), m.group(2))
        return (10**9, cn or "")

    rows = sorted(
        rows,
        key=lambda r: (
            r["set_code"],
            cn_sortkey(r["collector_number"]),
        ),
    )

    rarity_value = ",".join(sorted(rarity_set)) if rarity_set else ""
    from . import __version__
    meta = {
        # `kind` distinguishes inventory checklists (this writer) from missing
        # checklists (`mm query missing-set`). See feedback_checklist_artifacts.
        "kind": "inventory",
        # `mode` declares ingest semantics. See the XLSX writer for the full
        # rationale; same field, same semantics in markdown form.
        "mode": mode,
        "anchor_code": (anchor_code or codes[0]).lower(),
        "set_codes": ",".join(codes),
        "rarity_filter": rarity_value,
        "slug": slug or out_path.stem,
        "include_tokens": "1" if include_tokens else "0",
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "magic_manager_version": __version__,
    }

    out_lines: list[str] = []
    out_lines.append("---")
    for k, v in meta.items():
        out_lines.append(f"{k}: {v}")
    out_lines.append("---")
    out_lines.append("")

    title = anchor_code.upper() if anchor_code else codes[0].upper()
    if rarity_value:
        title += f" — {rarity_value}"
    out_lines.append(f"# {title}")
    out_lines.append("")
    if mode == "modify":
        out_lines.append(
            "> ⚠ **MODIFY checklist.** Ingest SETS each row to the quantity shown; "
            "the pre-filled numbers are your current collection. Editing a number "
            "changes your collection by the difference (`3`→`5` adds 2; `3`→`1` "
            "removes 2; `0` zeroes that row). Rows you don't touch are left alone — "
            "not a wipe. At ingest you'll be asked whether to also zero in-partition "
            "rows absent from the file (say yes only for a full audit)."
        )
    else:
        out_lines.append(
            "> **ADD checklist.** Cells start blank. Ingest ADDS the quantities you "
            "enter to your existing collection — it only increases, never overwrites "
            "or removes. Safe for new acquisitions."
        )
    out_lines.append("")
    out_lines.append(
        "Edit the `[N:k F:k]` brackets to record quantities. Save, then run "
        "`/ingest-new-inventory-list` (or `mm set ingest`) to apply."
    )
    out_lines.append("")

    cells_prefilled = 0
    current_set = None
    set_counts: dict[str, int] = {}
    for r in rows:
        set_counts[r["set_code"]] = set_counts.get(r["set_code"], 0) + 1

    from .treatments import compute_treatment, LEGEND
    for r in rows:
        set_code = r["set_code"]
        if set_code != current_set:
            current_set = set_code
            out_lines.append("")
            out_lines.append(f"## {set_code.upper()} ({set_counts[set_code]} cards)")
            out_lines.append("")

        flavor = r["flavor_name"]
        display_name = f"{flavor} / {r['name']}" if flavor else r["name"]
        # Escape pipes / brackets that would otherwise interfere with markdown
        # link/table syntax. ``[`` and ``]`` in a link's display text need to
        # be escaped; flavor and oracle names rarely contain them but a few
        # split-card names do (e.g. "Fire // Ice" wouldn't, but
        # "[Battlefield Forge]" hypothetically would).
        safe_name = display_name.replace("[", "\\[").replace("]", "\\]")
        uri = r["scryfall_uri"] or ""
        link = f"[{safe_name}]({uri})" if uri else safe_name

        qn = prepop.get((r["scryfall_id"], "nonfoil"), 0)
        qf = prepop.get((r["scryfall_id"], "foil"), 0)
        if qn > 0:
            cells_prefilled += 1
        if qf > 0:
            cells_prefilled += 1

        usd = r["prices_usd"]
        usd_foil = r["prices_usd_foil"]
        price_segment = f"${usd if usd is not None else '—'} / ${usd_foil if usd_foil is not None else '—'}"

        treatment = compute_treatment(r)
        # Treatment is rendered in `[...]` after the qty bracket; empty for
        # standard prints. Parser keys on `(SET) CN` so the position doesn't
        # affect ingest.
        treatment_seg = f" [{treatment}]" if treatment else ""

        out_lines.append(
            f"- ({r['set_code'].upper()}) {r['collector_number']} "
            f"[N:{qn} F:{qf}]{treatment_seg} — {link} — {price_segment}"
        )

    # Legend at the bottom — informational, ignored by the parser.
    out_lines.append("")
    out_lines.append("## Treatment legend")
    out_lines.append("")
    for code, meaning in LEGEND:
        out_lines.append(f"- `{code}` — {meaning}")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text("\n".join(out_lines) + "\n", encoding="utf-8")
    return (len(rows), cells_prefilled)


# ---------- Jumpstart pack-level checklist (kind=jumpstart) ----------

# A "jumpstart variant" is one MTGJSON deck file describing a sealed pack's
# 15-card content list (e.g. ``Toph_TLE``). Sets like TLE/J25/JMP/J22 publish
# 50+ variants. The user fills in one qty column + a deconstruct flag per row:
#   - ``qty``         → how many copies of this pack were opened
#   - ``deconstruct`` → bool; if False (default) one `pack:*` recipe is created
#                       and qty copies' worth of cards land in inventory. If
#                       True, NO recipe is created and all qty copies go to
#                       loose inventory.
#
# There is no per-copy kept/deconstructed split. V5 semantics write a pack's
# recipe exactly once regardless of copies (import_precon copies= only scales
# the inventory add, not the deck composition), so "kept 2, deconstructed 1"
# was indistinguishable in the DB from "kept 3": the recipe exists either way
# and 3 copies' cards are in inventory either way. The only representable
# distinction is "does a recipe exist for this pack" — hence a boolean, not a
# second qty column. (This is also why keeping qty>=2 leaves the extra copies
# as `available` in `mm deck find`: one pack's worth is pledged to the recipe,
# the rest are loose. That's expected, not a bug.)

def _rollup_deck_prices(deck_data: dict) -> tuple[int, float | None]:
    """Sum card_count + local-market USD across a precon/Jumpstart deck JSON.

    Walks the commander/main/side boards, pulling Scryfall USD per scryfall_id
    from the local cards table (foil price for foil cards, nonfoil otherwise).
    Printings missing from the cards table are skipped silently — the totals
    just under-report, and the user can still ingest. Returns ``(total_count,
    usd_total)`` where ``usd_total`` is ``None`` when nothing priced.
    """
    sids: list[tuple[str, int, bool]] = []  # (scryfall_id, count, is_foil)
    total_count = 0
    for board_key in ("commander", "mainBoard", "sideBoard"):
        for entry in deck_data.get(board_key) or []:
            sid = (entry.get("identifiers") or {}).get("scryfallId")
            count = int(entry.get("count", 1) or 1)
            total_count += count
            if sid:
                sids.append((sid, count, bool(entry.get("isFoil"))))

    usd_total = 0.0
    if sids:
        with db.connect() as conn:
            placeholders = ",".join("?" for _ in sids)
            rows = {
                r["scryfall_id"]: (r["prices_usd"], r["prices_usd_foil"])
                for r in conn.execute(
                    f"SELECT scryfall_id, prices_usd, prices_usd_foil "
                    f"FROM cards WHERE scryfall_id IN ({placeholders})",
                    [s[0] for s in sids],
                ).fetchall()
            }
        for sid, count, is_foil in sids:
            prices = rows.get(sid)
            if not prices:
                continue
            price = prices[1] if is_foil else prices[0]
            if price is not None:
                usd_total += float(price) * count
    return total_count, (round(usd_total, 2) if usd_total else None)


def _jumpstart_variant_summary(variant_meta: dict, *, anchor: str) -> dict:
    """Fetch one variant's MTGJSON deck file and roll up displayable stats.

    Returns ``{"file_name", "theme", "card_count", "usd_total"}``. Pulls
    Scryfall USD per scryfall_id from the local cards table to compute
    ``usd_total`` (nonfoil price × count); printings missing from cards are
    skipped silently — user can still ingest, the totals just under-report.

    ``usd_total`` also folds in the pack's front/title card price (looked up
    from the quarantined ``front_cards`` table by name match) — value-only;
    the front card is never counted toward ``card_count`` since it isn't a
    gameplay card and never lands in inventory.
    """
    from . import front_cards as _fc
    from . import mtgjson as mtgjson_mod
    file_name = variant_meta["fileName"]
    deck_data = mtgjson_mod.deck(file_name)
    total_count, usd_total = _rollup_deck_prices(deck_data)
    theme = variant_meta.get("name") or file_name
    fc = _fc.front_card_for_theme(anchor, theme)
    if fc is not None and fc["prices_usd"] is not None:
        usd_total = round((usd_total or 0.0) + float(fc["prices_usd"]), 2)
    return {
        "file_name": file_name,
        "theme": theme,
        "card_count": total_count,
        "usd_total": usd_total,
    }


def _build_jumpstart_rows(set_code: str) -> list[dict]:
    """Enumerate Jumpstart variants for ``set_code`` and roll each one up."""
    from . import front_cards as _fc
    from . import mtgjson as mtgjson_mod
    variants = mtgjson_mod.jumpstart_variants(set_code)
    if not variants:
        return []
    _fc.sync_front_cards(set_code)  # best-effort; front cards fold into usd_total below
    return [_jumpstart_variant_summary(v, anchor=set_code) for v in
            sorted(variants, key=lambda d: d.get("name") or d.get("fileName") or "")]


def _precon_variant_summary(variant_meta: dict) -> dict:
    """Fetch one precon's MTGJSON deck file and roll up displayable stats.

    Precon twin of ``_jumpstart_variant_summary``, carrying the extra
    descriptive columns a precon checklist surfaces: ``type`` (product type,
    e.g. "Commander Deck"), ``release_date`` (from the DeckList entry), and
    ``commander`` (the commander-board card name(s), joined with "; "; blank
    for non-commander products). ``card_count``/``usd_total`` share the
    ``_rollup_deck_prices`` core with Jumpstart.
    """
    from . import mtgjson as mtgjson_mod
    file_name = variant_meta["fileName"]
    deck_data = mtgjson_mod.deck(file_name)
    total_count, usd_total = _rollup_deck_prices(deck_data)
    commander = "; ".join(
        (c.get("name") or "") for c in (deck_data.get("commander") or [])
    )
    return {
        "file_name": file_name,
        # DeckList's ``code`` is the product's own set code (uppercase); lower
        # it for consistency with the rest of the codebase's code handling.
        "set": (variant_meta.get("code") or "").lower(),
        "deck_name": variant_meta.get("name") or file_name,
        "type": variant_meta.get("type") or "",
        "release_date": variant_meta.get("releaseDate") or "",
        "commander": commander,
        "card_count": total_count,
        "usd_total": usd_total,
    }


# Sentinel: default ``types`` for _build_precon_rows means "the modern-precon
# allow-set". We can't reference mtgjson.PRECON_MODERN_TYPES as a default arg
# because mtgjson is imported lazily (function-local) to avoid an import cycle.
_PRECON_TYPES_DEFAULT = object()


def unsynced_set_codes(codes: Iterable[str]) -> list[str]:
    """Return the subset of ``codes`` (lowercased) with zero rows in ``cards``.

    Used to decide which sets still need a Scryfall pull before a price
    rollup can see them.
    """
    wanted = sorted({c.lower() for c in codes if c})
    if not wanted:
        return []
    with db.connect() as conn:
        placeholders = ",".join("?" for _ in wanted)
        present = {
            row[0]
            for row in conn.execute(
                f"SELECT DISTINCT set_code FROM cards WHERE set_code IN ({placeholders})",
                wanted,
            ).fetchall()
        }
    return [c for c in wanted if c not in present]


def _build_precon_rows(
    set_code: str | None = None,
    *,
    only_type: str | None = None,
    types=_PRECON_TYPES_DEFAULT,
    include_collector: bool = False,
    sync_all: bool = False,
    prepopulate_from_counts: bool = False,
    progress=None,
) -> list[dict]:
    """Enumerate physical precon products and roll each up.

    ``set_code=None`` (the default) spans every set — the precon catalog is
    global. ``types`` defaults to ``mtgjson.PRECON_MODERN_TYPES``; pass
    ``types=None`` for every physical type. Rolling up
    ``card_count``/``commander`` fetches each precon's per-deck JSON from
    MTGJSON (cached forever), so a first all-sets run makes one request per
    precon.

    ``usd_total`` is best-effort: it prices only cards already present in the
    local ``cards`` table, so it's blank for sets not yet synced. With
    ``sync_all=True``, every set referenced by the catalog's cards is synced
    from Scryfall first (batched), so all totals populate — slower, and it
    grows the local cards table with sets you may not own. ``progress`` is an
    optional ``callable(str)`` for status lines.

    Each row carries ``constructed_qty``/``deconstructed_qty`` fill values.
    With ``prepopulate_from_counts=True`` (the ``modify`` flavor) they're filled
    from the REAL deck collection via ``decks.precon_unit_counts()`` (0 when the
    precon isn't owned); otherwise (the ``add`` flavor) they're ``None`` (blank
    cells). Sorted newest-first, then by ``(type, deck_name)``.
    """
    from . import mtgjson as mtgjson_mod
    if types is _PRECON_TYPES_DEFAULT:
        types = mtgjson_mod.PRECON_MODERN_TYPES
    variants = mtgjson_mod.precon_variants(
        set_code, only_type=only_type, types=types,
        include_collector=include_collector,
    )
    if not variants:
        return []

    if sync_all:
        # Pre-pass: fetch each deck once (cached after) to collect every set its
        # cards belong to — including cross-set reprints (a FIC deck's FIN
        # cards) — then sync only the sets not already local, in one batched
        # call. The rollup below re-reads the same cached deck JSON for free.
        referenced: set[str] = set()
        for v in variants:
            deck_data = mtgjson_mod.deck(v["fileName"])
            for board_key in ("commander", "mainBoard", "sideBoard"):
                for entry in deck_data.get(board_key) or []:
                    sc = entry.get("setCode")
                    if sc:
                        referenced.add(sc.lower())
        missing = unsynced_set_codes(referenced)
        if progress:
            progress(f"{len(referenced)} sets referenced; {len(missing)} not yet local — syncing…")
        if missing:
            n = sync(missing)
            if progress:
                progress(f"synced {len(missing)} sets → {n} cards upserted")

    from . import mtgjson as _mtg
    summaries = [_precon_variant_summary(v) for v in variants]

    # Fill columns: prefilled from the real deck collection for the `modify`
    # flavor, blank (None) for `add`. Derive all counts in one query, join by
    # fileName. Each row also carries a `suggested_state` (built|pool) so the
    # writer can steer pool products (Starter Collection, Scene Box) to the
    # pool column.
    counts = {}
    if prepopulate_from_counts:
        from . import decks as decks_mod
        counts = decks_mod.precon_unit_counts()
    for s in summaries:
        s["suggested_state"] = _mtg.default_precon_state(
            s["file_name"], name=s.get("deck_name"))
        if prepopulate_from_counts:
            c, d, p = counts.get(s["file_name"], (0, 0, 0))
            s["constructed_qty"] = c
            s["deconstructed_qty"] = d
            s["pool_qty"] = p
        else:
            s["constructed_qty"] = None
            s["deconstructed_qty"] = None
            s["pool_qty"] = None

    # Newest first (release_date is an ISO date string, so reverse-lex works),
    # then cluster like product types, then name.
    summaries.sort(key=lambda r: (r["release_date"], r["type"], r["deck_name"]),
                   reverse=True)
    return summaries


def write_jumpstart_list_xlsx(set_code: str, out_path: Path,
                              *, slug: str | None = None) -> int:
    """Emit a fillable XLSX of every Jumpstart pack variant for ``set_code``.

    Row schema: file_name | theme | card_count | usd_total | keep_qty | deconstructed_qty

    ``keep_qty`` (0 or 1) = copies kept *constructed*: one ``pack:*`` recipe is
    created and the sum of keep_qty + deconstructed_qty copies' worth of cards
    land in inventory, then one physical copy is auto-composed into
    deck_assignments. ``deconstructed_qty`` = copies torn into free/loose cards
    (no pledge). Hidden ``_meta`` sheet declares ``kind=jumpstart`` so ingest
    can dispatch the correct branch. Returns ``rows_written``.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    from . import __version__

    rows = _build_jumpstart_rows(set_code)
    if not rows:
        raise ValueError(
            f"no Jumpstart variants found for set {set_code!r}. "
            f"Check `mm mtgjson decks --set {set_code}` for available decks."
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "checklist"

    headers = ["file_name", "theme", "card_count", "usd_total",
               "keep_qty", "deconstructed_qty"]
    ws.append(headers)
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

    qty_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # keep_qty is a 0/1 flag: 0 = deconstruct all copies (no recipe),
    # 1 = keep one constructed (creates recipe + auto-composes one copy).
    keep_validator = DataValidation(type="whole", operator="between",
                                    formula1=0, formula2=1, allow_blank=True)
    keep_validator.error = "Enter 0 (deconstruct all copies, no recipe) or 1 (keep one constructed)."
    keep_validator.errorTitle = "Invalid keep_qty"
    ws.add_data_validation(keep_validator)

    # deconstructed_qty is a non-negative integer (blank = 0).
    decon_validator = DataValidation(type="whole", operator="greaterThanOrEqual",
                                     formula1=0, allow_blank=True)
    decon_validator.error = "Enter a non-negative integer (or leave blank for 0)."
    decon_validator.errorTitle = "Invalid deconstructed_qty"
    ws.add_data_validation(decon_validator)

    for r in rows:
        ws.append([
            r["file_name"],
            r["theme"],
            r["card_count"],
            r["usd_total"],
            None,
            None,
        ])
    last_row = ws.max_row

    # col 5 = keep_qty (0/1), col 6 = deconstructed_qty (non-negative int)
    keep_letter = get_column_letter(5)
    keep_validator.add(f"{keep_letter}2:{keep_letter}{last_row}")
    decon_letter = get_column_letter(6)
    decon_validator.add(f"{decon_letter}2:{decon_letter}{last_row}")
    for col_idx in (5, 6):
        for r in range(2, last_row + 1):
            ws.cell(row=r, column=col_idx).fill = qty_fill

    for row_idx in range(2, last_row + 1):
        ws.cell(row=row_idx, column=4).number_format = '"$"#,##0.00'

    widths = {1: 22, 2: 24, 3: 11, 4: 11, 5: 10, 6: 17}
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = "A2"

    meta_ws = wb.create_sheet("_meta")
    meta_ws.sheet_state = "hidden"
    meta_ws.append(["key", "value"])
    meta_ws["A1"].font = Font(bold=True)
    meta_ws["B1"].font = Font(bold=True)

    code = set_code.lower()
    meta = {
        # `kind` is the dispatch key for `mm set ingest`. New value 'jumpstart'
        # routes to the Jumpstart importer; existing values 'inventory' and
        # 'missing' route to their own paths.
        "kind": "jumpstart",
        "anchor_code": code,
        "set_codes": code,
        "slug": slug or out_path.stem,
        "mode": "add",
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "magic_manager_version": __version__,
    }
    for k, v in meta.items():
        meta_ws.append([k, v])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    for _ws in wb.worksheets:
        util.apply_base_font_size(_ws)
    wb.save(out_path)
    return last_row - 1


def write_jumpstart_list_md(set_code: str, out_path: Path,
                            *, slug: str | None = None) -> int:
    """Markdown twin of ``write_jumpstart_list_xlsx``.

    Line shape (after YAML frontmatter):

        - Toph_TLE — Toph — 15 cards — $4.20 [K:0 D:0]

    Parser keys on the leading file_name token, so prose changes don't break
    ingest. The ``[K:k D:d]`` bracket holds keep_qty (0 or 1, copies kept
    constructed) and deconstructed_qty (copies torn into free cards).
    """
    from . import __version__

    rows = _build_jumpstart_rows(set_code)
    if not rows:
        raise ValueError(
            f"no Jumpstart variants found for set {set_code!r}. "
            f"Check `mm mtgjson decks --set {set_code}` for available decks."
        )

    code = set_code.lower()
    meta = {
        "kind": "jumpstart",
        "anchor_code": code,
        "set_codes": code,
        "slug": slug or out_path.stem,
        "mode": "add",
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "magic_manager_version": __version__,
    }

    out_lines: list[str] = ["---"]
    for k, v in meta.items():
        out_lines.append(f"{k}: {v}")
    out_lines.append("---")
    out_lines.append("")
    out_lines.append(f"# {code.upper()} Jumpstart variants ({len(rows)} packs)")
    out_lines.append("")
    out_lines.append(
        "Edit the `[K:k D:d]` bracket per row: `K` = copies kept *constructed* "
        "(0 or 1 — creates a `pack:*` recipe and auto-composes one physical copy), "
        "`D` = copies deconstructed to free cards (no pledge). K+D = total packs "
        "opened. Save, then run `mm set ingest` to apply."
    )
    out_lines.append("")
    for r in rows:
        usd = r["usd_total"]
        usd_seg = f"${usd:.2f}" if usd is not None else "—"
        out_lines.append(
            f"- {r['file_name']} — {r['theme']} — {r['card_count']} cards — "
            f"{usd_seg} [K:0 D:0]"
        )
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text("\n".join(out_lines) + "\n", encoding="utf-8")
    return len(rows)


def _precon_list_meta(slug: str, out_stem: str, *, mode: str,
                      only_type: str | None,
                      all_physical: bool, include_collector: bool) -> dict:
    """Shared ``_meta`` for the (global, all-sets) precon catalog.

    No ``anchor_code``/``set_codes`` — the catalog spans every set, so ingest
    derives each row's set from its ``Words_CODE`` fileName and syncs on demand.
    ``mode`` is ``add`` (blank fill columns; ingest adds the entered counts) or
    ``modify`` (columns prefilled from the live deck counts; ingest applies the
    signed delta vs the prefilled value).
    """
    from . import __version__
    return {
        # `kind` is the dispatch key for `mm set ingest`. 'precon' routes to
        # the shared deck-checklist engine (as 'jumpstart' does); 'inventory'
        # and 'missing' route to their own paths.
        "kind": "precon",
        "mode": mode,
        "slug": slug or out_stem,
        # Provenance: which slice of Magic's precons this catalog covers.
        "only_type": only_type or "",
        "all_physical": "1" if all_physical else "0",
        "include_collector": "1" if include_collector else "0",
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "magic_manager_version": __version__,
    }


def _add_precon_banner_sheet(wb, mode: str) -> None:
    """Visible ``README`` sheet for the precon catalog, stating its ingest
    semantics — the precon twin of ``_add_mode_banner_sheet``."""
    from openpyxl.styles import Alignment, Font, PatternFill

    banner = wb.create_sheet("README", 0)
    if mode == "modify":
        color = "C00000"
        lines = [
            "⚠  MODIFY precon catalog — read before editing",
            "",
            "constructed_qty / deconstructed_qty / pool_qty are prefilled with",
            "your current precon units (counted live from your collection).",
            "",
            "Ingest applies the DIFFERENCE as a new transaction:",
            "  constructed 1 → 2  builds another copy (adds its cards + a deck).",
            "  constructed 2 → 1  is NOT applied here — removing a copy is an",
            "                     explicit action: run `mm deck delete <slug>`.",
            "                     The count updates automatically once you do.",
            "Untouched rows are left alone.",
            "",
            "pool_qty = card-POOL products (Starter Collection, Scene Box) —",
            "  never a playable deck. Their pool cell is tinted green; fill it,",
            "  not constructed. Cards go loose in inventory; the deck row is",
            "  just a 'you own one' marker.",
        ]
    else:
        color = "548235"
        lines = [
            "ADD precon catalog — record copies you acquired",
            "",
            "constructed_qty = built copies you're adding (each creates a deck",
            "  and adds its cards to inventory).",
            "deconstructed_qty = copies you tore down for parts (a deck row is",
            "  recorded, cards go loose).",
            "pool_qty = card-POOL products (Starter Collection, Scene Box) that",
            "  were never a deck — cards go loose, a marker deck row is kept.",
            "  Their pool cell is tinted green; fill it, not constructed.",
            "Ingest ADDS these; it never removes.",
        ]
    title_font = Font(bold=True, size=13, color="FFFFFF")
    title_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for i, text in enumerate(lines, start=1):
        cell = banner.cell(row=i, column=1, value=text)
        if i == 1:
            cell.font = title_font
            cell.fill = title_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    banner.column_dimensions["A"].width = 72
    banner.sheet_properties.tabColor = color
    if "checklist" in wb.sheetnames:
        wb.active = wb.sheetnames.index("checklist")


def write_precon_list_xlsx(out_path: Path, *,
                           slug: str | None = None,
                           mode: str = "add",
                           only_type: str | None = None,
                           all_physical: bool = False,
                           include_collector: bool = False,
                           sync_all: bool = False,
                           progress=None) -> int:
    """Emit a fillable XLSX cataloging preconstructed products across ALL sets.

    The precon catalog is global — there are only a handful of precons per set,
    so this is one master list you populate once, not a per-set file. Scope
    defaults to the modern-constructed product types (``PRECON_MODERN_TYPES``);
    ``only_type`` narrows to one type, ``all_physical=True`` opens it to every
    physical product. ``… Collector's Edition`` twins are excluded unless
    ``include_collector=True``.

    Row schema: file_name | set | deck_name | type | release_date | commander |
    card_count | usd_total | constructed_qty | deconstructed_qty

    ``constructed_qty`` and ``deconstructed_qty`` track precon decks AS UNITS:
    how many built vs torn-down copies of each product you have. In ``mode=add``
    (default) both cells are blank and ingest ADDS the entered counts. In
    ``mode=modify`` both are prefilled from the live deck counts (your current
    collection) and ingest applies the SIGNED DELTA vs the prefilled value
    — a new transaction, not a history rewrite. ``usd_total`` is best-effort —
    blank for sets not yet in the local cards table; pass ``sync_all=True`` to
    sync every referenced set first (slower; ``progress`` is a status callback).
    Hidden ``_meta`` declares ``kind=precon`` + ``mode``. Returns
    ``rows_written``.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    rows = _build_precon_rows(
        only_type=only_type,
        types=(None if all_physical else _PRECON_TYPES_DEFAULT),
        include_collector=include_collector,
        sync_all=sync_all,
        prepopulate_from_counts=(mode == "modify"),
        progress=progress,
    )
    if not rows:
        raise ValueError(
            "no precon variants found"
            + (f" of type {only_type!r}" if only_type else "")
            + ". Check `mm mtgjson decks` for available decks."
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "checklist"

    headers = ["file_name", "set", "deck_name", "type", "release_date",
               "commander", "card_count", "usd_total",
               "constructed_qty", "deconstructed_qty", "pool_qty"]
    ws.append(headers)
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

    qty_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    # Pool-suggested rows (Starter Collection, Scene Box) get their pool cell
    # tinted a distinct green so the user fills THAT column, not constructed.
    pool_hint_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

    # All three fill columns are non-negative counts (built / torn-down / pool
    # copies of the precon as a unit), not a 0/1 flag.
    count_validator = DataValidation(type="whole", operator="greaterThanOrEqual",
                                     formula1=0, allow_blank=True)
    count_validator.error = "Enter a non-negative integer (or leave blank for 0)."
    count_validator.errorTitle = "Invalid count"
    ws.add_data_validation(count_validator)

    for r in rows:
        ws.append([
            r["file_name"],
            r["set"].upper(),
            r["deck_name"],
            r["type"],
            r["release_date"],
            r["commander"],
            r["card_count"],
            r["usd_total"],
            r.get("constructed_qty"),
            r.get("deconstructed_qty"),
            r.get("pool_qty"),
        ])
    last_row = ws.max_row

    # cols 9/10/11 = constructed_qty / deconstructed_qty / pool_qty (non-neg int)
    for col_idx in (9, 10, 11):
        letter = get_column_letter(col_idx)
        count_validator.add(f"{letter}2:{letter}{last_row}")
        for r in range(2, last_row + 1):
            ws.cell(row=r, column=col_idx).fill = qty_fill
    # Tint the pool cell of pool-suggested rows (col 11) so it stands out.
    for i, r in enumerate(rows, start=2):
        if r.get("suggested_state") == "pool":
            ws.cell(row=i, column=11).fill = pool_hint_fill

    for row_idx in range(2, last_row + 1):
        ws.cell(row=row_idx, column=8).number_format = '"$"#,##0.00'

    widths = {1: 34, 2: 6, 3: 30, 4: 18, 5: 13, 6: 26, 7: 11, 8: 11, 9: 15, 10: 17, 11: 9}
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = "A2"

    meta_ws = wb.create_sheet("_meta")
    meta_ws.sheet_state = "hidden"
    meta_ws.append(["key", "value"])
    meta_ws["A1"].font = Font(bold=True)
    meta_ws["B1"].font = Font(bold=True)

    meta = _precon_list_meta(slug, out_path.stem, mode=mode, only_type=only_type,
                             all_physical=all_physical,
                             include_collector=include_collector)
    for k, v in meta.items():
        meta_ws.append([k, v])

    _add_precon_banner_sheet(wb, mode)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    for _ws in wb.worksheets:
        util.apply_base_font_size(_ws)
    wb.save(out_path)
    return last_row - 1


def write_precon_list_md(out_path: Path, *,
                         slug: str | None = None,
                         mode: str = "add",
                         only_type: str | None = None,
                         all_physical: bool = False,
                         include_collector: bool = False,
                         sync_all: bool = False,
                         progress=None) -> int:
    """Markdown twin of ``write_precon_list_xlsx`` (global, all-sets catalog).

    Line shape (after YAML frontmatter):

        - CounterBlitzFinalFantasyX_FIC — FIC — Counter Blitz — Commander Deck — 100 cards — $142.50 [C:0 D:0]

    The ``[C:c D:d]`` bracket holds constructed_qty and deconstructed_qty (both
    non-negative counts). In ``mode=modify`` the bracket is prefilled from the
    live deck counts. ``sync_all``/``progress`` behave as on
    ``write_precon_list_xlsx``.
    """
    rows = _build_precon_rows(
        only_type=only_type,
        types=(None if all_physical else _PRECON_TYPES_DEFAULT),
        include_collector=include_collector,
        sync_all=sync_all,
        prepopulate_from_counts=(mode == "modify"),
        progress=progress,
    )
    if not rows:
        raise ValueError(
            "no precon variants found"
            + (f" of type {only_type!r}" if only_type else "")
            + ". Check `mm mtgjson decks` for available decks."
        )

    meta = _precon_list_meta(slug, out_path.stem, mode=mode, only_type=only_type,
                             all_physical=all_physical,
                             include_collector=include_collector)

    out_lines: list[str] = ["---"]
    for k, v in meta.items():
        out_lines.append(f"{k}: {v}")
    out_lines.append("---")
    out_lines.append("")
    scope = only_type if only_type else ("all physical products" if all_physical
                                         else "modern constructed precons")
    out_lines.append(f"# Precon catalog — {scope} ({len(rows)} decks, all sets)")
    out_lines.append("")
    if mode == "modify":
        out_lines.append(
            "> ⚠ **MODIFY precon catalog.** The `[C:c D:d]` brackets are prefilled "
            "with your current precon decks (counted live from your collection). "
            "Ingest applies the DIFFERENCE: raising `C` builds another copy (adds its "
            "cards + a deck); lowering `C` is NOT applied here — removing a copy is an "
            "explicit action (`mm deck delete <slug>`), and the count updates once you "
            "do. Untouched rows are left alone."
        )
    else:
        out_lines.append(
            "> **ADD precon catalog.** Edit the `[C:c D:d]` bracket per row: `C` = "
            "built copies you acquired (each creates a deck + adds its cards), `D` = "
            "copies you tore down for parts (a deck row is recorded, cards go loose). "
            "Ingest ADDS these; it never removes anything."
        )
    out_lines.append("")
    out_lines.append("Save, then run `mm set ingest` to apply.")
    out_lines.append("")
    for r in rows:
        usd = r["usd_total"]
        usd_seg = f"${usd:.2f}" if usd is not None else "—"
        c = r.get("constructed_qty") or 0
        d = r.get("deconstructed_qty") or 0
        p = r.get("pool_qty") or 0
        pool_hint = "  ← pool (fill P)" if r.get("suggested_state") == "pool" else ""
        out_lines.append(
            f"- {r['file_name']} — {r['set'].upper()} — {r['deck_name']} — "
            f"{r['type']} — {r['card_count']} cards — {usd_seg} [C:{c} D:{d} P:{p}]{pool_hint}"
        )
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text("\n".join(out_lines) + "\n", encoding="utf-8")
    return len(rows)


def _slug_theme(theme: str, set_code: str) -> str:
    """Build a deck slug for a Jumpstart pack: ``pack:<theme>-<setcode>``."""
    raw_theme = "".join(c if c.isalnum() else "-" for c in theme.lower())
    while "--" in raw_theme:
        raw_theme = raw_theme.replace("--", "-")
    raw_theme = raw_theme.strip("-")
    return f"pack:{raw_theme}-{set_code.lower()}"


# ---------------------------------------------------------------------------
# Deck-checklist ingest engine.
#
# Taxonomy: a *precon* (preconstructed product) is the base concept; a
# Jumpstart pack is one species of precon. So the ingest is a single engine
# both kinds share — NOT a Jumpstart path with a precon clone bolted on. The
# only per-kind differences are (a) how a deck's slug is derived and (b) which
# ``format`` the created deck gets. ``ingest_jumpstart_from_path`` below is a
# thin specialization wrapper preserved for its callers.
# ---------------------------------------------------------------------------

# What each ``kind`` contributes to the shared engine:
#   slug_fn(deck_name)  -> the deck slug to create (or None to let
#                          ``import_precon`` derive it from the deck name), and
#   deck_format         -> the ``format`` passed to ``import_precon`` (None lets
#                          it derive ``commander`` for Commander decks, else null).
def _deck_checklist_kind_config(kind: str, set_code: str):
    if kind == "jumpstart":
        # pack:<theme>-<code> slugs; every pack is a 'jumpstart'-format deck.
        return (lambda name: _slug_theme(name, set_code)), "jumpstart"
    if kind == "precon":
        # Slug derived from the deck name (as `mm deck import-precon` does);
        # format left to import_precon's type-based default.
        return (lambda name: None), None
    raise ValueError(f"unknown deck-checklist kind: {kind!r}")


def _apply_deck_checklist(parsed, *, set_code: str, slug_fn, deck_format) -> dict:
    """Apply parsed keep/deconstruct rows to the DB (the shared engine).

    For each row with T = keep_qty + deconstructed_qty > 0:
      - ``keep_qty == 1``: creates one deck recipe, adds T copies' worth of
        cards to inventory via ``import_precon(deconstruct=False, copies=T)``,
        then auto-constructs one physical copy via
        ``deck_assign_from_composition``.
      - ``keep_qty == 0`` (D > 0): fully deconstructed — runs
        ``import_precon(deconstruct=True, copies=D)`` so all D copies' cards
        land in inventory with no recipe created.

    Raises ``ValueError`` (naming the offending file_name(s)) if any acted row
    has ``keep_qty > 1`` — validation runs before any DB writes.

    Returns a kind-neutral summary:
      - ``rows_total`` (int): rows present in the file
      - ``rows_acted`` (int): rows with T > 0
      - ``constructed`` (int): rows where keep_qty == 1 (recipe + composed)
      - ``loose_copies`` (int): total unpledged copies deconstructed to inventory
      - ``inv_qty_total`` (int): cumulative card-qty added to inventory
      - ``per_row``: list of ``{"file_name", "label", "keep_qty",
        "deconstructed_qty", "composed": bool, "slug": str|None,
        "slugs": list, "missing_sids": [...], "error": str|None}``
      - ``warnings`` (list[str]): non-fatal parse/lookup warnings
    """
    from . import decks as decks_mod, mtgjson as mtgjson_mod

    # Pre-scan validation: keep_qty must be 0 or 1 for every acted row.
    # Do this before any DB writes so we never partially commit.
    invalid = [
        r.file_name
        for r in parsed.rows
        if r.keep_qty + r.deconstructed_qty > 0 and r.keep_qty > 1
    ]
    if invalid:
        raise ValueError(
            f"keep_qty must be 0 or 1; offending deck(s): {', '.join(invalid)}"
        )

    summary: dict = {
        "rows_total": len(parsed.rows),
        "rows_acted": 0,
        "constructed": 0,
        "loose_copies": 0,
        "inv_qty_total": 0,
        "per_row": [],
        "warnings": list(parsed.warnings),
    }

    for row in parsed.rows:
        total = row.keep_qty + row.deconstructed_qty
        if total <= 0:
            continue
        summary["rows_acted"] += 1
        per_row: dict = {
            "file_name": row.file_name,
            "label": row.theme,
            "keep_qty": row.keep_qty,
            "deconstructed_qty": row.deconstructed_qty,
            "composed": False,
            "slug": None,
            "slugs": [],
            "missing_sids": [],
            "error": None,
        }

        # Resolve the display name (used both as a label and to derive the
        # slug) from MTGJSON when the parser couldn't capture it (md path, or
        # the precon XLSX 'deck_name' column the shared parser doesn't read).
        deck_name = row.theme
        if not deck_name:
            try:
                deck_data = mtgjson_mod.deck(row.file_name)
                deck_name = deck_data.get("name") or row.file_name
            except mtgjson_mod.MtgJsonError as e:
                per_row["error"] = f"could not fetch deck JSON: {e}"
                summary["per_row"].append(per_row)
                continue
            per_row["label"] = deck_name

        base_slug = slug_fn(deck_name)  # may be None → import_precon derives it

        try:
            if row.keep_qty == 1:
                # Create the recipe + add T copies' worth of cards to inventory,
                # then auto-construct one physical copy.
                r = decks_mod.import_precon(
                    row.file_name,
                    slug=base_slug,
                    format=deck_format,
                    copies=total,
                    add_inventory=True,
                    deconstruct=False,
                )
                per_row["slugs"].extend(r["effective_slugs"])
                per_row["missing_sids"].extend(r["missing_sids"])
                # import_precon derives the slug when slug_fn returned None, so
                # read the effective slug back rather than trusting base_slug.
                effective_slug = r["effective_slugs"][0] if r["effective_slugs"] else base_slug
                per_row["slug"] = effective_slug
                summary["inv_qty_total"] += r["inv_qty_total"]
                # Auto-construct one physical copy (pledges exactly one recipe's
                # worth into deck_assignments, leaving deconstructed_qty copies free).
                if effective_slug:
                    decks_mod.deck_assign_from_composition(effective_slug)
                    per_row["composed"] = True
                summary["constructed"] += 1
                summary["loose_copies"] += row.deconstructed_qty
            else:
                # Fully deconstructed: D copies loose, no recipe.
                r = decks_mod.import_precon(
                    row.file_name,
                    slug=base_slug,  # not used in deconstruct path
                    format=deck_format,
                    copies=row.deconstructed_qty,
                    add_inventory=True,
                    deconstruct=True,
                )
                per_row["missing_sids"].extend(r["missing_sids"])
                summary["inv_qty_total"] += r["inv_qty_total"]
                summary["loose_copies"] += row.deconstructed_qty
        except (mtgjson_mod.MtgJsonError, ValueError) as e:
            per_row["error"] = str(e)

        summary["per_row"].append(per_row)

    return summary


def _count_precon_deck_copies(base_slug: str) -> int:
    """How many decks already exist for ``base_slug`` (the bare slug plus any
    ``-2``/``-3``… copies). Used to pick the next free copy slug."""
    from . import decks as decks_mod
    n = 0
    if decks_mod.deck_get(base_slug) is not None:
        n += 1
    i = 2
    while decks_mod.deck_get(f"{base_slug}-{i}") is not None:
        n += 1
        i += 1
    return n


def _apply_precon_checklist(parsed, *, mode: str) -> dict:
    """Apply a filled-in PRECON checklist as a signed transaction against the
    ``decks`` table — the single source of truth (there is no ledger).

    Precon rows track built (``constructed_qty`` → ``row.keep_qty``), torn-down
    (``deconstructed_qty``) and card-pool (``pool_qty``) copies as UNITS. Current
    counts are DERIVED from decks via ``precon_unit_counts_for`` (built,
    deconstructed, pool):

      - ``add`` mode: the entered counts ARE the delta — add that many copies of
        each state on top of what's already owned.
      - ``modify`` mode: the file was prefilled from the real deck counts, so
        the delta is (entered − current) per column.

    Applying a positive delta creates that many deck rows via ``import_precon``
    (distinct ``-2``/``-3`` slugs) in the matching state: ``built`` pledges one
    physical copy; ``deconstructed`` and ``pool`` leave the cards loose. Any
    negative delta (``modify`` lowering a count) is NOT applied — it warns and
    points at ``mm deck delete <slug>``; the derived count updates when the user
    actually deletes. No history rewrite via the checklist.

    Returns a summary with ``rows_total``/``rows_acted``/``built``/
    ``deconstructed``/``pool``/``inv_qty_total``/``per_row``/``warnings``; each
    ``per_row`` carries ``count_before``/``count_after``/``delta`` (all 3-tuples)
    and any ``warning``/``error``.
    """
    from . import decks as decks_mod, mtgjson as mtgjson_mod

    summary: dict = {
        "rows_total": len(parsed.rows),
        "rows_acted": 0,
        "built": 0,            # decks built this ingest
        "deconstructed": 0,    # torn-down copies recorded this ingest
        "pool": 0,             # card-pool units recorded this ingest
        "inv_qty_total": 0,
        "per_row": [],
        "warnings": list(parsed.warnings),
    }

    for row in parsed.rows:
        entered = (row.keep_qty, row.deconstructed_qty, row.pool_qty)  # (c, d, p)
        before = decks_mod.precon_unit_counts_for(row.file_name)       # (c, d, p)

        # Compute target absolute counts + the delta to apply, per state.
        if mode == "modify":
            target = entered
        else:  # add: entered values stack on top of what's already owned
            target = tuple(b + e for b, e in zip(before, entered))
        delta = tuple(t - b for t, b in zip(target, before))

        if all(x == 0 for x in delta):
            continue  # untouched row — leave it alone (no history rewrite)
        summary["rows_acted"] += 1

        per_row: dict = {
            "file_name": row.file_name,
            "label": row.theme or row.file_name,
            "count_before": before,
            "count_after": target,
            "delta": delta,
            "built": 0,
            "torn_down": 0,
            "pooled": 0,
            "slugs": [],
            "missing_sids": [],
            "warning": None,
            "error": None,
        }

        # Resolve deck name for slug derivation + label.
        deck_name = row.theme
        if not deck_name:
            try:
                deck_name = mtgjson_mod.deck(row.file_name).get("name") or row.file_name
            except mtgjson_mod.MtgJsonError as e:
                per_row["error"] = f"could not fetch deck JSON: {e}"
                summary["per_row"].append(per_row)
                continue
            per_row["label"] = deck_name
        base_slug = decks_mod._slug(deck_name)

        # Build N copies in one state (distinct -2/-3 slugs); pledge only built.
        def _build(n: int, state: str) -> None:
            for _ in range(n):
                n_existing = _count_precon_deck_copies(base_slug)
                copy_slug = base_slug if n_existing == 0 else f"{base_slug}-{n_existing + 1}"
                r = decks_mod.import_precon(
                    row.file_name, slug=copy_slug, format=None, copies=1,
                    add_inventory=True,
                    deconstruct=(state != "built"),  # loose cards for decon + pool
                    precon_state=state,
                )
                per_row["slugs"].extend(r["effective_slugs"])
                per_row["missing_sids"].extend(r["missing_sids"])
                summary["inv_qty_total"] += r["inv_qty_total"]
                if state == "built":
                    eff = r["effective_slugs"][0] if r["effective_slugs"] else copy_slug
                    if eff:
                        decks_mod.deck_assign_from_composition(eff)
                    per_row["built"] += 1
                    summary["built"] += 1
                elif state == "deconstructed":
                    per_row["torn_down"] += 1
                    summary["deconstructed"] += 1
                else:  # pool
                    per_row["pooled"] += 1
                    summary["pool"] += 1

        try:
            delta_c, delta_d, delta_p = delta
            if delta_c > 0:
                _build(delta_c, "built")
            if delta_d > 0:
                _build(delta_d, "deconstructed")
            if delta_p > 0:
                _build(delta_p, "pool")

            if any(x < 0 for x in delta):
                per_row["warning"] = (
                    f"lowered counts (Δbuilt={delta_c}, Δdeconstructed={delta_d}, "
                    f"Δpool={delta_p}) were NOT applied — removing a precon copy is "
                    f"an explicit deck action: run `mm deck delete <slug>` (see "
                    f"`mm deck ls`). The count updates automatically once you do."
                )
        except (mtgjson_mod.MtgJsonError, ValueError) as e:
            per_row["error"] = str(e)

        summary["per_row"].append(per_row)

    return summary


def ingest_deck_checklist_from_path(path: Path, *, kind: str) -> dict:
    """Apply a filled-in deck checklist (precon or Jumpstart) to the local DB.

    The general (precon-level) entry point: parse the file, then run the shared
    ``_apply_deck_checklist`` engine with the per-``kind`` slug/format config.
    ``kind`` is one of ``"precon"`` or ``"jumpstart"``.

    A ``set_code`` is only meaningful for Jumpstart (it seeds the ``pack:*``
    slug). The precon catalog is global — every row can be from a different set,
    so no single ``set_code`` applies; the precon slug is derived per-row from
    the deck name instead, and each ``import_precon`` call self-syncs the sets
    its own cards reference. Returns the kind-neutral summary described on
    ``_apply_deck_checklist``.
    """
    from . import parsers
    from pathlib import Path as _Path

    path = _Path(path)
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        parsed = parsers.parse_jumpstart_list_xlsx(path)
    elif suffix == ".md":
        parsed = parsers.parse_jumpstart_list_md(path)
    else:
        raise ValueError(f"unsupported deck-checklist extension: {suffix!r}")

    meta = parsed.meta or {}

    # Precon: a signed ledger transaction (constructed/deconstructed units),
    # honoring the file's add/modify mode. Distinct from the jumpstart 0/1
    # recipe-flag engine.
    if kind == "precon":
        file_mode = (meta.get("mode") or "add").lower()
        return _apply_precon_checklist(parsed, mode=file_mode)

    set_code = (meta.get("set_codes") or meta.get("anchor_code") or "").lower()
    if not set_code:
        # Fall back to inferring from a fileName like ``Toph_TLE``.
        for r in parsed.rows:
            if "_" in r.file_name:
                set_code = r.file_name.rsplit("_", 1)[1].lower()
                break
    # Only Jumpstart needs a set_code (for its slug).
    if not set_code and kind == "jumpstart":
        raise ValueError("could not determine set_code from checklist _meta or rows")

    slug_fn, deck_format = _deck_checklist_kind_config(kind, set_code)
    return _apply_deck_checklist(
        parsed, set_code=set_code, slug_fn=slug_fn, deck_format=deck_format
    )


def ingest_jumpstart_from_path(path: Path) -> dict:
    """Apply a filled-in Jumpstart checklist to the local DB.

    Thin specialization of ``ingest_deck_checklist_from_path`` (Jumpstart is a
    species of precon): ``pack:*`` slugs and ``format='jumpstart'``. Behavior
    is unchanged from before the shared-engine refactor.
    """
    return ingest_deck_checklist_from_path(path, kind="jumpstart")
